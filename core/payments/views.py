import jdatetime
import logging
import mimetypes
import random
import re
import uuid
from difflib import SequenceMatcher
from openpyxl import Workbook
from urllib.parse import urlencode
import json

from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import DatabaseError, IntegrityError, transaction
from django.db.models import Count, Max, Prefetch, Q, Sum
from django.http import FileResponse, Http404, HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth import logout as auth_logout
from django.contrib.auth.views import LoginView
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.conf import settings
from django.core.files.base import ContentFile
from django.contrib.sessions.models import Session
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST
from django.views.decorators.cache import never_cache
from django.utils.decorators import method_decorator
from django_ratelimit.decorators import ratelimit
from axes.helpers import get_client_ip_address
from zoneinfo import ZoneInfo

from .forms import CounterpartyBankAccountFormSet, CounterpartyForm, CounterpartyManagementForm, CustomPasswordChangeForm, CustomerOrderForm, CustomerOrderItemFormSet, CustomerProfileUpdateForm, DailyPaymentAssignmentForm, DailyPaymentPlanForm, InvoiceCustomerNoteForm, InvoiceUploadForm, OrderProformaUploadForm, PaymentRecordForm, PriceListUploadForm, ProformaInvoiceForm, ReconciliationMessageForm, ReconciliationThreadForm, SalesAssignmentBulkForm, StaffOrderUpdateForm, StaffPaymentDetailsForm, StaffStatusUpdateForm, SystemLogoSettingsForm, SystemMenuSettingsForm, UserAccessManagementForm, UserAccountManagementForm
from .invoice_extraction import create_preview_extraction_job, flatten_fields, process_invoice_extraction_job
from .models import AgencyApplication, AgencyApplicationLog, Counterparty, CounterpartyBankAccount, CustomerOrder, CustomerOrderLog, CustomerSalesAssignment, DailyPaymentAssignment, DailyPaymentPlan, InvoiceExtractionJob, InvoiceRecord, LoginAdvertisement, PaymentActivityLog, PaymentRecord, PaymentReceipt, PriceList, ProductCatalog, ProfileChangeRequest, ProformaInvoice, ProformaInvoiceLog, ReconciliationMessage, ReconciliationReadState, ReconciliationThread, SystemActivityLog, SystemSettings, UploadSettings, UserNotification, UserProfile, WarrantyClaim, WarrantyClaimFile, WarrantyClaimLog
import os


STAFF_ROLES = {'staff', 'finance', 'finance_manager', 'commercial', 'commercial_manager', 'sales', 'sales_manager', 'data_entry'}
MANAGER_ROLES = {'finance_manager', 'commercial_manager', 'sales_manager'}
logger = logging.getLogger(__name__)
DISPLAY_TIME_ZONE = ZoneInfo(getattr(settings, 'APP_DISPLAY_TIME_ZONE', 'Asia/Tehran'))
STATUS_FLAG_META = {
    PaymentRecord.STATUS_COMMERCIAL_REVIEW: ('بررسی بازرگانی', 'flag-blue'),
    PaymentRecord.STATUS_TEMP_COMMERCIAL: ('ثبت موقت بازرگانی', 'flag-teal'),
    PaymentRecord.STATUS_APPROVED: ('ثبت بازرگانی', 'flag-orange'),
    PaymentRecord.STATUS_FINAL_APPROVED: ('تایید نهایی', 'flag-green'),
    PaymentRecord.STATUS_REJECTED: ('رد شده', 'flag-red'),
    PaymentRecord.STATUS_INCOMPLETE: ('ناقص', 'flag-yellow'),
    PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL: ('عودت به بازرگانی', 'flag-gray'),
}
STATUS_PROGRESS_FLOWS = {
    PaymentRecord.STATUS_COMMERCIAL_REVIEW: [PaymentRecord.STATUS_COMMERCIAL_REVIEW],
    PaymentRecord.STATUS_TEMP_COMMERCIAL: [PaymentRecord.STATUS_COMMERCIAL_REVIEW, PaymentRecord.STATUS_TEMP_COMMERCIAL],
    PaymentRecord.STATUS_APPROVED: [PaymentRecord.STATUS_COMMERCIAL_REVIEW, PaymentRecord.STATUS_APPROVED],
    PaymentRecord.STATUS_FINAL_APPROVED: [
        PaymentRecord.STATUS_COMMERCIAL_REVIEW,
        PaymentRecord.STATUS_APPROVED,
        PaymentRecord.STATUS_FINAL_APPROVED,
    ],
    PaymentRecord.STATUS_REJECTED: [PaymentRecord.STATUS_REJECTED],
    PaymentRecord.STATUS_INCOMPLETE: [PaymentRecord.STATUS_INCOMPLETE],
    PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL: [PaymentRecord.STATUS_COMMERCIAL_REVIEW, PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL],
}
CUSTOMER_STATUSES = [
    (PaymentRecord.STATUS_PENDING, 'در حال بررسی'),
    (PaymentRecord.STATUS_FINAL_APPROVED, 'تایید نهایی'),
    (PaymentRecord.STATUS_REJECTED, 'رد شده'),
    (PaymentRecord.STATUS_INCOMPLETE, 'ناقص'),
]


def _user_role(user):
    if not user.is_authenticated:
        return ''
    if user.is_superuser:
        return 'staff'
    try:
        return user.profile.role
    except UserProfile.DoesNotExist:
        return 'staff' if user.is_staff else 'customer'


def _staff_role_label(role):
    return {
        'commercial': 'بازرگانی',
        'commercial_manager': 'مدیر بازرگانی',
        'finance': 'مالی',
        'finance_manager': 'مدیر مالی',
        'sales': 'فروش',
        'sales_manager': 'مدیر فروش',
        'data_entry': 'تکمیل اطلاعات فیش',
        'staff': 'کارمندی',
    }.get(role, '')


def _department_role(role):
    return {
        'commercial_manager': 'commercial',
        'finance_manager': 'finance',
        'sales_manager': 'sales',
    }.get(role, role)


def _is_staff_user(user):
    if not user.is_authenticated:
        return False
    if user.is_staff or user.is_superuser:
        return True
    try:
        return user.profile.role in STAFF_ROLES
    except UserProfile.DoesNotExist:
        return False


def _can_upload_invoices(user):
    if user.is_superuser:
        return True
    try:
        role = user.profile.role
        if role in {'commercial_manager', 'finance_manager'}:
            return True
        return user.profile.can_upload_invoices
    except UserProfile.DoesNotExist:
        return False


def _can_view_invoices(user):
    if user.is_superuser:
        return True
    try:
        role = user.profile.role
        # مشتریان همگی دسترسی یکسان به مشاهده فاکتورهای خودشان دارند.
        if role == 'customer':
            return True
        if role in {'sales', 'sales_manager', 'commercial_manager', 'finance_manager'}:
            return True
        return user.profile.can_view_invoices
    except UserProfile.DoesNotExist:
        return False


def _can_upload_price_lists(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return _user_role(user) in {'commercial', 'commercial_manager', 'sales', 'sales_manager', 'finance', 'finance_manager'}


def _can_issue_proformas(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return _user_role(user) in {'commercial', 'commercial_manager', 'sales', 'sales_manager'}


def _can_view_orders(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return _user_role(user) in {'customer', 'sales', 'sales_manager', 'commercial', 'commercial_manager', 'finance', 'finance_manager'}


def _can_manage_orders(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return _user_role(user) in {'sales', 'sales_manager', 'commercial', 'commercial_manager'}


def _can_manage_sales_assignments(user):
    if not user or not user.is_authenticated:
        return False
    return user.is_superuser or _user_role(user) == 'sales_manager'


def _can_access_reconciliation(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    # کلیه کارکنان و مشتریان به بخش مغایرت‌گیری دسترسی دارند
    return _is_staff_user(user) or _user_role(user) == 'customer'


def _reconciliation_threads_for_user(user):
    qs = (
        ReconciliationThread.objects
        .select_related('customer', 'customer__profile', 'created_by')
        .prefetch_related('staff_participants', 'messages')
        .order_by('-updated_at', '-id')
    )
    if user.is_superuser:
        return qs
    if _user_role(user) == 'customer':
        return qs.filter(customer=user)
    return qs.filter(staff_participants=user).distinct()


def _can_access_reconciliation_thread(user, thread):
    if user.is_superuser:
        return True
    if _user_role(user) == 'customer':
        return thread.customer_id == user.id
    return thread.staff_participants.filter(id=user.id).exists()


def _reconciliation_document_url(document_type, document_id, thread_id=None):
    if not document_type or not document_id:
        return ''
    routes = {
        ReconciliationThread.DOC_PAYMENT: ('payment_timeline', [document_id]),
        ReconciliationThread.DOC_ORDER: ('order_detail', [document_id]),
        ReconciliationThread.DOC_PROFORMA: ('proforma_detail', [document_id]),
        ReconciliationThread.DOC_INVOICE: ('invoice_detail', [document_id]),
        ReconciliationThread.DOC_DAILY_PAYMENT: ('daily_payment_plan_detail', [document_id]),
    }
    route = routes.get(document_type)
    if not route:
        return ''
    try:
        url = reverse(route[0], args=route[1])
    except Exception:
        return ''
    if thread_id:
        # با ارسال آدرس گفتگو به‌عنوان next، دکمه «بازگشت» در صفحه سند
        # کاربر را به همین گفتگو برمی‌گرداند نه صفحه پیش‌فرض.
        next_url = f"{reverse('reconciliation_center')}?thread={thread_id}"
        url = f"{url}?{urlencode({'next': next_url})}"
    return url


def _reconciliation_unread_count(user):
    if not user or not user.is_authenticated or not _can_access_reconciliation(user):
        return 0
    count = 0
    for thread in _reconciliation_threads_for_user(user).prefetch_related('read_states'):
        state = next((item for item in thread.read_states.all() if item.user_id == user.id), None)
        messages_qs = thread.messages.exclude(sender_id=user.id)
        if state:
            messages_qs = messages_qs.filter(created_at__gt=state.last_read_at)
        count += messages_qs.count()
    return count


def _mark_reconciliation_thread_read(thread, user):
    if not user or not user.is_authenticated:
        return
    ReconciliationReadState.objects.update_or_create(
        thread=thread,
        user=user,
        defaults={'last_read_at': timezone.now()},
    )
    # خواندن thread → شمارش پیام‌های نخوانده این کاربر کاهش می‌یابد
    from django.core.cache import cache
    from .context_processors import recon_unread_cache_key
    cache.delete(recon_unread_cache_key(user.id))


def _apply_reconciliation_filters(threads, request):
    filters = {
        'q': (request.GET.get('q') or '').strip(),
        'status': (request.GET.get('status') or '').strip(),
        'document_type': (request.GET.get('document_type') or '').strip(),
    }

    query = filters['q']
    if query:
        # هر کلمه به‌صورت مستقل با فیلدهای مختلف مقایسه می‌شود تا جستجوی نام کامل
        # (که در دیتابیس به‌صورت نام و نام‌خانوادگی جدا ذخیره شده) هم کار کند.
        condition = Q()
        for word in query.split():
            word_condition = (
                Q(title__icontains=word) |
                Q(customer__first_name__icontains=word) |
                Q(customer__last_name__icontains=word) |
                Q(customer__username__icontains=word) |
                Q(customer__profile__organization__icontains=word) |
                Q(messages__body__icontains=word)
            )
            digits = word.replace(',', '')
            if digits.isdigit():
                word_condition |= Q(document_id=int(digits)) | Q(messages__document_id=int(digits))
            condition &= word_condition
        threads = threads.filter(condition).distinct()

    valid_statuses = {choice[0] for choice in ReconciliationThread.STATUS_CHOICES}
    if filters['status'] in valid_statuses:
        threads = threads.filter(status=filters['status'])

    valid_doc_types = {choice[0] for choice in ReconciliationThread.DOCUMENT_CHOICES}
    if filters['document_type'] in valid_doc_types:
        threads = threads.filter(document_type=filters['document_type'])

    return threads, filters


def _can_view_price_list_history(user):
    return _is_staff_user(user)


def _can_delete_customer_documents(user):
    return _is_staff_user(user)


def _active_customer_profiles():
    return UserProfile.objects.filter(
        role='customer',
        user__is_active=True,
        suspended=False,
    ).select_related('user')


def _is_counterparty_user(user):
    """بررسی می‌کند که آیا کاربر یک طرف حساب است (از طریق رابطه یا نقش)."""
    if not user or not user.is_authenticated:
        return False
    if getattr(user, 'counterparty_account', None):
        return True
    try:
        return user.profile.role == 'counterparty'
    except Exception:
        return False


def _get_user_counterparty(user):
    """طرف حساب مرتبط با کاربر را برمی‌گرداند."""
    return getattr(user, 'counterparty_account', None)


def _assigned_customer_ids_for_sales(user):
    if not user or not user.is_authenticated:
        return []
    return list(CustomerSalesAssignment.objects.filter(sales_user=user).values_list('customer_id', flat=True))


def _customer_limited_queryset_for_user(qs, user, customer_field='customer'):
    if _user_role(user) != 'sales':
        return qs
    assigned_ids = _assigned_customer_ids_for_sales(user)
    lookup = {f'{customer_field}_id__in': assigned_ids}
    return qs.filter(**lookup)


def _can_staff_access_customer(user, customer_id):
    if _user_role(user) != 'sales':
        return True
    return customer_id in _assigned_customer_ids_for_sales(user)


def _can_manage_users(user):
    return bool(user and user.is_authenticated and user.is_superuser)


def _can_import_customer_accounting_codes(user):
    if not user or not user.is_authenticated:
        return False
    if not SystemSettings.load().accounting_code_import_enabled:
        return False
    if user.is_superuser:
        return True
    return _user_role(user) in {'finance_manager', 'commercial_manager'}


# نگاشت نقش مدیر هر بخش به نقش کارکنان همان بخش — برای «مدیریت دسترسی‌ها»
ACCESS_DEPARTMENT_MANAGER_ROLES = {
    'commercial_manager': 'commercial',
    'finance_manager':    'finance',
    'sales_manager':      'sales',
    'warranty_manager':   'warranty',
}


def _can_manage_access(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return _user_role(user) in ACCESS_DEPARTMENT_MANAGER_ROLES


def _access_manageable_users(user, query='', role=''):
    qs = (
        User.objects.select_related('profile')
        .filter(is_superuser=False)
        .exclude(profile__role='customer')
        .exclude(counterparty_account__isnull=False)
        .order_by('profile__role', 'first_name', 'last_name', 'username')
    )
    if not user.is_superuser:
        dept_role = ACCESS_DEPARTMENT_MANAGER_ROLES.get(_user_role(user))
        if not dept_role:
            return qs.none()
        qs = qs.filter(profile__role=dept_role)
    if query:
        qs = qs.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(profile__mobile__icontains=query) |
            Q(profile__organization__icontains=query)
        )
    if role:
        qs = qs.filter(profile__role=role)
    return qs


def _can_manage_access_for_target(user, target):
    if not _can_manage_access(user):
        return False
    if user.is_superuser:
        return not target.is_superuser and target.profile.role != 'customer' and not getattr(target, 'counterparty_account', None)
    dept_role = ACCESS_DEPARTMENT_MANAGER_ROLES.get(_user_role(user))
    return bool(dept_role) and target.profile.role == dept_role


def _can_edit_payment_details(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    try:
        return bool(user.profile.can_edit_payment_details)
    except UserProfile.DoesNotExist:
        return False


def _can_access_payment(user, payment):
    if _is_staff_user(user):
        return True
    if payment.user_id == user.id:
        return True
    # طرف حساب فقط به فیش‌هایی که به خودش اختصاص دارند دسترسی دارد
    if _is_counterparty_user(user):
        cp = _get_user_counterparty(user)
        return cp is not None and payment.counterparty_id == cp.id
    return False


def _can_access_invoice(user, invoice):
    if _is_staff_user(user):
        return _can_view_invoices(user) and _can_staff_access_customer(user, invoice.customer_id)
    return invoice.customer_id == user.id


def _safe_next_url(request, default=''):
    next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return next_url
    return default


def _return_link_label(request, default_label):
    # وقتی کاربر از طریق گفتگوی مغایرت‌گیری به این صفحه آمده، دکمه «بازگشت»
    # او را به همان گفتگو برمی‌گرداند نه مقصد پیش‌فرض، پس برچسب باید همین را نشان بدهد.
    next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()
    if next_url:
        try:
            recon_path = reverse('reconciliation_center')
        except Exception:
            recon_path = ''
        if recon_path and next_url.startswith(recon_path):
            return 'بازگشت به گفتگو'
    return default_label


def _file_response(field_file, as_attachment=False, filename=None):
    if not field_file:
        raise Http404
    try:
        field_file.open('rb')
    except FileNotFoundError as exc:
        raise Http404 from exc
    download_name = filename or field_file.name.rsplit('/', 1)[-1]
    content_type, _ = mimetypes.guess_type(download_name)
    return FileResponse(
        field_file,
        as_attachment=as_attachment,
        filename=download_name,
        content_type=content_type or 'application/octet-stream',
    )


@require_POST
def safe_logout(request):
    try:
        auth_logout(request)
        return redirect(settings.LOGOUT_REDIRECT_URL)
    except Exception:
        logger.exception('Logout failed; clearing client cookies as a fallback.')
        response = redirect(settings.LOGOUT_REDIRECT_URL)
        response.delete_cookie(
            settings.SESSION_COOKIE_NAME,
            path=settings.SESSION_COOKIE_PATH,
            domain=settings.SESSION_COOKIE_DOMAIN,
            samesite=settings.SESSION_COOKIE_SAMESITE,
        )
        response.delete_cookie(
            settings.CSRF_COOKIE_NAME,
            path=settings.CSRF_COOKIE_PATH,
            domain=settings.CSRF_COOKIE_DOMAIN,
            samesite=settings.CSRF_COOKIE_SAMESITE,
        )
        return response


class SafeLoginView(LoginView):
    template_name = 'registration/login.html'

    @method_decorator(never_cache)
    @method_decorator(ratelimit(key='ip', rate='10/m', method='POST', block=False))
    def post(self, request, *args, **kwargs):
        if getattr(request, 'limited', False):
            from django.contrib import messages as _msg
            _msg.error(request, 'تعداد درخواست‌های ورود بیش از حد مجاز است. لطفاً چند دقیقه صبر کنید.')
            return self.get(request, *args, **kwargs)
        return super().post(request, *args, **kwargs)

    def get_success_url(self):
        user = self.request.user
        if _is_counterparty_user(user):
            return reverse('counterparty_dashboard')
        return super().get_success_url()

    def form_valid(self, form):
        try:
            response = super().form_valid(form)
        except DatabaseError:
            logger.exception('Login failed because the session/database could not be written.')
            form.add_error(
                None,
                'ورود انجام نشد، چون سامانه در حال حاضر امکان ثبت نشست کاربر در دیتابیس را ندارد. لطفا با مدیر سیستم تماس بگیرید.',
            )
            return self.form_invalid(form)

        # SMS MFA — اگر کاربر SMS MFA فعال داشته باشد و پیامک سیستم روشن باشد
        user = self.request.user
        try:
            profile = user.profile
            if profile.sms_mfa_enabled and _sms_mfa_is_active() and profile.sms_number:
                from .sms_service import send_otp
                otp = send_otp(user, purpose='mfa')
                if otp:
                    self.request.session['sms_mfa_pending'] = True
                    self.request.session['sms_mfa_otp_key'] = otp.pk
                    self.request.session['sms_mfa_next'] = self.get_success_url()
                    return redirect(reverse('sms_otp_verify'))
        except Exception:
            logger.exception('SMS MFA check failed for user %s', user.username)

        return response

    def form_invalid(self, form):
        from axes.helpers import get_lockout_response
        from axes.signals import user_locked_out
        if getattr(self.request, 'axes_locked_out', False):
            cooloff = getattr(settings, 'AXES_COOLOFF_TIME', None)
            minutes = int(cooloff.total_seconds() // 60) if cooloff else 15
            form.add_error(
                None,
                f'حساب کاربری به دلیل تلاش‌های ناموفق مکرر به مدت {minutes} دقیقه قفل شده است.',
            )
        return super().form_invalid(form)


def _suggest_five_digit_password():
    lower = 'abcdefghijklmnopqrstuvwxyz'
    upper = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    digits = '0123456789'
    all_chars = lower + upper + digits

    categories = [lower, upper, digits]
    selected = []
    # ensure at least two categories are represented
    chosen_categories = random.sample(categories, 2)
    for category in chosen_categories:
        selected.append(random.choice(category))
    while len(selected) < 5:
        selected.append(random.choice(all_chars))
    random.shuffle(selected)
    return ''.join(selected)


def _staff_status_choices_for_role(role):
    """
    وضعیت‌های مجاز برای فلگ بازرگانی — بر اساس نقش.
    فلگ مالی مستقل است و از طریق finance_unified_action مدیریت می‌شود.
    """
    # وضعیت‌های مجاز بازرگانی — مستقل از نقش مالی
    COMMERCIAL_CHOICES = [
        (PaymentRecord.STATUS_COMMERCIAL_REVIEW, 'در حال بررسی بازرگانی'),
        (PaymentRecord.STATUS_TEMP_COMMERCIAL,   'ثبت موقت بازرگانی'),
        (PaymentRecord.STATUS_APPROVED,           'ثبت بازرگانی'),
        (PaymentRecord.STATUS_INCOMPLETE,         'ناقص'),
        (PaymentRecord.STATUS_REJECTED,           'رد شده'),
    ]

    dept = _department_role(role)

    if dept == 'commercial':
        return COMMERCIAL_CHOICES

    if dept == 'finance':
        # مالی فقط می‌تواند عودت به بازرگانی کند (از طریق status اصلی)
        return [
            (PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL, 'عودت به بازرگانی'),
        ]

    if dept == 'sales':
        # فروش همان اختیارات بازرگانی برای بررسی اولیه
        return COMMERCIAL_CHOICES

    # سایر نقش‌ها (staff, data_entry, ...) — بدون تغییر وضعیت از این dropdown
    # ادمین سیستم — اختیارات بازرگانی + امکان بازگشت از ناقص به بررسی
    return COMMERCIAL_CHOICES + [
        (PaymentRecord.STATUS_PENDING, 'بازگشت به صف بررسی (رفع نقص ادمین)'),
    ]


def _can_staff_act_on_payment(role, payment, is_system_admin=False):
    """
    آیا کاربر می‌تواند روی فلگ بازرگانی (status اصلی) اقدام کند؟

    قوانین قفل:
    - بازرگانی/فروش: فقط در pending, commercial_review, returned_commercial
    - بازرگانی بعد از ثبت بازرگانی (approved): قفل است
    - مالی: فقط در approved (برای عودت)
    - ادمین: هم دسترسی بازرگانی هم مالی دارد ولی با رعایت منطق
    """
    dept = _department_role(role)

    COMMERCIAL_ACTIVE_STATUSES = {
        PaymentRecord.STATUS_PENDING,
        PaymentRecord.STATUS_COMMERCIAL_REVIEW,
        PaymentRecord.STATUS_TEMP_COMMERCIAL,
        PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL,
    }

    if dept == 'commercial' or dept == 'sales':
        return payment.status in COMMERCIAL_ACTIVE_STATUSES

    if dept == 'finance':
        # مالی فقط عودت به بازرگانی را از طریق این فلگ انجام می‌دهد
        return payment.status == PaymentRecord.STATUS_APPROVED

    if is_system_admin:
        # ادمین در هر وضعیتی می‌تواند اقدام کند (برای مدیریت)
        return payment.status not in {
            PaymentRecord.STATUS_FINAL_APPROVED,
        }

    return False


def _commercial_can_revise(payment, logs=None):
    """
    بازرگانی/فروش می‌تواند وضعیت رد/ناقص/ثبت‌بازرگانی را تجدیدنظر کند اگر:
    - جدیدترین اقدام واقعی روی سند از همین بخش باشد
    - هیچ بخش دیگری پس از آن آخرین اقدام بازرگانی تغییری نداده باشد

    لاگ‌ها به ترتیب جدیدترین-اول هستند (ordering = ['-created_at', '-id']).
    اقدامات نادیده‌گرفته‌شده: VIEWED، CUSTOMER_NOTE، مدیر سیستم.
    """
    REVISABLE = {
        PaymentRecord.STATUS_APPROVED,
        PaymentRecord.STATUS_REJECTED,
        PaymentRecord.STATUS_INCOMPLETE,
    }
    if payment.status not in REVISABLE:
        return False

    if logs is None:
        logs = list(payment.activity_logs.all())

    IGNORABLE = {PaymentActivityLog.ACTION_VIEWED, PaymentActivityLog.ACTION_CUSTOMER_NOTE}

    # پیدا کردن جدیدترین (اولین در لیست) اقدام واقعی بازرگانی/فروش
    most_recent_commercial_idx = None
    for i, log in enumerate(logs):
        if log.action in IGNORABLE:
            continue
        if log.actor and log.actor.is_superuser:
            continue
        dept = _department_role(_user_role(log.actor)) if log.actor else ''
        if dept in {'commercial', 'sales'}:
            most_recent_commercial_idx = i
            break  # جدیدترین یافت شد

    if most_recent_commercial_idx is None:
        return False

    # آیا اقدام واقعی سایر بخش‌ها جدیدتر از آن هست؟ (ایندکس کوچک‌تر = جدیدتر)
    for log in logs[:most_recent_commercial_idx]:
        if log.action in IGNORABLE:
            continue
        if log.actor and log.actor.is_superuser:
            continue
        dept = _department_role(_user_role(log.actor)) if log.actor else ''
        if dept not in {'commercial', 'sales'}:
            return False

    return True


def _can_finance_register(role, payment, is_system_admin=False):
    """آیا مالی می‌تواند ثبت مالی انجام دهد؟ — فلگ مستقل."""
    if is_system_admin:
        return True
    role = _department_role(role)
    if role != 'finance':
        return False
    # مالی می‌تواند ثبت کند مگر:
    # - سند رد شده (rejected)
    # - سند تأیید نهایی شده (final_approved)
    # - سند ناقص است و در انتظار اصلاح مشتری (incomplete)
    # - قبلاً ثبت مالی انجام شده
    return (
        payment.status not in {
            PaymentRecord.STATUS_REJECTED,
            PaymentRecord.STATUS_FINAL_APPROVED,
            PaymentRecord.STATUS_INCOMPLETE,
        }
        and not payment.is_finance_registered
    )


def _ready_for_final_q():
    """فیلتر ORM اسناد در انتظار تأیید نهایی — از روی فلگ ذخیره‌شده"""
    return Q(pending_final_approval=True)


def _sync_pending_final_flag(payment):
    """
    همگام‌سازی فلگ pending_final_approval با وضعیت واقعی سند.
    شیء payment را در حافظه آپدیت می‌کند؛ ذخیره توسط فراخواننده انجام می‌شود.
    """
    should_be = (
        payment.status == PaymentRecord.STATUS_APPROVED
        and payment.finance_status == PaymentRecord.FINANCE_STATUS_APPROVED
        and (payment.counterparty_id is None or payment.counterparty_status == PaymentRecord.CP_STATUS_APPROVED)
    )
    if should_be and not payment.pending_final_approval:
        payment.pending_final_approval = True
        if not payment.pending_final_approval_since:
            payment.pending_final_approval_since = timezone.now()
    elif not should_be and payment.pending_final_approval:
        payment.pending_final_approval = False
        payment.pending_final_approval_since = None


def _can_see_pending_final_approval(user):
    """آیا کاربر به صفحه «در انتظار تأیید نهایی» دسترسی دارد؟"""
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    role = _user_role(user)
    if role == 'finance_manager':
        return True
    from .models import FinalApprovalDelegate
    return FinalApprovalDelegate.objects.filter(delegated_user=user, is_active=True).exists()


def _can_final_approve(role, payment, is_system_admin=False, user=None):
    """
    آیا تأیید نهایی مجاز است؟
    - مدیر مالی: همیشه (اگر هر دو فلگ آماده باشند)
    - کاربر تفویض‌شده (از FinalApprovalDelegate): اگر delegation فعال باشد
    - ادمین: همیشه
    """
    if not payment.ready_for_final_approval:
        return False
    if is_system_admin or role == 'finance_manager':
        return True
    # بررسی تفویض جهانی
    if user:
        from .models import FinalApprovalDelegate
        return FinalApprovalDelegate.objects.filter(
            delegated_user=user, is_active=True
        ).exists()
    return False


def _can_delegate_final_approval(role, is_system_admin=False):
    """آیا کاربر می‌تواند تأیید نهایی را تفویض کند؟"""
    return is_system_admin or role == 'finance_manager'


def _records_for_user(user):
    qs = PaymentRecord.objects.select_related('counterparty', 'user', 'user__profile').prefetch_related(
        'receipts',
        Prefetch('activity_logs', queryset=PaymentActivityLog.objects.select_related('actor', 'actor__profile')),
    )
    if _is_staff_user(user):
        return qs.order_by('-created_at', '-id')
    return qs.filter(user=user).order_by('-created_at', '-id')


def _active_payment_records_for_user(user):
    records = _records_for_user(user)
    if not _is_staff_user(user):
        return records  # مشتریان: همه فیش‌های خودشان

    ready_q = _ready_for_final_q()

    if user.is_superuser:
        # ادمین: همه به‌جز در‌انتظار‌تأیید‌نهایی (→ صف جداگانه)، تأیید‌نهایی‌شده و رد‌شده (→ سوابق)
        return records.filter(pending_final_approval=False).exclude(status__in=[
            PaymentRecord.STATUS_FINAL_APPROVED,
            PaymentRecord.STATUS_REJECTED,
        ])

    role = _department_role(_user_role(user))
    if role == 'commercial':
        return records.filter(status__in=[
            PaymentRecord.STATUS_PENDING,
            PaymentRecord.STATUS_COMMERCIAL_REVIEW,
            PaymentRecord.STATUS_TEMP_COMMERCIAL,
            PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL,
        ])
    if role == 'finance':
        return records.filter(status__in=[
            PaymentRecord.STATUS_PENDING,
            PaymentRecord.STATUS_COMMERCIAL_REVIEW,
            PaymentRecord.STATUS_TEMP_COMMERCIAL,
            PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL,
            PaymentRecord.STATUS_APPROVED,
            PaymentRecord.STATUS_INCOMPLETE,
        ], pending_final_approval=False)
    if role == 'data_entry':
        return records.exclude(status__in=[
            PaymentRecord.STATUS_FINAL_APPROVED,
            PaymentRecord.STATUS_REJECTED,
        ]).filter(pending_final_approval=False)
    if role == 'sales':
        filtered = records.exclude(status__in=[
            PaymentRecord.STATUS_FINAL_APPROVED,
            PaymentRecord.STATUS_REJECTED,
        ]).filter(pending_final_approval=False)
        return _customer_limited_queryset_for_user(filtered, user, customer_field='user')
    return records.none()


def _history_payment_records_for_user(user):
    if not _is_staff_user(user):
        return PaymentRecord.objects.none()

    records = _records_for_user(user)
    role = _department_role(_user_role(user))
    if user.is_superuser or role in {'commercial', 'finance', 'data_entry', 'sales'}:
        return records
    return records.none()


def _mark_commercial_records_seen(records, actor):
    if not actor.is_authenticated or actor.is_superuser or _department_role(_user_role(actor)) != 'commercial':
        return

    pending_records = list(records.filter(status=PaymentRecord.STATUS_PENDING))
    for payment in pending_records:
        from_status = payment.status
        payment.status = PaymentRecord.STATUS_COMMERCIAL_REVIEW
        payment.save(update_fields=['status'])
        _log_activity(
            payment,
            actor,
            PaymentActivityLog.ACTION_STATUS_CHANGED,
            from_status=from_status,
            to_status=payment.status,
            note='رویت توسط بازرگانی',
        )
        _log_activity(
            payment,
            actor,
            PaymentActivityLog.ACTION_VIEWED,
            note='مشاهده در صف بررسی بازرگانی',
        )


def _parse_jalali_date(date_text):
    if not date_text:
        return None
    try:
        return jdatetime.datetime.strptime(date_text, '%Y/%m/%d').date()
    except ValueError:
        return None


def _today_jalali_date():
    return jdatetime.date.fromgregorian(date=timezone.localdate(timezone=DISPLAY_TIME_ZONE))


def _format_jalali_date(value):
    if not value:
        return ''
    return value.strftime('%Y/%m/%d')


def _format_jalali_datetime(value, date_format='%Y/%m/%d %H:%M'):
    if not value:
        return ''
    if timezone.is_aware(value):
        value = timezone.localtime(value, DISPLAY_TIME_ZONE)
    return jdatetime.datetime.fromgregorian(datetime=value).strftime(date_format)


def _build_query_string(request, remove_keys=None):
    query_params = request.GET.copy()
    for key in remove_keys or []:
        query_params.pop(key, None)
    return query_params.urlencode()


def _get_page_size(request, page_param='page', default=10):
    page_size_key = 'per_page' if page_param == 'page' else page_param.replace('_page', '_per_page')
    page_size_value = (request.GET.get(page_size_key) or '').strip().lower()
    if page_size_value == 'all':
        return 'all', page_size_key
    try:
        page_size = int(page_size_value)
    except (TypeError, ValueError):
        return default, page_size_key
    return (page_size if page_size in {10, 20, 50, 100} else default), page_size_key


def _paginate_queryset(request, queryset, per_page=10, page_param='page'):
    page_size, _ = _get_page_size(request, page_param=page_param, default=per_page)
    if page_size == 'all':
        paginator = Paginator(queryset, max(len(queryset), 1))
    else:
        paginator = Paginator(queryset, page_size)
    page_number = request.GET.get(page_param) or 1
    try:
        page_obj = paginator.page(page_number)
    except (PageNotAnInteger, EmptyPage):
        page_obj = paginator.page(1)
    return page_obj


def _excel_response(filename, sheets):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{_timestamped_excel_filename(filename)}"'
    wb = Workbook()
    first = True
    for title, headers, rows in sheets:
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = title[:31]
        ws.append(headers)
        for row in rows:
            ws.append(row)
    wb.save(response)
    return response


def _timestamped_excel_filename(filename):
    base, extension = os.path.splitext(filename)
    if not extension:
        extension = '.xlsx'
    now = timezone.localtime(timezone.now(), DISPLAY_TIME_ZONE)
    timestamp = now.strftime('%Y%m%d_%H%M%S')
    return f'{base}_{timestamp}{extension}'


def _field(key, label, getter):
    return {'key': key, 'label': label, 'getter': getter}


def _selected_export_fields(request, fields):
    selected = set(request.GET.getlist('fields'))
    if not selected:
        return fields
    return [field for field in fields if field['key'] in selected] or fields


def _export_response(filename, sheet_title, fields, records):
    headers = [field['label'] for field in fields]
    rows = [
        [field['getter'](record) for field in fields]
        for record in records
    ]
    return _excel_response(filename, [(sheet_title, headers, rows)])


def _export_scope_records(request, records, page_param='page'):
    if request.GET.get('scope') != 'page':
        return records

    page_size, _ = _get_page_size(request, page_param=page_param, default=10)
    if page_size == 'all':
        return records

    paginator = Paginator(records, page_size)
    page_number = request.GET.get(page_param) or 1
    try:
        return paginator.page(page_number).object_list
    except (PageNotAnInteger, EmptyPage):
        return paginator.page(1).object_list


PAYMENT_EXPORT_FIELDS = [
    _field('id', 'ID', lambda p: p.id),
    _field('accounting_code', 'کد تفضیلی', lambda p: getattr(getattr(p.user, 'profile', None), 'accounting_code', '') if p.user else ''),
    _field('username', 'نام کاربری', lambda p: p.user.username if p.user else ''),
    _field('customer_name', 'نام مشتری', lambda p: p.user.get_full_name() if p.user else ''),
    _field('first_name', 'نام', lambda p: p.first_name),
    _field('last_name', 'نام خانوادگی', lambda p: p.last_name),
    _field('organization', 'مجموعه', lambda p: p.organization),
    _field('city', 'شهر', lambda p: p.city),
    _field('phone', 'شماره تماس', lambda p: p.phone),
    _field('payer_full_name', 'نام واریز کننده', lambda p: p.payer_full_name),
    _field('payer_account_number', 'شماره حساب واریز کننده', lambda p: p.payer_account_number),
    _field('payer_bank_name', 'بانک واریز کننده', lambda p: p.payer_bank_name),
    _field('beneficiary_bank_name', 'بانک مقصد', lambda p: p.beneficiary_bank_name),
    _field('beneficiary_account_number', 'شماره حساب مقصد', lambda p: p.beneficiary_account_number),
    _field('beneficiary_account_owner', 'صاحب حساب مقصد', lambda p: p.beneficiary_account_owner),
    _field('amount', 'مبلغ', lambda p: p.amount),
    _field('pay_date', 'تاریخ واریز', lambda p: _format_jalali_date(p.pay_date)),
    _field('tracking_code', 'کد پیگیری', lambda p: p.tracking_code or ''),
    _field('status', 'وضعیت', lambda p: p.get_status_display()),
    _field('counterparty', 'طرف حساب', lambda p: p.counterparty.name if p.counterparty else ''),
    _field('last_staff_note', 'آخرین توضیح کارشناس', lambda p: p.last_staff_note),
    _field('customer_notes', 'توضیح مشتری', lambda p: p.customer_notes),
    _field('created_at', 'تاریخ ثبت', lambda p: _format_jalali_datetime(p.created_at)),
    _field('updated_seen_at', 'زمان مشاهده مشتری', lambda p: _format_jalali_datetime(p.customer_seen_at)),
]

INVOICE_EXPORT_FIELDS = [
    _field('id', 'ID', lambda i: i.id),
    _field('accounting_code', 'کد تفضیلی', lambda i: getattr(i.customer.profile, 'accounting_code', '')),
    _field('customer_username', 'نام کاربری مشتری', lambda i: i.customer.username),
    _field('customer_name', 'نام مشتری', lambda i: i.customer.get_full_name() or i.customer.username),
    _field('organization', 'مجموعه', lambda i: getattr(i.customer.profile, 'organization', '')),
    _field('invoice_number', 'شماره فاکتور', lambda i: i.invoice_number),
    _field('invoice_date', 'تاریخ فاکتور', lambda i: _format_jalali_date(i.invoice_date)),
    _field('amount', 'مبلغ', lambda i: i.amount),
    _field('reference_number', 'شماره حواله', lambda i: i.reference_number),
    _field('uploaded_by', 'بارگذاری کننده', lambda i: i.uploaded_by.get_full_name() if i.uploaded_by else ''),
    _field('seen', 'وضعیت مشاهده', lambda i: 'دیده شده' if i.is_seen_by_customer else 'دیده نشده'),
    _field('customer_visible_note', 'توضیح قابل مشاهده مشتری', lambda i: i.customer_visible_note),
    _field('internal_note', 'توضیح داخلی', lambda i: i.internal_note),
    _field('customer_note', 'یادداشت مشتری', lambda i: i.customer_note),
    _field('created_at', 'تاریخ ثبت', lambda i: _format_jalali_datetime(i.created_at)),
]

CUSTOMER_EXPORT_FIELDS = [
    _field('accounting_code', 'کد تفضیلی', lambda c: c['profile'].accounting_code),
    _field('username', 'نام کاربری', lambda c: c['user'].username),
    _field('full_name', 'نام و نام خانوادگی', lambda c: f"{c['profile'].first_name or c['user'].first_name} {c['profile'].last_name or c['user'].last_name}".strip()),
    _field('organization', 'مجموعه', lambda c: c['profile'].organization),
    _field('city', 'شهر', lambda c: c['profile'].city),
    _field('province', 'استان', lambda c: c['profile'].province),
    _field('phone', 'شماره تماس', lambda c: c['profile'].phone),
    _field('mobile', 'شماره همراه', lambda c: c['profile'].mobile),
    _field('second_mobile', 'شماره همراه دوم', lambda c: c['profile'].second_mobile),
    _field('payment_count', 'تعداد فیش ها', lambda c: c['payment_count']),
    _field('invoice_count', 'تعداد فاکتورها', lambda c: c['invoice_count']),
    _field('invoice_total', 'جمع فاکتورها', lambda c: c['invoice_total']),
    _field('total_amount', 'جمع واریزی ها', lambda c: c['total_amount']),
    _field('review_debt', 'بدهی ممیزی نشده', lambda c: c['review_debt']),
    _field('confirmed_debt', 'بدهی تایید شده', lambda c: c['confirmed_debt']),
    _field('latest_payment_date', 'آخرین سند', lambda c: _format_jalali_datetime(c['latest_payment_date'])),
    _field('status', 'وضعیت', lambda c: 'معلق' if c['profile'].suspended else ('فعال' if c['user'].is_active else 'غیرفعال')),
]

USER_EXPORT_FIELDS = [
    _field('accounting_code', 'کد تفضیلی', lambda u: getattr(u.profile, 'accounting_code', '')),
    _field('username', 'نام کاربری', lambda u: u.username),
    _field('first_name', 'نام', lambda u: u.first_name),
    _field('last_name', 'نام خانوادگی', lambda u: u.last_name),
    _field('role', 'نقش', lambda u: u.profile.get_role_display()),
    _field('email', 'ایمیل', lambda u: u.email),
    _field('phone', 'شماره تماس', lambda u: u.profile.phone),
    _field('mobile', 'شماره همراه', lambda u: u.profile.mobile),
    _field('second_mobile', 'شماره همراه دوم', lambda u: u.profile.second_mobile),
    _field('organization', 'مجموعه', lambda u: u.profile.organization),
    _field('city', 'شهر', lambda u: u.profile.city),
    _field('province', 'استان', lambda u: u.profile.province),
    _field('active_from', 'تاریخ آغاز', lambda u: _format_jalali_date(u.profile.active_from)),
    _field('valid_until', 'تاریخ اعتبار', lambda u: _format_jalali_date(u.profile.valid_until)),
    _field('is_active', 'فعال', lambda u: 'بله' if u.is_active else 'خیر'),
    _field('suspended', 'معلق', lambda u: 'بله' if u.profile.suspended else 'خیر'),
    _field('can_edit_payment_details', 'دسترسی تکمیل اطلاعات فیش‌ها', lambda u: 'بله' if u.profile.can_edit_payment_details else 'خیر'),
]

COUNTERPARTY_EXPORT_FIELDS = [
    _field('id', 'ID', lambda c: c.id),
    _field('name', 'نام', lambda c: c.name),
    _field('description', 'توضیحات', lambda c: c.description),
    _field('created_at', 'تاریخ ثبت', lambda c: _format_jalali_datetime(c.created_at)),
    _field('updated_at', 'آخرین بروزرسانی', lambda c: _format_jalali_datetime(c.updated_at)),
]

DAILY_PLAN_EXPORT_FIELDS = [
    _field('id', 'ID', lambda p: p.id),
    _field('deposit_date', 'تاریخ', lambda p: _format_jalali_date(p.deposit_date)),
    _field('bank_name', 'بانک', lambda p: p.bank_name),
    _field('account_number', 'شماره حساب', lambda p: p.account_number),
    _field('account_owner', 'صاحب حساب', lambda p: p.account_owner),
    _field('total_expected_amount', 'مبلغ کل اعلامی', lambda p: p.total_expected_amount),
    _field('assigned_expected_total', 'جمع تخصیص مشتریان', lambda p: getattr(p, 'assigned_expected_total', 0)),
    _field('paid_total', 'واریز ممیزی نشده', lambda p: getattr(p, 'paid_total', 0)),
    _field('confirmed_total', 'واریز تایید شده', lambda p: getattr(p, 'confirmed_total', 0)),
    _field('remaining_total', 'کسری', lambda p: getattr(p, 'remaining_total', 0)),
    _field('assignment_count', 'تعداد مشتری', lambda p: getattr(p, 'assignment_count', 0)),
    _field('note', 'توضیح', lambda p: p.note),
]

DAILY_ASSIGNMENT_EXPORT_FIELDS = [
    _field('customer_username', 'نام کاربری مشتری', lambda a: a.customer.username),
    _field('customer_name', 'مشتری', lambda a: a.customer.get_full_name() or a.customer.username),
    _field('organization', 'مجموعه', lambda a: getattr(a.customer.profile, 'organization', '')),
    _field('expected_amount', 'مبلغ مورد انتظار', lambda a: a.expected_amount),
    _field('paid_amount', 'واریز ممیزی نشده', lambda a: a.report['paid_amount']),
    _field('confirmed_amount', 'واریز تایید شده', lambda a: a.report['confirmed_amount']),
    _field('remaining_amount', 'کسری ممیزی نشده', lambda a: a.remaining_amount),
    _field('confirmed_remaining_amount', 'کسری تایید شده', lambda a: a.confirmed_remaining_amount),
    _field('paid_percent', 'درصد تحقق', lambda a: a.paid_percent),
    _field('plan_status', 'وضعیت برنامه', lambda a: a.plan_status_label),
    _field('payment_count', 'تعداد فیش', lambda a: a.report['payment_count']),
    _field('latest_payment', 'آخرین فیش', lambda a: a.latest_payment_text),
    _field('note', 'توضیح', lambda a: a.note),
]


ORDER_EXPORT_FIELDS = [
    _field('order_number', 'شماره سفارش', lambda o: o.order_number),
    _field('customer_username', 'نام کاربری مشتری', lambda o: o.customer.username),
    _field('customer_name', 'مشتری', lambda o: o.customer.get_full_name() or o.customer.username),
    _field('organization', 'مجموعه', lambda o: getattr(o.customer.profile, 'organization', '')),
    _field('city', 'شهر', lambda o: getattr(o.customer.profile, 'city', '')),
    _field('province', 'استان', lambda o: getattr(o.customer.profile, 'province', '')),
    _field('title', 'عنوان سفارش', lambda o: o.title),
    _field('status', 'وضعیت', lambda o: o.get_status_display()),
    _field('sales_expert', 'کارشناس فروش', lambda o: o.sales_expert.get_full_name() or o.sales_expert.username if o.sales_expert else ''),
    _field('requested_sales_expert', 'کارشناس انتخابی مشتری', lambda o: o.requested_sales_expert.get_full_name() or o.requested_sales_expert.username if o.requested_sales_expert else ''),
    _field('items', 'اقلام', lambda o: ' | '.join(f'{item.product_name} - {item.quantity:g} {item.unit}'.strip() for item in o.items.all())),
    _field('item_count', 'تعداد اقلام', lambda o: o.items.count()),
    _field('proforma_count', 'تعداد پیش فاکتور', lambda o: o.proformas.count()),
    _field('customer_note', 'توضیح مشتری', lambda o: o.customer_note),
    _field('staff_note', 'توضیح داخلی فروش', lambda o: o.staff_note),
    _field('created_at', 'زمان ثبت', lambda o: _format_jalali_datetime(o.created_at)),
    _field('updated_at', 'آخرین بروزرسانی', lambda o: _format_jalali_datetime(o.updated_at)),
]

SALES_ASSIGNMENT_EXPORT_FIELDS = [
    _field('customer_username', 'نام کاربری مشتری', lambda r: r['customer'].username),
    _field('customer_name', 'مشتری', lambda r: r['customer'].get_full_name() or r['customer'].username),
    _field('organization', 'مجموعه', lambda r: r['profile'].organization),
    _field('city', 'شهر', lambda r: r['profile'].city),
    _field('province', 'استان', lambda r: r['profile'].province),
    _field('sales_user', 'کارشناس فروش', lambda r: r['sales_user'].get_full_name() or r['sales_user'].username if r['sales_user'] else ''),
    _field('assigned_by', 'تخصیص دهنده', lambda r: r['assigned_by'].get_full_name() or r['assigned_by'].username if r['assigned_by'] else ''),
    _field('open_orders', 'سفارش های باز', lambda r: r['open_orders']),
    _field('updated_at', 'آخرین بروزرسانی', lambda r: _format_jalali_datetime(r['updated_at'])),
    _field('note', 'توضیح', lambda r: r['note']),
]


def _customer_list_rows(request):
    customers = UserProfile.objects.filter(role='customer').select_related('user').order_by('user__username')
    filters = {
        'q': (request.GET.get('q') or '').strip(),
        'status': (request.GET.get('status') or '').strip(),
    }
    if filters['q']:
        customers = customers.filter(
            Q(user__username__icontains=filters['q']) |
            Q(user__first_name__icontains=filters['q']) |
            Q(user__last_name__icontains=filters['q']) |
            Q(first_name__icontains=filters['q']) |
            Q(last_name__icontains=filters['q']) |
            Q(organization__icontains=filters['q']) |
            Q(city__icontains=filters['q']) |
            Q(province__icontains=filters['q']) |
            Q(phone__icontains=filters['q']) |
            Q(mobile__icontains=filters['q']) |
            Q(representative_name__icontains=filters['q']) |
            Q(representative_mobile__icontains=filters['q'])
        )
    if filters['status'] == 'active':
        customers = customers.filter(suspended=False, user__is_active=True)
    elif filters['status'] == 'suspended':
        customers = customers.filter(suspended=True)
    elif filters['status'] == 'inactive':
        customers = customers.filter(user__is_active=False)

    payment_stats = {
        row['user']: row
        for row in (
            PaymentRecord.objects
            .filter(user__profile__role='customer')
            .values('user')
            .annotate(payment_count=Count('id'), total_amount=Sum('amount'), latest_payment_date=Max('created_at'))
        )
    }
    confirmed_payment_totals = {
        row['user']: row['total'] or 0
        for row in (
            PaymentRecord.objects
            .filter(user__profile__role='customer', status__in=[PaymentRecord.STATUS_APPROVED, PaymentRecord.STATUS_FINAL_APPROVED])
            .values('user')
            .annotate(total=Sum('amount'))
        )
    }
    review_payment_totals = {
        row['user']: row['total'] or 0
        for row in (
            PaymentRecord.objects
            .filter(user__profile__role='customer')
            .exclude(status=PaymentRecord.STATUS_REJECTED)
            .values('user')
            .annotate(total=Sum('amount'))
        )
    }
    invoice_counts = {
        row['customer']: row['invoice_count']
        for row in InvoiceRecord.objects.filter(customer__profile__role='customer').values('customer').annotate(invoice_count=Count('id'))
    }
    invoice_totals = {
        row['customer']: row['total'] or 0
        for row in InvoiceRecord.objects.filter(customer__profile__role='customer').values('customer').annotate(total=Sum('amount'))
    }

    customer_data = []
    for profile in customers:
        stats = payment_stats.get(profile.user_id, {})
        invoice_total = invoice_totals.get(profile.user_id, 0)
        confirmed_payment_total = confirmed_payment_totals.get(profile.user_id, 0)
        review_payment_total = review_payment_totals.get(profile.user_id, 0)
        customer_data.append({
            'profile': profile,
            'user': profile.user,
            'payment_count': stats.get('payment_count') or 0,
            'invoice_count': invoice_counts.get(profile.user_id, 0),
            'total_amount': stats.get('total_amount') or 0,
            'invoice_total': invoice_total,
            'confirmed_debt': invoice_total - confirmed_payment_total,
            'review_debt': invoice_total - review_payment_total,
            'latest_payment_date': stats.get('latest_payment_date'),
        })
    return customer_data, filters


def _enrich_daily_plans(plans):
    for plan in plans:
        assignments = list(plan.assignments.all())
        stats = _daily_assignment_stats(assignments)
        plan.assignment_count = len(assignments)
        plan.assigned_expected_total = sum(assignment.expected_amount for assignment in assignments)
        plan.paid_total = sum(stats.get(assignment.id, {}).get('paid_amount', 0) for assignment in assignments)
        plan.confirmed_total = sum(stats.get(assignment.id, {}).get('confirmed_amount', 0) for assignment in assignments)
        plan.remaining_total = plan.assigned_expected_total - plan.paid_total
    return plans


def _daily_plans_for_period(start_date, end_date):
    plans = list(
        DailyPaymentPlan.objects
        .select_related('created_by')
        .prefetch_related('assignments')
        .filter(deposit_date__gte=start_date, deposit_date__lte=end_date)
        .order_by('-deposit_date', '-id')
    )
    return _enrich_daily_plans(plans)


def _daily_plans_for_date(selected_date):
    return _daily_plans_for_period(selected_date, selected_date)


def _daily_payment_period(request):
    mode = (request.GET.get('mode') or 'day').strip()
    if mode not in {'day', 'week', 'month', 'range'}:
        mode = 'day'
    selected_date = _parse_jalali_date((request.GET.get('date') or '').strip()) or _today_jalali_date()

    if mode == 'range':
        start_date = _parse_jalali_date((request.GET.get('start_date') or '').strip()) or selected_date
        end_date = _parse_jalali_date((request.GET.get('end_date') or '').strip()) or start_date
        if end_date < start_date:
            start_date, end_date = end_date, start_date
        span_days = (end_date.togregorian() - start_date.togregorian()).days + 1
        shift = jdatetime.timedelta(days=span_days)
        label = f'بازه {start_date.strftime("%Y/%m/%d")} تا {end_date.strftime("%Y/%m/%d")}'
        previous_date = start_date - shift
        next_date = start_date + shift
        previous_start_date = start_date - shift
        previous_end_date = end_date - shift
        next_start_date = start_date + shift
        next_end_date = end_date + shift
    elif mode == 'week':
        end_date = selected_date
        start_date = selected_date - jdatetime.timedelta(days=6)
        label = f'۷ روز گذشته {start_date.strftime("%Y/%m/%d")} تا {end_date.strftime("%Y/%m/%d")}'
        previous_date = selected_date - jdatetime.timedelta(days=7)
        next_date = selected_date + jdatetime.timedelta(days=7)
        previous_start_date = None
        previous_end_date = None
        next_start_date = None
        next_end_date = None
    elif mode == 'month':
        end_date = selected_date
        start_date = selected_date - jdatetime.timedelta(days=29)
        label = f'۳۰ روز گذشته {start_date.strftime("%Y/%m/%d")} تا {end_date.strftime("%Y/%m/%d")}'
        previous_date = selected_date - jdatetime.timedelta(days=30)
        next_date = selected_date + jdatetime.timedelta(days=30)
        previous_start_date = None
        previous_end_date = None
        next_start_date = None
        next_end_date = None
    else:
        start_date = selected_date
        end_date = selected_date
        label = f'روز {selected_date.strftime("%Y/%m/%d")}'
        previous_date = selected_date - jdatetime.timedelta(days=1)
        next_date = selected_date + jdatetime.timedelta(days=1)
        previous_start_date = None
        previous_end_date = None
        next_start_date = None
        next_end_date = None

    return {
        'mode': mode,
        'selected_date': selected_date,
        'start_date': start_date,
        'end_date': end_date,
        'label': label,
        'previous_date': previous_date,
        'next_date': next_date,
        'previous_start_date': previous_start_date,
        'previous_end_date': previous_end_date,
        'next_start_date': next_start_date,
        'next_end_date': next_end_date,
    }


def _daily_period_query(mode, date_value, start_date=None, end_date=None):
    query = {'mode': mode, 'date': _format_jalali_date(date_value)}
    if mode == 'range':
        query['start_date'] = _format_jalali_date(start_date or date_value)
        query['end_date'] = _format_jalali_date(end_date or start_date or date_value)
    return urlencode(query)


def _daily_assignments_for_plan(plan):
    assignments = list(
        plan.assignments
        .select_related('customer', 'customer__profile')
        .prefetch_related('payments', 'payments__receipts')
        .all()
    )
    return _enrich_daily_assignments(assignments)


def _enrich_daily_assignments(assignments):
    stats = _daily_assignment_stats(assignments)
    for assignment in assignments:
        assignment.report = stats.get(assignment.id, {
            'paid_amount': 0,
            'payment_count': 0,
            'confirmed_amount': 0,
            'confirmed_count': 0,
            'latest_payment': None,
        })
        assignment.remaining_amount = assignment.expected_amount - assignment.report['paid_amount']
        assignment.confirmed_remaining_amount = assignment.expected_amount - assignment.report['confirmed_amount']
        paid_amount = assignment.report['paid_amount']
        assignment.paid_percent = round((paid_amount / assignment.expected_amount) * 100, 1) if assignment.expected_amount else 0
        if paid_amount <= 0:
            assignment.plan_status_label = 'بدون واریز'
            assignment.plan_status_class = 'flag-red'
        elif paid_amount < assignment.expected_amount:
            assignment.plan_status_label = 'واریز ناقص'
            assignment.plan_status_class = 'flag-yellow'
        elif paid_amount == assignment.expected_amount:
            assignment.plan_status_label = 'تکمیل شده'
            assignment.plan_status_class = 'flag-green'
        else:
            assignment.plan_status_label = 'بیش از برنامه'
            assignment.plan_status_class = 'flag-blue'
        assignment.latest_payment_text = _format_jalali_datetime(assignment.report.get('latest_payment'))
    return assignments


def _customer_daily_assignments_for_user(user, request=None):
    assignments = (
        DailyPaymentAssignment.objects
        .select_related('plan', 'customer', 'customer__profile')
        .prefetch_related('payments', 'payments__receipts')
        .filter(customer=user)
        .order_by('-plan__deposit_date', '-id')
    )
    if request:
        start_date = _parse_jalali_date((request.GET.get('start_date') or '').strip())
        end_date = _parse_jalali_date((request.GET.get('end_date') or '').strip())
        if start_date:
            assignments = assignments.filter(plan__deposit_date__gte=start_date)
        if end_date:
            assignments = assignments.filter(plan__deposit_date__lte=end_date)
    return _enrich_daily_assignments(list(assignments))


def _log_activity(payment, actor, action, from_status='', to_status='', note=''):
    PaymentActivityLog.objects.create(
        payment=payment,
        actor=actor if actor and actor.is_authenticated else None,
        action=action,
        from_status=from_status or '',
        to_status=to_status or '',
        note=note or '',
    )


PAYMENT_CUSTOMER_EDIT_BLOCKING_ACTIONS = {
    PaymentActivityLog.ACTION_EDITED,
    PaymentActivityLog.ACTION_STATUS_CHANGED,
    PaymentActivityLog.ACTION_FINANCE_REGISTERED,
    PaymentActivityLog.ACTION_FINAL_APPROVED,
    PaymentActivityLog.ACTION_CP_APPROVED,
    PaymentActivityLog.ACTION_CP_RETURNED,
    PaymentActivityLog.ACTION_CP_REJECTED,
}


def _payment_has_non_customer_operation(payment):
    if payment.status != PaymentRecord.STATUS_PENDING:
        return True
    if payment.finance_status == PaymentRecord.FINANCE_STATUS_APPROVED:
        return True
    if payment.finance_registered_at or payment.finance_registered_by_id:
        return True
    if payment.counterparty_status:
        return True

    logs = getattr(payment, '_prefetched_objects_cache', {}).get('activity_logs')
    if logs is None:
        logs = payment.activity_logs.select_related('actor').all()
    for log in logs:
        if log.action not in PAYMENT_CUSTOMER_EDIT_BLOCKING_ACTIONS:
            continue
        if log.actor_id and log.actor_id == payment.user_id:
            continue
        return True
    return False


def _can_customer_edit_payment(payment):
    return not _payment_has_non_customer_operation(payment)


def _notification_payload(notification):
    # آیکون بر اساس دسته‌بندی
    icon_map = {
        UserNotification.CATEGORY_PAYMENT: '💳',
        UserNotification.CATEGORY_INVOICE: '📄',
        UserNotification.CATEGORY_SYSTEM:  '🔔',
    }
    icon = icon_map.get(notification.category, '🔔')

    # متن زمانی نسبی
    now = timezone.now()
    delta = now - notification.created_at
    if delta.total_seconds() < 60:
        time_label = 'همین الان'
    elif delta.total_seconds() < 3600:
        mins = int(delta.total_seconds() / 60)
        time_label = f'{mins} دقیقه پیش'
    elif delta.total_seconds() < 86400:
        hours = int(delta.total_seconds() / 3600)
        time_label = f'{hours} ساعت پیش'
    else:
        time_label = _format_jalali_datetime(notification.created_at)

    return {
        'id':         notification.id,
        'title':      notification.title,
        'message':    notification.message,
        'url':        notification.resolved_url,
        'category':   notification.category,
        'icon':       icon,
        'time_label': time_label,
        'created_at': _format_jalali_datetime(notification.created_at),
    }


def _mark_notifications_read_for_url(user, url):
    if not user.is_authenticated or not url:
        return 0

    path = url.split('?', 1)[0]
    if not path:
        return 0

    condition = Q(url=path) | Q(url__startswith=f'{path}?')
    payment_match = re.search(r'/payments/(\d+)/timeline/?$', path)
    if payment_match:
        condition |= Q(
            url=reverse('submit'),
            title__contains='طرف حساب',
            message__contains=f'#{payment_match.group(1)}',
        )

    now = timezone.now()
    return UserNotification.objects.filter(
        condition,
        user=user,
        is_read=False,
    ).update(is_read=True, read_at=now)


def _notify_users(users, title, message, url='', category=UserNotification.CATEGORY_SYSTEM, actor=None, sms_message=None):
    from .sms_service import notify_sms
    seen_user_ids = set()
    notifications = []
    for user in users:
        if not user or not user.id or user.id in seen_user_ids or not user.is_active:
            continue
        seen_user_ids.add(user.id)
        notifications.append(UserNotification(
            user=user,
            actor=actor if actor and actor.is_authenticated else None,
            title=title,
            message=message,
            url=url,
            category=category,
        ))
        # ارسال پیامک اطلاع‌رسانی (اگر فعال باشد)
        if sms_message:
            notify_sms(user, sms_message, purpose='notification')
    if notifications:
        UserNotification.objects.bulk_create(notifications)


def _staff_notification_users(roles=None, exclude_user=None):
    role_filter = Q(is_superuser=True)
    if roles:
        expanded_roles = set(roles)
        if 'commercial' in expanded_roles:
            expanded_roles.add('commercial_manager')
        if 'finance' in expanded_roles:
            expanded_roles.add('finance_manager')
        if 'sales' in expanded_roles:
            expanded_roles.add('sales_manager')
        role_filter |= Q(profile__role__in=expanded_roles)
    else:
        role_filter |= Q(profile__role__in=STAFF_ROLES)
    users = User.objects.filter(role_filter, is_active=True).distinct()
    if exclude_user and exclude_user.is_authenticated:
        users = users.exclude(id=exclude_user.id)
    return users


def _notify_payment_created(payment, actor):
    customer_name = f"{payment.first_name} {payment.last_name}".strip() or (payment.user.username if payment.user else '-')
    if payment.daily_assignment_id:
        title = 'واریز برنامه‌ریزی‌شده ثبت شد'
        plan = payment.daily_assignment.plan
        message = f'{customer_name} برای برنامه واریز {plan.deposit_date} فیش ثبت کرد.'
        roles = {'commercial', 'sales', 'finance'}
    else:
        title = 'فیش واریزی جدید'
        message = f'یک فیش واریزی جدید برای {customer_name} ثبت شد.'
        roles = {'commercial', 'finance'}
    _notify_users(
        _staff_notification_users(roles=roles, exclude_user=actor),
        title,
        message,
        reverse('payment_timeline', args=[payment.id]),
        category=UserNotification.CATEGORY_PAYMENT,
        actor=actor,
    )


def _notify_payment_status_changed(payment, actor, from_status, to_status):
    status_labels = dict(PaymentRecord.STATUS_CHOICES)
    status_text = status_labels.get(to_status, to_status)
    recipients = []
    if payment.user_id and (not actor or payment.user_id != actor.id):
        recipients.append(payment.user)

    if to_status == PaymentRecord.STATUS_APPROVED:
        recipients.extend(_staff_notification_users(roles={'finance'}, exclude_user=actor))
    elif to_status == PaymentRecord.STATUS_TEMP_COMMERCIAL:
        recipients.extend(_staff_notification_users(roles={'finance'}, exclude_user=actor))
    elif to_status == PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL:
        recipients.extend(_staff_notification_users(roles={'commercial'}, exclude_user=actor))
    elif to_status in {PaymentRecord.STATUS_FINAL_APPROVED, PaymentRecord.STATUS_REJECTED, PaymentRecord.STATUS_INCOMPLETE}:
        recipients.extend(_staff_notification_users(roles={'commercial', 'finance'}, exclude_user=actor))

    customer_name = f"{payment.first_name} {payment.last_name}".strip() or (payment.user.username if payment.user else f'#{payment.id}')
    _notify_users(
        recipients,
        'تغییر وضعیت فیش',
        f'وضعیت فیش #{payment.id} مشتری {customer_name} به «{status_text}» تغییر کرد.',
        reverse('payment_timeline', args=[payment.id]),
        category=UserNotification.CATEGORY_PAYMENT,
        actor=actor,
    )


def _notify_payment_edited(payment, actor, title='ویرایش فیش واریزی'):
    recipients = []
    if payment.user_id and (not actor or payment.user_id != actor.id):
        recipients.append(payment.user)
    recipients.extend(_staff_notification_users(roles={'commercial', 'finance'}, exclude_user=actor))
    customer_name = f"{payment.first_name} {payment.last_name}".strip() or (payment.user.username if payment.user else f'#{payment.id}')
    _notify_users(
        recipients,
        title,
        f'اطلاعات فیش #{payment.id} مشتری {customer_name} بروزرسانی شد.',
        reverse('payment_timeline', args=[payment.id]),
        category=UserNotification.CATEGORY_PAYMENT,
        actor=actor,
    )


def _notify_invoice_created(invoice, actor):
    invoice_ref = invoice.invoice_number or f'#{invoice.id}'
    _notify_users(
        [invoice.customer],
        'فاکتور جدید',
        f'فاکتور {invoice_ref} برای شما ثبت شد.',
        reverse('invoice_detail', args=[invoice.id]),
        category=UserNotification.CATEGORY_INVOICE,
        actor=actor,
    )


def _invoice_staff_notification_users(exclude_user=None):
    role_filter = (
        Q(is_superuser=True)
        | Q(profile__role__in={'sales', 'sales_manager', 'commercial_manager', 'finance_manager'})
        | Q(profile__can_view_invoices=True)
    )
    users = User.objects.filter(role_filter, is_active=True).distinct()
    if exclude_user and exclude_user.is_authenticated:
        users = users.exclude(id=exclude_user.id)
    return users


def _notify_payment_customer_note(payment, actor, note_text):
    customer_name = f"{payment.first_name} {payment.last_name}".strip() or (payment.user.username if payment.user else '-')
    _notify_users(
        _staff_notification_users(roles={'commercial', 'finance'}, exclude_user=actor),
        'یادداشت جدید مشتری روی فیش',
        f'{customer_name} یک توضیح جدید روی فیش واریزی ثبت کرد: «{note_text[:120]}»',
        reverse('payment_timeline', args=[payment.id]),
        category=UserNotification.CATEGORY_PAYMENT,
        actor=actor,
    )


def _notify_invoice_customer_note(invoice, actor):
    customer_name = invoice.customer.get_full_name().strip() or invoice.customer.username
    _notify_users(
        _invoice_staff_notification_users(exclude_user=actor),
        'یادداشت جدید مشتری روی فاکتور',
        f'{customer_name} یک یادداشت جدید روی فاکتور ثبت کرد.',
        reverse('invoice_detail', args=[invoice.id]),
        category=UserNotification.CATEGORY_INVOICE,
        actor=actor,
    )


def _is_superuser_actor(log):
    """True اگر log توسط مدیر سیستم ثبت شده باشد — برای حذف از نمایش تاریخچه."""
    return bool(log.actor_id and log.actor and log.actor.is_superuser)


def _role_title(user):
    if not user:
        return 'کاربر'
    try:
        role = user.profile.role
    except UserProfile.DoesNotExist:
        role = ''
    return {
        'commercial': 'کاربر بازرگانی',
        'finance': 'کاربر مالی',
        'sales': 'کاربر فروش',
        'data_entry': 'کاربر تکمیل اطلاعات فیش',
        'staff': 'کاربر کارمند',
        'customer': 'مشتری',
    }.get(role, 'کاربر')


def _display_name(user):
    if not user:
        return 'سیستم'
    full_name = f"{user.first_name} {user.last_name}".strip()
    return full_name or user.username


def _group_consecutive_views(logs):
    """مشاهده‌های پیاپی یک کاربر را زیر یک رویداد جمع می‌کند تا تاریخچه شلوغ نشود."""
    grouped = []
    for log in logs:
        log['children'] = []
        previous = grouped[-1] if grouped else None
        if (
            previous
            and log['action'] == PaymentActivityLog.ACTION_VIEWED
            and previous['action'] == PaymentActivityLog.ACTION_VIEWED
            and previous['actor_id'] == log['actor_id']
        ):
            previous['children'].append(log)
        else:
            grouped.append(log)
    return grouped


def _log_text(log):
    """متن کامل رویداد برای کارکنان — با نام، نقش و جزئیات."""
    actor = _display_name(log.actor)
    role = _role_title(log.actor)
    status_labels = dict(PaymentRecord.STATUS_CHOICES)

    if log.action == PaymentActivityLog.ACTION_CREATED:
        return f"📄 {role} ({actor}) سند را بارگذاری کرد."

    if log.action == PaymentActivityLog.ACTION_EDITED:
        from_text = status_labels.get(log.from_status, '') if log.from_status else ''
        return f"✏️ {role} ({actor}) سند را ویرایش و مجدد ارسال کرد." + (f" (از وضعیت «{from_text}»)" if from_text else "")

    if log.action == PaymentActivityLog.ACTION_STATUS_CHANGED:
        from_text = status_labels.get(log.from_status, log.from_status or '')
        to_text   = status_labels.get(log.to_status,   log.to_status or '')
        icon = {
            PaymentRecord.STATUS_APPROVED:   '🏬',
            PaymentRecord.STATUS_TEMP_COMMERCIAL: '📋',
            PaymentRecord.STATUS_REJECTED:   '🚫',
            PaymentRecord.STATUS_INCOMPLETE: '⚠',
            PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL: '↩',
            PaymentRecord.STATUS_COMMERCIAL_REVIEW: '🔍',
        }.get(log.to_status, '🔄')
        base = f"{icon} {role} ({actor}) فلگ بازرگانی را"
        if from_text and to_text:
            return f"{base} از «{from_text}» به «{to_text}» تغییر داد."
        return f"{base} به «{to_text}» تغییر داد."

    if log.action == PaymentActivityLog.ACTION_FINANCE_REGISTERED:
        return f"💰 {role} ({actor}) ثبت مالی انجام داد."

    if log.action == PaymentActivityLog.ACTION_FINAL_APPROVED:
        return f"✅ {role} ({actor}) سند را تأیید نهایی کرد."

    if log.action == PaymentActivityLog.ACTION_CUSTOMER_NOTE:
        return f"💬 مشتری ({actor}) توضیح اضافه کرد."

    if log.action == PaymentActivityLog.ACTION_VIEWED:
        return f"👁 {role} ({actor}) سند را مشاهده کرد."

    if log.action == PaymentActivityLog.ACTION_CP_APPROVED:
        return f"✅ طرف حساب ({actor}) فیش را تایید کرد."
    if log.action == PaymentActivityLog.ACTION_CP_RETURNED:
        return f"⚠ طرف حساب ({actor}) فیش را عودت داد."
    if log.action == PaymentActivityLog.ACTION_CP_REJECTED:
        return f"🚫 طرف حساب ({actor}) فیش را رد کرد."

    return f"🔄 {role} ({actor}) عملیاتی انجام داد."


def _customer_log_text(log):
    """متن ساده‌شده رویداد برای مشتری — بدون نام کارکنان و جزئیات."""
    if log.action == PaymentActivityLog.ACTION_CREATED:
        return 'فیش شما با موفقیت ثبت شد.'
    if log.action == PaymentActivityLog.ACTION_EDITED:
        return 'فیش شما ویرایش و مجدد ارسال شد.'
    if log.action == PaymentActivityLog.ACTION_STATUS_CHANGED:
        if log.to_status == PaymentRecord.STATUS_INCOMPLETE:
            return 'فیش نیاز به تکمیل مدارک دارد.'
        if log.to_status == PaymentRecord.STATUS_REJECTED:
            return 'فیش رد شد.'
        # سایر تغییرات وضعیت داخلی برای مشتری نمایش نمی‌یابد
        return ''
    if log.action == PaymentActivityLog.ACTION_FINAL_APPROVED:
        return 'فیش تأیید نهایی شد.'
    # ثبت مالی و بررسی‌های داخلی نشان داده نمی‌شود
    return ''


def _customer_visible_logs(logs):
    """
    فیلتر و فرمت لاگ‌ها برای مشتری:
    - وضعیت‌های مهم (ثبت، ناقص، رد، تأیید نهایی)
    - توضیحات خود مشتری
    - بدون نام کارکنان یا جزئیات داخلی
    """
    visible = []
    seen_keys = set()
    for log in logs:
        # توضیحات مشتری — همه نشان داده می‌شوند
        if log.action == PaymentActivityLog.ACTION_CUSTOMER_NOTE:
            visible.append({
                'text': '💬 توضیح شما:',
                'note': log.note,
                'time': _format_jalali_datetime(log.created_at),
                'is_customer_note': True,
                'icon': '💬',
            })
            continue

        text = _customer_log_text(log)
        if not text:
            continue

        # dedup: یک نوع رویداد یک بار نشان داده می‌شود
        key = (log.action, log.to_status or text)
        if key in seen_keys:
            continue
        seen_keys.add(key)

        # یادداشت کارشناس فقط برای ناقص یا رد
        customer_note = ''
        if log.to_status in {PaymentRecord.STATUS_INCOMPLETE, PaymentRecord.STATUS_REJECTED} and log.note:
            customer_note = log.note

        icon = {
            PaymentActivityLog.ACTION_CREATED:        '📄',
            PaymentActivityLog.ACTION_EDITED:         '✏️',
            PaymentActivityLog.ACTION_FINAL_APPROVED: '✅',
        }.get(log.action, '🔄')
        if log.to_status == PaymentRecord.STATUS_INCOMPLETE:
            icon = '⚠'
        elif log.to_status == PaymentRecord.STATUS_REJECTED:
            icon = '🚫'

        visible.append({
            'text': text,
            'note': customer_note,
            'time': _format_jalali_datetime(log.created_at),
            'is_customer_note': False,
            'icon': icon,
        })
    return visible


def _enrich_records(records, staff_role='', is_system_admin=False, can_edit_payment_details=False, acting_user=None):
    status_order = [
        PaymentRecord.STATUS_COMMERCIAL_REVIEW,
        PaymentRecord.STATUS_TEMP_COMMERCIAL,
        PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL,
        PaymentRecord.STATUS_APPROVED,
        PaymentRecord.STATUS_FINAL_APPROVED,
        PaymentRecord.STATUS_REJECTED,
        PaymentRecord.STATUS_INCOMPLETE,
    ]
    records = list(records)
    for payment in records:
        reached = set()
        for log in payment.activity_logs.all():
            if log.to_status in STATUS_FLAG_META:
                reached.add(log.to_status)
        if payment.status in STATUS_FLAG_META:
            reached.add(payment.status)
            for step in STATUS_PROGRESS_FLOWS.get(payment.status, []):
                reached.add(step)

        payment.row_flags = [
            {
                'label': STATUS_FLAG_META[code][0],
                'css': STATUS_FLAG_META[code][1],
            }
            for code in status_order
            if code in reached
        ]
        if staff_role:
            raw_lines = [
                {
                    'time': _format_jalali_datetime(log.created_at),
                    'text': _log_text(log),
                    'note': log.note,
                    'action': log.action,
                    'actor_id': log.actor_id,
                }
                for log in list(payment.activity_logs.all()[:20])
                if not _is_superuser_actor(log)
            ]
            payment.timeline_lines = _group_consecutive_views(raw_lines)[:5]
        else:
            payment.timeline_lines = _customer_visible_logs(payment.activity_logs.all())[:5]
        payment.staff_can_act = _can_staff_act_on_payment(
            staff_role, payment, is_system_admin=is_system_admin,
        ) if staff_role else False
        payment.staff_allowed_choices = _staff_status_choices_for_role(staff_role) if staff_role else []
        payment.can_edit_details = bool(can_edit_payment_details)
        payment.can_customer_edit = _can_customer_edit_payment(payment)

        # فلگ‌های مستقل مالی
        payment.can_finance_register = _can_finance_register(
            staff_role, payment, is_system_admin=is_system_admin,
        ) if staff_role else False

        # اقدام بازرگانی — فقط بازرگانی/فروش/ادمین (نه مالی)
        _dept = _department_role(staff_role) if staff_role else ''

        # تجدیدنظر بازرگانی: اگر آخرین اقدام واقعی از بازرگانی بوده، امکان اصلاح وجود دارد
        if not payment.staff_can_act and _dept in {'commercial', 'sales'}:
            payment.staff_can_act = _commercial_can_revise(payment)

        payment.can_commercial_act = (
            _dept in {'commercial', 'sales'} and payment.staff_can_act
        ) or (
            is_system_admin and payment.staff_can_act and _dept not in {'finance'}
        )
        payment.can_final_approve = _can_final_approve(
            staff_role, payment, is_system_admin=is_system_admin, user=acting_user,
        ) if staff_role else False
        payment.can_delegate = _can_delegate_final_approval(staff_role, is_system_admin)

        # وضعیت‌های مجاز مالی (برای dropdown)
        dept = _department_role(staff_role) if staff_role else ''
        # مالی، مدیر مالی، و ادمین به بخش مالی دسترسی دارند
        is_finance_actor = dept == 'finance' or is_system_admin
        finance_choices = []
        if is_finance_actor:
            if payment.can_finance_register:
                finance_choices.append(('finance_register', 'ثبت مالی'))
            # عودت وقتی بازرگانی سند را به یکی از وضعیت‌های ثبت بازرگانی، ناقص یا رد تغییر داده باشد
            if payment.status in {
                PaymentRecord.STATUS_APPROVED,
                PaymentRecord.STATUS_INCOMPLETE,
                PaymentRecord.STATUS_REJECTED,
            }:
                finance_choices.append(('return_to_commercial', 'عودت به بازرگانی'))
        payment.finance_choices = finance_choices
        payment.is_finance_actor = is_finance_actor

        # وضعیت طرف حساب
        cs = payment.counterparty_status
        if cs == PaymentRecord.CP_STATUS_APPROVED:
            payment.cp_label      = 'تایید'
            payment.cp_class      = 'flag-green'
            payment.cp_needs_act  = False
        elif cs == PaymentRecord.CP_STATUS_RETURNED:
            payment.cp_label      = 'عودت — نیاز به اقدام'
            payment.cp_class      = 'flag-orange'
            payment.cp_needs_act  = True   # بازرگانی باید اقدام کند
        elif cs == PaymentRecord.CP_STATUS_REJECTED:
            payment.cp_label      = 'رد / ابطال'
            payment.cp_class      = 'flag-red'
            payment.cp_needs_act  = False
        elif payment.counterparty_id:
            payment.cp_label      = 'در انتظار'
            payment.cp_class      = 'flag-gray'
            payment.cp_needs_act  = False
        else:
            payment.cp_label      = ''
            payment.cp_class      = ''
            payment.cp_needs_act  = False
        payment.cp_note = payment.counterparty_note or ''
        payment.cp_icon = {'تایید': '✅', 'عودت — نیاز به اقدام': '⚠', 'رد / ابطال': '🚫', 'در انتظار': '⏳'}.get(payment.cp_label, '')

    return records


def _apply_record_filters(records, request, is_staff_user):
    filters = {
        'first_name': (request.GET.get('first_name') or '').strip(),
        'last_name': (request.GET.get('last_name') or '').strip(),
        'phone': (request.GET.get('phone') or '').strip(),
        'city': (request.GET.get('city') or '').strip(),
        'tracking_code': (request.GET.get('tracking_code') or '').strip(),
        'payer_account_number': (request.GET.get('payer_account_number') or '').strip(),
        'payer_full_name': (request.GET.get('payer_full_name') or '').strip(),
        'payer_bank_name': (request.GET.get('payer_bank_name') or '').strip(),
        'amount': (request.GET.get('amount') or '').replace(',', '').strip(),
        'pay_date': (request.GET.get('pay_date') or '').strip(),
        'pay_date_from': (request.GET.get('pay_date_from') or '').strip(),
        'pay_date_to': (request.GET.get('pay_date_to') or '').strip(),
        'status': (request.GET.get('status') or '').strip(),
        'counterparty':    (request.GET.get('counterparty') or '').strip(),
        'cp_status':       (request.GET.get('cp_status') or '').strip(),
        'serial':          (request.GET.get('serial') or '').strip(),
        'accounting_code': (request.GET.get('accounting_code') or '').strip(),
        'rejection_reason': (request.GET.get('rejection_reason') or '').strip(),
    }

    if is_staff_user:
        if filters['accounting_code']:
            records = records.filter(user__profile__accounting_code__icontains=filters['accounting_code'])
        if filters['first_name']:
            records = records.filter(first_name__icontains=filters['first_name'])
        if filters['last_name']:
            records = records.filter(last_name__icontains=filters['last_name'])
        if filters['phone']:
            records = records.filter(phone__icontains=filters['phone'])
        if filters['city']:
            records = records.filter(city__icontains=filters['city'])
        if filters['tracking_code']:
            records = records.filter(tracking_code__icontains=filters['tracking_code'])
        if filters['payer_account_number']:
            records = records.filter(payer_account_number__icontains=filters['payer_account_number'])
        if filters['payer_full_name']:
            records = records.filter(payer_full_name__icontains=filters['payer_full_name'])
        if filters['payer_bank_name']:
            records = records.filter(payer_bank_name__icontains=filters['payer_bank_name'])
        if filters['counterparty'].isdigit():
            records = records.filter(counterparty_id=int(filters['counterparty']))
    else:
        if filters['payer_full_name']:
            records = records.filter(payer_full_name__icontains=filters['payer_full_name'])
        if filters['payer_account_number']:
            records = records.filter(payer_account_number__icontains=filters['payer_account_number'])
        if filters['payer_bank_name']:
            records = records.filter(payer_bank_name__icontains=filters['payer_bank_name'])

    if filters['amount'].isdigit():
        parsed_amount = int(filters['amount'])
        records = records.filter(amount=parsed_amount)
        filters['amount'] = _format_thousand_separator(parsed_amount)

    parsed_date = _parse_jalali_date(filters['pay_date'])
    if parsed_date:
        records = records.filter(pay_date=parsed_date)
    else:
        parsed_date_from = _parse_jalali_date(filters['pay_date_from'])
        parsed_date_to = _parse_jalali_date(filters['pay_date_to'])
        if parsed_date_from and parsed_date_to and parsed_date_to < parsed_date_from:
            parsed_date_from, parsed_date_to = parsed_date_to, parsed_date_from
            filters['pay_date_from'], filters['pay_date_to'] = filters['pay_date_to'], filters['pay_date_from']
        if parsed_date_from:
            records = records.filter(pay_date__gte=parsed_date_from)
        if parsed_date_to:
            records = records.filter(pay_date__lte=parsed_date_to)

    valid_statuses = {choice[0] for choice in PaymentRecord.STATUS_CHOICES}
    if is_staff_user:
        if filters['status'] == 'finance_ok':
            # فیلتر ثبت مالی — فلگ مستقل مالی
            records = records.filter(finance_status=PaymentRecord.FINANCE_STATUS_APPROVED)
        elif filters['status'] == PaymentRecord.STAFF_FILTER_FINANCE_PENDING:
            # کلیه اسنادی که فلگ مستقل مالی آن‌ها «در انتظار ثبت مالی» است
            records = records.exclude(status__in=[
                PaymentRecord.STATUS_FINAL_APPROVED,
                PaymentRecord.STATUS_REJECTED,
                PaymentRecord.STATUS_INCOMPLETE,
            ]).exclude(finance_status=PaymentRecord.FINANCE_STATUS_APPROVED)
        elif filters['status'] == PaymentRecord.STAFF_FILTER_COMMERCIAL_APPROVED_FINANCE_PENDING:
            records = records.filter(
                status=PaymentRecord.STATUS_APPROVED,
            ).filter(Q(finance_status__isnull=True) | Q(finance_status=''))
        elif filters['status'] in valid_statuses:
            records = records.filter(status=filters['status'])
    else:
        customer_status_map = {
            PaymentRecord.STATUS_PENDING: [
                PaymentRecord.STATUS_PENDING,
                PaymentRecord.STATUS_COMMERCIAL_REVIEW,
                PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL,
            ],
            PaymentRecord.STATUS_FINAL_APPROVED: [
                PaymentRecord.STATUS_APPROVED,
                PaymentRecord.STATUS_FINAL_APPROVED,
            ],
            PaymentRecord.STATUS_REJECTED: [PaymentRecord.STATUS_REJECTED],
            PaymentRecord.STATUS_INCOMPLETE: [PaymentRecord.STATUS_INCOMPLETE],
        }
        if filters['status'] in customer_status_map:
            records = records.filter(status__in=customer_status_map[filters['status']])

    # فیلتر سریال سند (ID)
    serial_f = filters.get('serial', '')
    if serial_f.isdigit():
        records = records.filter(id=int(serial_f))

    # فیلتر وضعیت طرف حساب
    cp_f = filters.get('cp_status', '')
    if cp_f == 'pending':
        records = records.filter(counterparty__isnull=False, counterparty_status__isnull=True)
    elif cp_f == 'approved':
        records = records.filter(counterparty_status=PaymentRecord.CP_STATUS_APPROVED)
    elif cp_f == 'returned':
        records = records.filter(counterparty_status=PaymentRecord.CP_STATUS_RETURNED)
    elif cp_f == 'rejected':
        records = records.filter(counterparty_status=PaymentRecord.CP_STATUS_REJECTED)
    elif cp_f == 'has_cp':
        records = records.filter(counterparty__isnull=False)

    # فیلتر دلیل رد فیش/سند
    valid_rejection_reasons = {choice[0] for choice in PaymentRecord.REJECTION_REASON_CHOICES}
    if filters['rejection_reason'] in valid_rejection_reasons:
        records = records.filter(rejection_reason=filters['rejection_reason'])

    return records, filters

def _format_thousand_separator(value):
    try:
        return '{:,}'.format(int(str(value).replace(',', '').strip()))
    except (ValueError, TypeError):
        return value


def _apply_record_sort(records, request):
    sortable_fields = {
        'payer_full_name': 'payer_full_name',
        'pay_date': 'pay_date',
        'tracking_code': 'tracking_code',
        'amount': 'amount',
        'payer_bank_name': 'payer_bank_name',
        'created_at': 'created_at',
        'status': 'status',
    }
    current_sort = (request.GET.get('sort') or '').strip()
    current_dir = (request.GET.get('dir') or 'desc').strip().lower()
    if current_dir not in {'asc', 'desc'}:
        current_dir = 'desc'

    sort_field = sortable_fields.get(current_sort)
    if sort_field:
        prefix = '' if current_dir == 'asc' else '-'
        records = records.order_by(f'{prefix}{sort_field}', '-id')
    else:
        records = records.order_by('-created_at', '-id')
        current_sort = ''
        current_dir = 'desc'

    query_params = request.GET.copy()
    query_params.pop('sort', None)
    query_params.pop('dir', None)
    base_query = urlencode(query_params, doseq=True)

    return records, current_sort, current_dir, base_query

def _check_duplicate_payment(form, user):
    """
    بررسی تکراری بودن فیش با فیلدهایی که مشتری پر کرده است.

    قوانین:
    ۱. اگر کد پیگیری وارد شده → به تنهایی کافی است (یکتاست)
    ۲. اگر کد پیگیری نبود → از ترکیب فیلدهای موجود استفاده می‌شود
    ۳. حداقل ۲ فیلد معنادار باید پر باشد تا بررسی انجام شود
    ۴. فیش‌های «رد شده» نادیده گرفته می‌شوند
    """
    amount           = form.cleaned_data.get('amount')
    tracking_code    = (form.cleaned_data.get('tracking_code') or '').strip()
    payer_account    = (form.cleaned_data.get('payer_account_number') or '').replace(' ', '').strip()
    beneficiary_acct = (form.cleaned_data.get('beneficiary_account_number') or '').replace(' ', '').strip()
    pay_date         = form.cleaned_data.get('pay_date')

    base_qs = PaymentRecord.objects.filter(
        user=user,
    ).exclude(
        status=PaymentRecord.STATUS_REJECTED,
    )

    duplicate = None

    # ── مسیر ۱: کد پیگیری وارد شده — به تنهایی یکتاست ─────────
    if tracking_code:
        qs = base_qs.filter(tracking_code=tracking_code)
        if amount:
            qs = qs.filter(amount=amount)
        duplicate = qs.first()

    # ── مسیر ۲: بدون کد پیگیری — ترکیب فیلدهای موجود ──────────
    if not duplicate and not tracking_code:
        filters = {}
        if amount:
            filters['amount'] = amount
        if payer_account:
            filters['payer_account_number'] = payer_account
        if beneficiary_acct:
            filters['beneficiary_account_number'] = beneficiary_acct
        if pay_date:
            filters['pay_date'] = pay_date

        # حداقل ۲ فیلد معنادار لازم است
        meaningful = [v for v in [amount, payer_account, beneficiary_acct] if v]
        if len(meaningful) >= 2 and filters:
            duplicate = base_qs.filter(**filters).first()

    if duplicate:
        form.add_error(
            None,
            f'⚠️ این فیش احتمالاً قبلاً ثبت شده است (سریال #{duplicate.id} — '
            f'وضعیت: {duplicate.get_status_display() if hasattr(duplicate, "get_status_display") else duplicate.status}). '
            f'پیش از ارسال مجدد، وضعیت فیش قبلی را بررسی کنید.'
        )


def _account_initial_data(user, profile, payment=None):
    payment = payment or PaymentRecord()
    return {
        'first_name': user.first_name or payment.first_name,
        'last_name': user.last_name or payment.last_name,
        'organization': (profile.organization if profile else '') or payment.organization,
        'city': (profile.city if profile else '') or payment.city,
        'phone': (profile.phone if profile else '') or payment.phone,
    }


def _save_receipts(payment, form):
    payload = form.receipt_payload()
    if not payload:
        return

    receipts = [
        PaymentReceipt(payment=payment, image=uploaded, file_hash=file_hash)
        for uploaded, file_hash in payload
    ]
    PaymentReceipt.objects.bulk_create(receipts)


def _source_profiles_for_user(user):
    if not user or not user.is_authenticated:
        return []
    records = (
        PaymentRecord.objects
        .filter(user=user)
        .values(
            'payer_account_number',
            'payer_full_name',
            'payer_bank_name',
        )
        .order_by('-id')
    )
    seen = set()
    profiles = []
    for row in records:
        values = {
            'payer_account_number': (row.get('payer_account_number') or '').strip(),
            'payer_full_name': (row.get('payer_full_name') or '').strip(),
            'payer_bank_name': (row.get('payer_bank_name') or '').strip(),
        }
        if not all(values.values()):
            continue
        if 'Z' in values.values():
            continue
        key = tuple(values[field] for field in (
            'payer_account_number',
            'payer_full_name',
            'payer_bank_name',
        ))
        if key in seen:
            continue
        seen.add(key)
        profiles.append(values)
    return profiles


def _destination_profiles_for_user(user):
    if not user or not user.is_authenticated:
        return []
    records = (
        PaymentRecord.objects
        .filter(user=user)
        .values(
            'beneficiary_bank_name',
            'beneficiary_account_number',
            'beneficiary_account_owner',
        )
        .order_by('-id')
    )
    seen = set()
    profiles = []
    for row in records:
        values = {
            'beneficiary_bank_name': (row.get('beneficiary_bank_name') or '').strip(),
            'beneficiary_account_number': (row.get('beneficiary_account_number') or '').strip(),
            'beneficiary_account_owner': (row.get('beneficiary_account_owner') or '').strip(),
        }
        if not all(values.values()):
            continue
        if 'Z' in values.values():
            continue
        key = tuple(values[field] for field in (
            'beneficiary_bank_name',
            'beneficiary_account_number',
            'beneficiary_account_owner',
        ))
        if key in seen:
            continue
        seen.add(key)
        profiles.append(values)
    return profiles


def _invoice_records_for_user(user):
    qs = InvoiceRecord.objects.select_related('customer', 'customer__profile', 'uploaded_by')
    if _is_staff_user(user):
        # Staff can see all invoices only if they have permission
        if not _can_view_invoices(user):
            return InvoiceRecord.objects.none()
        if _user_role(user) == 'sales':
            return _customer_limited_queryset_for_user(qs, user).order_by('-created_at', '-id')
        return qs.order_by('-created_at', '-id')
    # Customers can always see their own invoices
    return qs.filter(customer=user).order_by('-created_at', '-id')


def _price_lists_for_user(user):
    qs = PriceList.objects.select_related('customer', 'customer__profile', 'uploaded_by')
    if _can_view_price_list_history(user):
        if _user_role(user) == 'sales':
            return _customer_limited_queryset_for_user(qs, user).order_by('-created_at', '-id')
        return qs.order_by('-created_at', '-id')
    latest = qs.filter(customer=user).order_by('-created_at', '-id').first()
    if latest:
        return qs.filter(customer=user, batch_id=latest.batch_id).order_by('-created_at', '-id')
    return PriceList.objects.none()


def _proformas_for_user(user):
    qs = ProformaInvoice.objects.select_related('customer', 'customer__profile', 'issued_by', 'order')
    if _is_staff_user(user):
        if _user_role(user) == 'sales':
            return _customer_limited_queryset_for_user(qs, user).order_by('-created_at', '-id')
        return qs.order_by('-created_at', '-id')
    return qs.filter(customer=user).order_by('-created_at', '-id')


def _orders_for_user(user):
    qs = (
        CustomerOrder.objects
        .select_related('customer', 'customer__profile', 'sales_expert', 'requested_sales_expert')
        .prefetch_related('items', 'proformas')
        .order_by('-created_at', '-id')
    )
    if user.is_superuser or _user_role(user) in {'commercial', 'commercial_manager', 'finance', 'finance_manager', 'sales_manager'}:
        return qs
    if _user_role(user) == 'sales':
        # فقط سفارش‌های مشتریانی که به این کارشناس اختصاص دارند + سفارش‌هایی که مستقیم به او ارجاع شده
        return qs.filter(
            Q(sales_expert=user) | Q(customer__sales_assignment__sales_user=user)
        ).distinct()
    return qs.filter(customer=user)


def _apply_order_filters(records, request, is_staff_user=False):
    filters = {
        'q': (request.GET.get('q') or '').strip(),
        'status': (request.GET.get('status') or '').strip(),
        'customer': (request.GET.get('customer') or '').strip(),
        'sales': (request.GET.get('sales') or '').strip(),
        'city': (request.GET.get('city') or '').strip(),
        'province': (request.GET.get('province') or '').strip(),
    }
    if filters['q']:
        records = records.filter(
            Q(title__icontains=filters['q']) |
            Q(customer_note__icontains=filters['q']) |
            Q(staff_note__icontains=filters['q']) |
            Q(items__product_name__icontains=filters['q']) |
            Q(items__note__icontains=filters['q'])
        ).distinct()
    if filters['status']:
        records = records.filter(status=filters['status'])
    if is_staff_user and filters['customer']:
        records = records.filter(
            Q(customer__username__icontains=filters['customer']) |
            Q(customer__first_name__icontains=filters['customer']) |
            Q(customer__last_name__icontains=filters['customer']) |
            Q(customer__profile__organization__icontains=filters['customer'])
        )
    if is_staff_user and filters['sales']:
        records = records.filter(
            Q(sales_expert__username__icontains=filters['sales']) |
            Q(sales_expert__first_name__icontains=filters['sales']) |
            Q(sales_expert__last_name__icontains=filters['sales'])
        )
    if is_staff_user and filters['city']:
        records = records.filter(customer__profile__city__icontains=filters['city'])
    if is_staff_user and filters['province']:
        records = records.filter(customer__profile__province__icontains=filters['province'])
    return records, filters


def _sales_assignment_rows(request):
    profiles = _active_customer_profiles().order_by('organization', 'user__last_name', 'user__first_name', 'user__username')
    filters = {
        'q': (request.GET.get('q') or '').strip(),
        'sales': (request.GET.get('sales') or '').strip(),
        'city': (request.GET.get('city') or '').strip(),
        'province': (request.GET.get('province') or '').strip(),
        'status': (request.GET.get('status') or '').strip(),
    }
    if filters['q']:
        profiles = profiles.filter(
            Q(user__username__icontains=filters['q']) |
            Q(user__first_name__icontains=filters['q']) |
            Q(user__last_name__icontains=filters['q']) |
            Q(organization__icontains=filters['q']) |
            Q(phone__icontains=filters['q']) |
            Q(mobile__icontains=filters['q'])
        )
    if filters['city']:
        profiles = profiles.filter(city__icontains=filters['city'])
    if filters['province']:
        profiles = profiles.filter(province__icontains=filters['province'])

    assignments = {
        assignment.customer_id: assignment
        for assignment in CustomerSalesAssignment.objects.select_related('customer', 'sales_user', 'assigned_by')
    }
    open_order_counts = {
        row['customer']: row['count']
        for row in (
            CustomerOrder.objects
            .exclude(status__in=[CustomerOrder.STATUS_COMPLETED, CustomerOrder.STATUS_CANCELLED])
            .values('customer')
            .annotate(count=Count('id'))
        )
    }
    rows = []
    for profile in profiles:
        assignment = assignments.get(profile.user_id)
        sales_user = assignment.sales_user if assignment else None
        if filters['status'] == 'assigned' and not sales_user:
            continue
        if filters['status'] == 'unassigned' and sales_user:
            continue
        if filters['sales'] and not (
            sales_user and (
                filters['sales'].lower() in sales_user.username.lower()
                or filters['sales'].lower() in sales_user.get_full_name().lower()
            )
        ):
            continue
        rows.append({
            'customer': profile.user,
            'profile': profile,
            'assignment': assignment,
            'sales_user': sales_user,
            'assigned_by': assignment.assigned_by if assignment else None,
            'updated_at': assignment.updated_at if assignment else None,
            'note': assignment.note if assignment else '',
            'open_orders': open_order_counts.get(profile.user_id, 0),
        })
    return rows, filters


def _customer_home_summary(user):
    invoices = list(_invoice_records_for_user(user)[:5])
    payments = list(_records_for_user(user).filter(user=user).order_by('-created_at', '-id')[:5])
    payments = _enrich_records(payments)
    price_lists = list(_price_lists_for_user(user)[:3])
    proformas = list(_proformas_for_user(user)[:5])
    orders = list(_orders_for_user(user)[:5])
    daily_assignments = _customer_daily_assignments_for_user(user)[:5]
    return {
        'invoices': invoices,
        'payments': payments,
        'price_lists': price_lists,
        'proformas': proformas,
        'orders': orders,
        'daily_assignments': daily_assignments,
        'invoice_count': _invoice_records_for_user(user).count(),
        'payment_count': _records_for_user(user).filter(user=user).count(),
        'price_list_count': _price_lists_for_user(user).count(),
        'proforma_count': _proformas_for_user(user).count(),
        'order_count': _orders_for_user(user).count(),
        'daily_assignment_count': DailyPaymentAssignment.objects.filter(customer=user).count(),
    }


def _can_access_proforma(user, proforma):
    return (_is_staff_user(user) and _can_staff_access_customer(user, proforma.customer_id)) or proforma.customer_id == user.id


def _log_proforma(proforma, actor, action, note=''):
    ProformaInvoiceLog.objects.create(
        proforma=proforma,
        actor=actor if actor and actor.is_authenticated else None,
        action=action,
        note=note or '',
    )


def _delete_file_field(file_field):
    if file_field:
        try:
            file_field.delete(save=False)
        except OSError:
            logger.debug('Could not delete file %s from storage.', getattr(file_field, 'name', ''))


def _invoice_customer_rows():
    rows = []
    profiles = _active_customer_profiles().order_by(
        'user__first_name', 'user__last_name', 'user__username'
    )
    for profile in profiles:
        full_name = profile.user.get_full_name().strip() or profile.user.username
        rows.append({
            'profile_id': profile.id,
            'user_id': profile.user_id,
            'full_name': full_name,
            'username': profile.user.username,
            'organization': profile.organization or '-',
            'city': profile.city or '-',
            'phone': profile.phone or '-',
            'search_blob': ' '.join(
                filter(None, [full_name, profile.user.username, profile.organization, profile.city, profile.phone])
            ).lower(),
        })
    return rows


@login_required
def notifications_feed(request):
    notifications = UserNotification.objects.filter(user=request.user, is_read=False)[:10]
    return JsonResponse({
        'unread_count': UserNotification.objects.filter(user=request.user, is_read=False).count(),
        'items': [_notification_payload(notification) for notification in notifications],
    })


@login_required
def reconciliation_messages_feed(request):
    if not _can_access_reconciliation(request.user):
        return JsonResponse({'unread_count': 0, 'items': []})
    threads = _reconciliation_threads_for_user(request.user)[:5]
    items = []
    for thread in threads:
        last_message = thread.messages.exclude(sender_id=request.user.id).order_by('-created_at').first()
        if not last_message:
            continue
        items.append({
            'thread_id': thread.id,
            'title': thread.title,
            'message': last_message.body[:120],
            'url': f"{reverse('reconciliation_center')}?thread={thread.id}",
        })
    return JsonResponse({
        'unread_count': _reconciliation_unread_count(request.user),
        'items': items,
    })


_RECONCILIATION_INLINE_ATTACHMENT_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.pdf'}


@login_required
def reconciliation_attachment_download(request, message_id):
    message = get_object_or_404(ReconciliationMessage.objects.select_related('thread'), id=message_id)
    if not _can_access_reconciliation_thread(request.user, message.thread):
        return HttpResponseForbidden('شما به این پیوست دسترسی ندارید.')
    if not message.attachment:
        raise Http404
    name = message.attachment_name or message.attachment.name
    ext = os.path.splitext(name)[1].lower()
    as_attachment = request.GET.get('download') == '1' or ext not in _RECONCILIATION_INLINE_ATTACHMENT_EXTENSIONS
    return _file_response(message.attachment, as_attachment=as_attachment, filename=message.attachment_name or None)


@login_required
@require_POST
def notifications_mark_read(request):
    queryset = UserNotification.objects.filter(user=request.user, is_read=False)
    notification_id = request.POST.get('id')
    if notification_id:
        queryset = queryset.filter(id=notification_id)
    updated = queryset.update(is_read=True, read_at=timezone.now())
    unread_count = UserNotification.objects.filter(user=request.user, is_read=False).count()
    return JsonResponse({'ok': True, 'updated': updated, 'unread_count': unread_count})


def _customer_debt_summary(user):
    invoice_total = InvoiceRecord.objects.filter(customer=user).aggregate(total=Sum('amount'))['total'] or 0
    confirmed_payment_total = PaymentRecord.objects.filter(
        user=user,
        status__in=[
            PaymentRecord.STATUS_APPROVED,
            PaymentRecord.STATUS_FINAL_APPROVED,
        ],
    ).aggregate(total=Sum('amount'))['total'] or 0
    review_payment_total = PaymentRecord.objects.filter(user=user).exclude(
        status=PaymentRecord.STATUS_REJECTED,
    ).aggregate(total=Sum('amount'))['total'] or 0

    return {
        'invoice_total': invoice_total,
        'confirmed_payment_total': confirmed_payment_total,
        'review_payment_total': review_payment_total,
        'confirmed_debt': invoice_total - confirmed_payment_total,
        'review_debt': invoice_total - review_payment_total,
        'pending_review_payment_total': review_payment_total - confirmed_payment_total,
    }


def _can_manage_daily_payments(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return _user_role(user) in {'staff', 'commercial', 'commercial_manager', 'finance_manager'}


def _can_view_daily_payments(user):
    return _is_staff_user(user)


def _active_daily_assignment_for_user(user):
    if not user or not user.is_authenticated:
        return None
    today = _today_jalali_date()
    return (
        DailyPaymentAssignment.objects
        .select_related('plan', 'customer')
        .filter(customer=user, plan__deposit_date=today)
        .order_by('-id')
        .first()
    )


def _latest_expired_daily_assignment_for_user(user):
    if not user or not user.is_authenticated:
        return None
    today = _today_jalali_date()
    return (
        DailyPaymentAssignment.objects
        .select_related('plan', 'customer')
        .filter(customer=user, plan__deposit_date__lt=today)
        .order_by('-plan__deposit_date', '-id')
        .first()
    )


def _daily_assignment_stats(assignments):
    assignment_ids = [assignment.id for assignment in assignments]
    if not assignment_ids:
        return {}

    paid_rows = (
        PaymentRecord.objects
        .filter(daily_assignment_id__in=assignment_ids)
        .exclude(status=PaymentRecord.STATUS_REJECTED)
        .values('daily_assignment')
        .annotate(total=Sum('amount'), count=Count('id'), latest_payment=Max('created_at'))
    )
    confirmed_rows = (
        PaymentRecord.objects
        .filter(
            daily_assignment_id__in=assignment_ids,
            status__in=[
                PaymentRecord.STATUS_APPROVED,
                PaymentRecord.STATUS_FINAL_APPROVED,
            ],
        )
        .values('daily_assignment')
        .annotate(total=Sum('amount'), count=Count('id'))
    )
    stats = {
        assignment_id: {
            'paid_amount': 0,
            'payment_count': 0,
            'confirmed_amount': 0,
            'confirmed_count': 0,
            'latest_payment': None,
        }
        for assignment_id in assignment_ids
    }
    for row in paid_rows:
        data = stats[row['daily_assignment']]
        data['paid_amount'] = row['total'] or 0
        data['payment_count'] = row['count'] or 0
        data['latest_payment'] = row['latest_payment']
    for row in confirmed_rows:
        data = stats[row['daily_assignment']]
        data['confirmed_amount'] = row['total'] or 0
        data['confirmed_count'] = row['count'] or 0
    return stats


def _managed_users(query='', role='', status=''):
    users = User.objects.select_related('profile').order_by('username')
    if query:
        users = users.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query) |
            Q(profile__phone__icontains=query) |
            Q(profile__mobile__icontains=query) |
            Q(profile__representative_name__icontains=query) |
            Q(profile__representative_mobile__icontains=query) |
            Q(profile__organization__icontains=query) |
            Q(profile__city__icontains=query) |
            Q(profile__province__icontains=query)
        )
    if role:
        users = users.filter(profile__role=role)
    if status == 'active':
        users = users.filter(is_active=True)
    elif status == 'inactive':
        users = users.filter(is_active=False)
    elif status == 'suspended':
        users = users.filter(profile__suspended=True)
    return users


def _apply_invoice_filters(records, request, is_staff_user):
    filters = {
        'customer': (request.GET.get('customer') or '').strip(),
        'invoice_number': (request.GET.get('invoice_number') or '').strip(),
        'reference_number': (request.GET.get('reference_number') or '').strip(),
        'amount': (request.GET.get('amount') or '').replace(',', '').strip(),
        'invoice_date': (request.GET.get('invoice_date') or '').strip(),
        'invoice_date_from': (request.GET.get('invoice_date_from') or '').strip(),
        'invoice_date_to': (request.GET.get('invoice_date_to') or '').strip(),
        'seen': (request.GET.get('seen') or '').strip(),
    }

    if is_staff_user and filters['customer']:
        records = records.filter(
            Q(customer__first_name__icontains=filters['customer']) |
            Q(customer__last_name__icontains=filters['customer']) |
            Q(customer__username__icontains=filters['customer']) |
            Q(customer__profile__organization__icontains=filters['customer'])
        )

    if filters['invoice_number']:
        records = records.filter(invoice_number__icontains=filters['invoice_number'])

    if filters['reference_number']:
        records = records.filter(reference_number__icontains=filters['reference_number'])

    if filters['amount'].isdigit():
        parsed_amount = int(filters['amount'])
        records = records.filter(amount=parsed_amount)
        filters['amount'] = _format_thousand_separator(parsed_amount)

    parsed_date = _parse_jalali_date(filters['invoice_date'])
    if parsed_date:
        records = records.filter(invoice_date=parsed_date)
    else:
        parsed_date_from = _parse_jalali_date(filters['invoice_date_from'])
        parsed_date_to = _parse_jalali_date(filters['invoice_date_to'])
        if parsed_date_from and parsed_date_to and parsed_date_to < parsed_date_from:
            parsed_date_from, parsed_date_to = parsed_date_to, parsed_date_from
            filters['invoice_date_from'], filters['invoice_date_to'] = filters['invoice_date_to'], filters['invoice_date_from']
        if parsed_date_from:
            records = records.filter(invoice_date__gte=parsed_date_from)
        if parsed_date_to:
            records = records.filter(invoice_date__lte=parsed_date_to)

    if filters['seen'] == 'seen':
        records = records.filter(customer_seen_at__isnull=False)
    elif filters['seen'] == 'unseen':
        records = records.filter(customer_seen_at__isnull=True)

    return records, filters


@login_required
def daily_payment_plans(request):
    if not _can_view_daily_payments(request.user):
        return HttpResponseForbidden('این بخش فقط برای کاربران واحدهای شرکت قابل دسترسی است.')

    can_manage = _can_manage_daily_payments(request.user)
    period = _daily_payment_period(request)

    if request.method == 'POST':
        if not can_manage:
            return HttpResponseForbidden('شما دسترسی ایجاد برنامه واریز روزانه را ندارید.')
        plan_form = DailyPaymentPlanForm(request.POST)
        if plan_form.is_valid():
            plan = plan_form.save(commit=False)
            plan.created_by = request.user
            plan.save()
            messages.success(request, 'برنامه واریز روزانه ثبت شد.')
            detail_url = f"{reverse('daily_payment_plan_detail', kwargs={'plan_id': plan.id})}?{urlencode({'next': request.get_full_path()})}"
            return redirect(detail_url)
    else:
        plan_form = DailyPaymentPlanForm(initial={'deposit_date': _today_jalali_date()})

    plans = _daily_plans_for_period(period['start_date'], period['end_date'])

    return render(request, 'payments/daily_payment_plans.html', {
        'form': plan_form,
        'plans': plans,
        'selected_date': period['selected_date'],
        'selected_date_text': _format_jalali_date(period['selected_date']),
        'period_mode': period['mode'],
        'period_label': period['label'],
        'period_start_text': _format_jalali_date(period['start_date']),
        'period_end_text': _format_jalali_date(period['end_date']),
        'previous_period_query': _daily_period_query(
            period['mode'],
            period['previous_date'],
            period['previous_start_date'],
            period['previous_end_date'],
        ),
        'next_period_query': _daily_period_query(
            period['mode'],
            period['next_date'],
            period['next_start_date'],
            period['next_end_date'],
        ),
        'today_period_query': _daily_period_query(period['mode'], _today_jalali_date()),
        'can_manage_daily_payments': can_manage,
        'user_display_name': f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username,
        'export_dataset': 'daily_plans',
        'export_fields': DAILY_PLAN_EXPORT_FIELDS,
    })


@login_required
def daily_payment_plan_detail(request, plan_id):
    if not _can_view_daily_payments(request.user):
        return HttpResponseForbidden('این بخش فقط برای کاربران واحدهای شرکت قابل دسترسی است.')

    plan = get_object_or_404(DailyPaymentPlan.objects.select_related('created_by'), id=plan_id)
    _mark_notifications_read_for_url(request.user, request.path)
    can_manage = _can_manage_daily_payments(request.user)
    return_url = _safe_next_url(request, default=f"{reverse('daily_payment_plans')}?date={_format_jalali_date(plan.deposit_date)}")
    return_label = _return_link_label(request, 'بازگشت به برنامه ها')
    detail_url = f"{request.path}?{urlencode({'next': return_url})}"

    if request.method == 'POST':
        if not can_manage:
            return HttpResponseForbidden('شما دسترسی ویرایش تخصیص واریز روزانه را ندارید.')
        if request.POST.get('action') == 'remove_assignment':
            assignment_id = request.POST.get('assignment_id')
            assignment = get_object_or_404(DailyPaymentAssignment, id=assignment_id, plan=plan)
            assignment.delete()
            messages.success(request, 'تخصیص مشتری حذف شد.')
            return redirect(detail_url)

        assignment_form = DailyPaymentAssignmentForm(request.POST)
        if assignment_form.is_valid():
            created_count = 0
            duplicate_count = 0
            for customer in assignment_form.cleaned_data['customers']:
                assignment = DailyPaymentAssignment(
                    plan=plan,
                    customer=customer,
                    expected_amount=assignment_form.cleaned_data['expected_amount'],
                    note=assignment_form.cleaned_data.get('note') or '',
                )
                try:
                    assignment.save()
                except IntegrityError:
                    duplicate_count += 1
                else:
                    created_count += 1
            if created_count:
                messages.success(request, f'تخصیص برای {created_count} مشتری ثبت شد.')
                return redirect(detail_url)
            assignment_form.add_error('customers', 'برای مشتریان انتخاب شده در این برنامه قبلاً تخصیص ثبت شده است.')
    else:
        assignment_form = DailyPaymentAssignmentForm()

    assignments = _daily_assignments_for_plan(plan)

    totals = {
        'expected': sum(assignment.expected_amount for assignment in assignments),
        'paid': sum(assignment.report['paid_amount'] for assignment in assignments),
        'confirmed': sum(assignment.report['confirmed_amount'] for assignment in assignments),
        'payment_count': sum(assignment.report['payment_count'] for assignment in assignments),
        'no_payment_count': sum(1 for assignment in assignments if assignment.report['paid_amount'] <= 0),
        'partial_count': sum(1 for assignment in assignments if 0 < assignment.report['paid_amount'] < assignment.expected_amount),
        'complete_count': sum(1 for assignment in assignments if assignment.expected_amount > 0 and assignment.report['paid_amount'] >= assignment.expected_amount),
    }
    totals['remaining'] = totals['expected'] - totals['paid']
    totals['confirmed_remaining'] = totals['expected'] - totals['confirmed']

    return render(request, 'payments/daily_payment_plan_detail.html', {
        'plan': plan,
        'assignments': assignments,
        'assignment_form': assignment_form,
        'totals': totals,
        'can_manage_daily_payments': can_manage,
        'user_display_name': f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username,
        'return_url': return_url,
        'return_label': return_label,
        'export_dataset': 'daily_assignments',
        'export_fields': DAILY_ASSIGNMENT_EXPORT_FIELDS,
        'export_extra_params': {'plan_id': plan.id},
    })


@login_required
def customer_daily_payments(request):
    if _is_staff_user(request.user):
        return redirect('daily_payment_plans')

    assignments = _customer_daily_assignments_for_user(request.user, request=request)
    status_filter = (request.GET.get('status') or '').strip()
    if status_filter:
        assignments = [
            assignment for assignment in assignments
            if (
                (status_filter == 'none' and assignment.report['paid_amount'] <= 0) or
                (status_filter == 'partial' and 0 < assignment.report['paid_amount'] < assignment.expected_amount) or
                (status_filter == 'complete' and assignment.expected_amount > 0 and assignment.report['paid_amount'] >= assignment.expected_amount)
            )
        ]

    page_obj = _paginate_queryset(request, assignments, per_page=10, page_param='page')
    page_base_query = _build_query_string(request, remove_keys=['page'])
    filters = {
        'start_date': (request.GET.get('start_date') or '').strip(),
        'end_date': (request.GET.get('end_date') or '').strip(),
        'status': status_filter,
    }
    return render(request, 'payments/customer_daily_payments.html', {
        'assignments': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'filters': filters,
        'export_dataset': 'customer_daily_assignments',
        'export_fields': DAILY_ASSIGNMENT_EXPORT_FIELDS,
        'user_display_name': request.user.get_full_name().strip() or request.user.username,
    })


@login_required
def create_payment(request):
    profile = None
    try:
        profile = request.user.profile
    except UserProfile.DoesNotExist:
        profile = None

    initial_data = _account_initial_data(request.user, profile)
    is_staff_user = _is_staff_user(request.user)
    show_payment_form = bool(request.resolver_match and request.resolver_match.url_name == 'payment_create')
    staff_role = _user_role(request.user) if is_staff_user else ''
    is_system_admin = request.user.is_superuser
    active_daily_assignment = _active_daily_assignment_for_user(request.user) if not is_staff_user else None
    expired_daily_assignment = None
    if not is_staff_user and not active_daily_assignment:
        expired_daily_assignment = _latest_expired_daily_assignment_for_user(request.user)
    form_initial = initial_data.copy()
    if active_daily_assignment and request.method != 'POST':
        form_initial.update({
            'beneficiary_bank_name': active_daily_assignment.plan.bank_name,
            'beneficiary_account_number': active_daily_assignment.plan.account_number,
            'beneficiary_account_owner': active_daily_assignment.plan.account_owner,
            'pay_date': active_daily_assignment.plan.deposit_date,
        })

    if request.method == 'POST':
        if not show_payment_form:
            return redirect('payment_create')
        if is_staff_user:
            return HttpResponseForbidden('کاربران واحدها امکان ثبت سند از این فرم را ندارند.')
        form = PaymentRecordForm(request.POST, request.FILES, initial=form_initial)
        valid = form.is_valid()
        if not valid:
            try:
                log_path = os.path.join(getattr(settings, 'BASE_DIR', '.'), 'payment_submit_debug.log')
                with open(log_path, 'a', encoding='utf-8') as fh:
                    fh.write(f"USER:{request.user.username} VALID={valid} ERRORS:{getattr(form.errors, 'as_json', lambda: str(form.errors))()} FILES:[")
                    first = True
                    for key in request.FILES:
                        for f in request.FILES.getlist(key):
                            if not first:
                                fh.write(',')
                            fh.write(f"{key}:{getattr(f, 'name', '')}:{getattr(f, 'size', 0)}")
                            first = False
                    fh.write("]\n")
            except Exception:
                logger.exception('Failed to write payment_submit_debug.log')

        if valid:
            submitted_account = (form.cleaned_data.get('beneficiary_account_number') or '').replace(' ', '').strip()
            if active_daily_assignment:
                expected_date = active_daily_assignment.plan.deposit_date
                submitted_date = form.cleaned_data.get('pay_date')
                if submitted_date and submitted_date != expected_date:
                    form.add_error('pay_date', 'این شماره حساب فقط برای تاریخ اعلام شده امروز معتبر است.')
                expected_account = (active_daily_assignment.plan.account_number or '').replace(' ', '').strip()
                if submitted_account and submitted_account != expected_account:
                    form.add_error('beneficiary_account_number', 'شماره حساب مقصد باید همان شماره حساب اعلام شده امروز باشد.')
            elif submitted_account:
                assigned_accounts = (
                    DailyPaymentAssignment.objects
                    .select_related('plan')
                    .filter(customer=request.user, plan__account_number__iexact=form.cleaned_data.get('beneficiary_account_number'))
                )
                if assigned_accounts.exists():
                    form.add_error('beneficiary_account_number', 'این شماره حساب برای امروز معتبر نیست و امکان ثبت فیش با آن وجود ندارد.')

            # ── بررسی تکراری بودن فیش ──────────────────────────────
            if not form.errors:
                _check_duplicate_payment(form, request.user)

            if not form.errors:
                payment = form.save(commit=False)
                payment.user = request.user
                payment.first_name = initial_data['first_name']
                payment.last_name = initial_data['last_name']
                payment.organization = initial_data['organization']
                payment.city = initial_data['city']
                payment.phone = initial_data['phone']
                payment.status = PaymentRecord.STATUS_PENDING
                payment.daily_assignment = active_daily_assignment
                payment.save()
                _save_receipts(payment, form)
                _log_activity(payment, request.user, PaymentActivityLog.ACTION_CREATED, to_status=payment.status)
                _notify_payment_created(payment, request.user)
                return redirect('success')
    else:
        form = PaymentRecordForm(initial=form_initial)

    records = _active_payment_records_for_user(request.user)
    _mark_commercial_records_seen(records, request.user)
    records = _active_payment_records_for_user(request.user)
    records, active_filters = _apply_record_filters(records, request, is_staff_user)
    records, current_sort, current_sort_dir, sort_base_query = _apply_record_sort(records, request)
    page_obj = _paginate_queryset(request, records, per_page=10, page_param='page')
    page_obj.object_list = _enrich_records(
        page_obj.object_list,
        staff_role=staff_role,
        is_system_admin=is_system_admin,
        can_edit_payment_details=_can_edit_payment_details(request.user),
        acting_user=request.user,
    )
    page_base_query = _build_query_string(request, remove_keys=['page'])
    user_display_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username

    return render(request, 'payments/form.html', {
        'form': form,
        'records': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'is_staff_user': is_staff_user,
        'filters': active_filters,
        'status_choices': PaymentRecord.STAFF_FILTER_CHOICES if is_staff_user else CUSTOMER_STATUSES,
        'counterparties': Counterparty.objects.all() if is_staff_user else [],
        'staff_user_role': staff_role,
        'staff_role_label': _staff_role_label(staff_role),
        'can_manage_counterparties': is_system_admin,
        'can_export_records': (not is_staff_user) or is_system_admin or staff_role in {'finance', 'finance_manager', 'commercial', 'commercial_manager'},
        'is_system_admin': is_system_admin,
        'user_display_name': user_display_name,
        'source_profiles': _source_profiles_for_user(request.user) if not is_staff_user else [],
        'destination_profiles': _destination_profiles_for_user(request.user) if not is_staff_user else [],
        'current_sort': current_sort,
        'current_sort_dir': current_sort_dir,
        'sort_base_query': sort_base_query,
        'is_history_mode': False,
        'records_url_name': 'payment_create' if show_payment_form else 'submit',
        'export_dataset': 'payments',
        'export_fields': PAYMENT_EXPORT_FIELDS,
        'customer_info': initial_data,
        'customer_debt': _customer_debt_summary(request.user) if not is_staff_user else None,
        'customer_home_summary': _customer_home_summary(request.user) if not is_staff_user else None,
        'active_daily_assignment': active_daily_assignment,
        'expired_daily_assignment': expired_daily_assignment,
        'show_payment_form': show_payment_form,
        'finance_users': list(User.objects.filter(profile__role__in=['finance', 'finance_manager'], is_active=True).select_related('profile')) if is_staff_user else [],
        'can_bulk_final_approve': _can_delegate_final_approval(staff_role, is_system_admin) or _can_final_approve(staff_role, PaymentRecord(), is_system_admin, user=request.user),
        'cp_returned_count': PaymentRecord.objects.filter(
            counterparty_status=PaymentRecord.CP_STATUS_RETURNED
        ).count() if is_staff_user and _user_role(request.user) in {'commercial', 'commercial_manager'} else 0,
        'rejection_reason_choices': PaymentRecord.REJECTION_REASON_CHOICES,
    })


@login_required
def payment_history(request):
    if not _is_staff_user(request.user):
        return HttpResponseForbidden('این بخش فقط برای پرسنل واحدها قابل دسترسی است.')

    staff_role = _user_role(request.user)
    is_system_admin = request.user.is_superuser
    records = _history_payment_records_for_user(request.user)
    records, active_filters = _apply_record_filters(records, request, True)
    records, current_sort, current_sort_dir, sort_base_query = _apply_record_sort(records, request)
    page_obj = _paginate_queryset(request, records, per_page=10, page_param='page')
    page_obj.object_list = _enrich_records(
        page_obj.object_list,
        staff_role=staff_role,
        is_system_admin=is_system_admin,
        can_edit_payment_details=_can_edit_payment_details(request.user),
        acting_user=request.user,
    )
    page_base_query = _build_query_string(request, remove_keys=['page'])
    user_display_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username

    return render(request, 'payments/form.html', {
        'form': PaymentRecordForm(),
        'records': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'is_staff_user': True,
        'filters': active_filters,
        'status_choices': PaymentRecord.STAFF_FILTER_CHOICES,
        'counterparties': Counterparty.objects.all(),
        'staff_user_role': staff_role,
        'staff_role_label': _staff_role_label(staff_role),
        'can_manage_counterparties': is_system_admin,
        'can_export_records': is_system_admin or staff_role in {'finance', 'finance_manager', 'commercial', 'commercial_manager'},
        'is_system_admin': is_system_admin,
        'user_display_name': user_display_name,
        'source_profiles': [],
        'destination_profiles': [],
        'current_sort': current_sort,
        'current_sort_dir': current_sort_dir,
        'sort_base_query': sort_base_query,
        'is_history_mode': True,
        'records_url_name': 'payment_history',
        'export_dataset': 'payment_history',
        'export_fields': PAYMENT_EXPORT_FIELDS,
        'customer_info': {},
        'customer_debt': None,
        'active_daily_assignment': None,
        'expired_daily_assignment': None,
        'rejection_reason_choices': PaymentRecord.REJECTION_REASON_CHOICES,
    })


@login_required
def pending_final_approval_queue(request):
    """صف «در انتظار تأیید نهایی» — فقط مدیر مالی و کاربران تفویض‌شده"""
    if not _can_see_pending_final_approval(request.user):
        return HttpResponseForbidden('دسترسی به این بخش فقط برای مدیر مالی و کاربران تفویض‌شده مجاز است.')

    staff_role = _user_role(request.user)
    is_system_admin = request.user.is_superuser

    records = (
        PaymentRecord.objects
        .filter(pending_final_approval=True)
        .select_related('counterparty', 'user', 'user__profile')
        .prefetch_related(
            'receipts',
            Prefetch('activity_logs', queryset=PaymentActivityLog.objects.select_related('actor', 'actor__profile')),
        )
        .order_by('pending_final_approval_since', 'id')
    )
    records, active_filters = _apply_record_filters(records, request, True)
    records, current_sort, current_sort_dir, sort_base_query = _apply_record_sort(records, request)
    page_obj = _paginate_queryset(request, records, per_page=10, page_param='page')
    page_obj.object_list = _enrich_records(
        page_obj.object_list,
        staff_role=staff_role,
        is_system_admin=is_system_admin,
        can_edit_payment_details=_can_edit_payment_details(request.user),
        acting_user=request.user,
    )
    page_base_query = _build_query_string(request, remove_keys=['page'])
    user_display_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username

    return render(request, 'payments/form.html', {
        'form': PaymentRecordForm(),
        'records': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'is_staff_user': True,
        'filters': active_filters,
        'status_choices': PaymentRecord.STAFF_FILTER_CHOICES,
        'counterparties': Counterparty.objects.all(),
        'staff_user_role': staff_role,
        'staff_role_label': _staff_role_label(staff_role),
        'can_manage_counterparties': is_system_admin,
        'can_export_records': is_system_admin or staff_role in {'finance', 'finance_manager'},
        'is_system_admin': is_system_admin,
        'user_display_name': user_display_name,
        'source_profiles': [],
        'destination_profiles': [],
        'current_sort': current_sort,
        'current_sort_dir': current_sort_dir,
        'sort_base_query': sort_base_query,
        'is_history_mode': False,
        'is_pending_final_mode': True,
        'records_url_name': 'pending_final_approval',
        'export_dataset': 'payments',
        'export_fields': PAYMENT_EXPORT_FIELDS,
        'customer_info': {},
        'customer_debt': None,
        'active_daily_assignment': None,
        'expired_daily_assignment': None,
        'show_payment_form': False,
        'finance_users': list(User.objects.filter(
            profile__role__in=['finance', 'finance_manager'], is_active=True
        ).select_related('profile')),
        'can_bulk_final_approve': (
            _can_delegate_final_approval(staff_role, is_system_admin) or
            _can_final_approve(staff_role, PaymentRecord(), is_system_admin, user=request.user)
        ),
        'cp_returned_count': 0,
        'rejection_reason_choices': PaymentRecord.REJECTION_REASON_CHOICES,
    })


@login_required
def success(request):
    records = _records_for_user(request.user)
    return render(request, 'payments/success.html', {'records': records})


@login_required
def profile_password_change(request):
    profile = getattr(request.user, 'profile', None)
    is_force_change = bool(profile and profile.force_password_change)
    show_initial_password_change_note = bool(
        is_force_change and request.session.get('show_initial_password_change_note')
    )

    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            if profile and profile.force_password_change:
                profile.force_password_change = False
                profile.save(update_fields=['force_password_change'])
            request.session.pop('show_initial_password_change_note', None)
            messages.success(request, 'رمز عبور با موفقیت تغییر کرد.')
            return redirect('submit')
        messages.error(request, 'تغییر رمز انجام نشد. لطفا خطاها را بررسی کنید.')
    else:
        form = CustomPasswordChangeForm(request.user)

    return render(request, 'payments/profile_password_change.html', {
        'form': form,
        'is_force_change': is_force_change,
        'show_initial_password_change_note': show_initial_password_change_note,
        'username': request.user.username,
    })


def _log_system_activity(actor, target_user, action, description=''):
    SystemActivityLog.objects.create(
        actor=actor if actor and actor.is_authenticated else None,
        target_user=target_user,
        action=action,
        description=description,
    )


def _can_review_profile_change(user):
    if user.is_superuser:
        return True
    return _user_role(user) in STAFF_ROLES


def _user_card_payload(user):
    profile = getattr(user, 'profile', None)
    organization = getattr(profile, 'organization', '') or ''
    counterparty = getattr(user, 'counterparty_account', None)
    if counterparty and not organization:
        organization = counterparty.name
    location = ' / '.join(
        part for part in [
            (getattr(profile, 'city', '') or '').strip(),
            (getattr(profile, 'province', '') or '').strip(),
        ]
        if part
    )
    role_label = getattr(profile, 'get_role_display', lambda: 'کاربر')()
    if counterparty:
        role_label = f'طرف حساب: {counterparty.name}'
    return {
        'display_name': profile.display_name if profile else (user.get_full_name().strip() or user.username),
        'username': user.username,
        'role_label': role_label,
        'organization': organization or '-',
        'mobile': getattr(profile, 'mobile', '') or '-',
        'phone': getattr(profile, 'phone', '') or '-',
        'email': user.email or '-',
        'location': location or '-',
        'representative_name': getattr(profile, 'representative_name', '') or '-',
        'representative_mobile': getattr(profile, 'representative_mobile', '') or '-',
        'delegate_sms_to_representative': bool(getattr(profile, 'delegate_sms_to_representative', False)),
        'avatar_url': profile.avatar_url if profile else '',
        'avatar_icon': profile.avatar_icon if profile else '👤',
        'avatar_class': profile.avatar_class if profile else 'avatar-neutral_1',
    }


@login_required
def user_business_card(request, user_id):
    target_user = get_object_or_404(User.objects.select_related('profile'), id=user_id, is_active=True)
    return JsonResponse(_user_card_payload(target_user))


@login_required
def profile_edit(request):
    profile = get_object_or_404(UserProfile, user=request.user)
    pending_change = (
        ProfileChangeRequest.objects
        .filter(user=request.user, status=ProfileChangeRequest.STATUS_PENDING)
        .order_by('-created_at', '-id')
        .first()
    )

    if request.method == 'POST':
        form = CustomerProfileUpdateForm(request.POST, request.FILES, instance=profile, user=request.user)
        if form.is_valid():
            avatar_changed = form.save_avatar_fields()
            changes = form.changed_profile_fields()
            if changes:
                if pending_change:
                    pending_change.changes = form.changes_payload()
                    pending_change.requested_by = request.user
                    pending_change.save(update_fields=['changes', 'requested_by'])
                else:
                    pending_change = ProfileChangeRequest.objects.create(
                        user=request.user,
                        requested_by=request.user,
                        changes=form.changes_payload(),
                    )
                change_text = '؛ '.join(
                    f"{item['field']}: از «{item['old']}» به «{item['new']}»"
                    for item in changes
                )
                _log_system_activity(
                    request.user,
                    request.user,
                    SystemActivityLog.ACTION_PROFILE_UPDATED,
                    f'درخواست تغییر مشخصات کاربر ثبت شد و در انتظار تایید است. {change_text}',
                )
                messages.success(request, 'درخواست تغییر مشخصات ثبت شد و تا زمان تایید، به صورت تایید نشده نمایش داده می‌شود.')
            elif avatar_changed:
                messages.success(request, 'نمایه کاربری با موفقیت به‌روزرسانی شد.')
            else:
                messages.info(request, 'تغییری برای ثبت وجود نداشت.')
            return redirect('profile_edit')
        messages.error(request, 'ذخیره مشخصات انجام نشد. لطفا خطاها را بررسی کنید.')
    else:
        form = CustomerProfileUpdateForm(instance=profile, user=request.user)

    return render(request, 'payments/profile_edit.html', {
        'form': form,
        'username': request.user.username,
        'pending_change': pending_change,
    })


@login_required
@require_POST
def profile_change_request_review(request, request_id):
    if not _can_review_profile_change(request.user):
        return HttpResponseForbidden('شما دسترسی تایید تغییرات مشخصات را ندارید.')
    change_request = get_object_or_404(
        ProfileChangeRequest.objects.select_related('user', 'user__profile', 'requested_by'),
        id=request_id,
        status=ProfileChangeRequest.STATUS_PENDING,
    )
    action = request.POST.get('action')
    if action == 'approve':
        change_request.apply_changes(request.user)
        change_text = '؛ '.join(
            f"{item['label']}: از «{item['old']}» به «{item['new']}»"
            for item in change_request.change_items
        )
        _log_system_activity(
            request.user,
            change_request.user,
            SystemActivityLog.ACTION_PROFILE_UPDATED,
            f'درخواست تغییر مشخصات تایید شد. {change_text}',
        )
        messages.success(request, 'درخواست تغییر مشخصات تایید و روی پروفایل اعمال شد.')
    elif action == 'reject':
        change_request.status = ProfileChangeRequest.STATUS_REJECTED
        change_request.reviewed_by = request.user
        change_request.reviewed_at = timezone.now()
        change_request.review_note = (request.POST.get('review_note') or '').strip()
        change_request.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'review_note'])
        _log_system_activity(
            request.user,
            change_request.user,
            SystemActivityLog.ACTION_PROFILE_UPDATED,
            'درخواست تغییر مشخصات رد شد.',
        )
        messages.warning(request, 'درخواست تغییر مشخصات رد شد.')
    else:
        messages.error(request, 'عملیات نامعتبر است.')
    return redirect(request.POST.get('next') or 'users_manage')


def _send_temporary_password_email(user, temp_password):
    if not user.email:
        return False, 'برای این کاربر ایمیل ثبت نشده است.'

    subject = 'رمز عبور جدید سامانه'
    message = (
        f'{user.get_full_name() or user.username} عزیز،\n\n'
        f'رمز عبور جدید شما در سامانه: {temp_password}\n\n'
        'پس از ورود، لازم است رمز عبور خود را تغییر دهید.'
    )
    send_mail(
        subject,
        message,
        settings.DEFAULT_FROM_EMAIL,
        [user.email],
        fail_silently=False,
    )
    return True, ''


@login_required
def profile_password_cancel(request):
    profile = getattr(request.user, 'profile', None)
    is_force_change = bool(profile and profile.force_password_change)
    request.session.pop('show_initial_password_change_note', None)
    if is_force_change:
        auth_logout(request)
        return redirect('login')
    return redirect('submit')


@login_required
@require_POST
def finance_bulk_final_approve(request):
    """تأیید نهایی گروهی — مدیر مالی یا تفویض‌شده."""
    redirect_target = _safe_next_url(request, default=reverse('submit'))
    role = _user_role(request.user)
    if not _can_delegate_final_approval(role, request.user.is_superuser) and \
       not _can_final_approve(role, PaymentRecord(), request.user.is_superuser, user=request.user):
        # بررسی اینکه آیا کاربر اصلاً مجاز به تأیید نهایی است
        from .models import FinalApprovalDelegate
        if not FinalApprovalDelegate.objects.filter(delegated_user=request.user, is_active=True).exists():
            messages.error(request, 'شما مجاز به تأیید نهایی نیستید.')
            return redirect(redirect_target)

    payment_ids = request.POST.getlist('payment_ids')
    note = (request.POST.get('note') or '').strip()

    if not payment_ids:
        messages.error(request, 'هیچ سندی انتخاب نشده است.')
        return redirect(redirect_target)

    approved = 0
    skipped = 0
    for pid in payment_ids:
        try:
            payment = PaymentRecord.objects.get(id=int(pid))
            if not payment.ready_for_final_approval:
                skipped += 1
                continue
            old_status = payment.status
            payment.status = PaymentRecord.STATUS_FINAL_APPROVED
            payment.pending_final_approval = False
            payment.pending_final_approval_since = None
            payment.save(update_fields=['status', 'pending_final_approval', 'pending_final_approval_since'])
            _log_activity(payment, request.user, PaymentActivityLog.ACTION_FINAL_APPROVED,
                          from_status=old_status, to_status=payment.status,
                          note=note or 'تأیید نهایی گروهی')
            if payment.user:
                customer_name = f"{payment.first_name} {payment.last_name}".strip() or payment.user.username
                _notify_users([payment.user], 'تأیید نهایی سند',
                              f'سند #{pid} مشتری {customer_name} تأیید نهایی شد.',
                              reverse('submit'), category=UserNotification.CATEGORY_SYSTEM,
                              actor=request.user)
            approved += 1
        except (PaymentRecord.DoesNotExist, ValueError):
            skipped += 1

    if approved:
        messages.success(request, f'✅ {approved} سند تأیید نهایی شد.' + (f' ({skipped} سند رد شد)' if skipped else ''))
    else:
        messages.error(request, f'هیچ سندی تأیید نشد. ({skipped} سند آماده نبود)')
    return redirect(redirect_target)


@login_required
def final_approval_delegation_page(request):
    """صفحه مستقل مدیریت تفویض اختیار تأیید نهایی."""
    role = _user_role(request.user)
    if not _can_delegate_final_approval(role, request.user.is_superuser):
        return HttpResponseForbidden('فقط مدیر مالی می‌تواند تفویض اختیار انجام دهد.')

    from .models import FinalApprovalDelegate
    if request.method == 'POST':
        action = request.POST.get('action')
        user_id = (request.POST.get('user_id') or '').strip()
        note = (request.POST.get('note') or '').strip()

        if action == 'add' and user_id.isdigit():
            target_user = get_object_or_404(User, id=int(user_id), is_active=True)
            obj, created = FinalApprovalDelegate.objects.get_or_create(
                delegated_user=target_user,
                defaults={'granted_by': request.user, 'note': note, 'is_active': True},
            )
            if not created:
                obj.is_active = True
                obj.granted_by = request.user
                obj.note = note
                obj.save(update_fields=['is_active', 'granted_by', 'note', 'updated_at'])
            messages.success(request, f'اختیار تأیید نهایی به {target_user.get_full_name() or target_user.username} تفویض شد.')

        elif action == 'deactivate' and user_id.isdigit():
            FinalApprovalDelegate.objects.filter(delegated_user_id=int(user_id)).update(is_active=False)
            messages.success(request, 'تفویض اختیار لغو شد.')

        return redirect('final_approval_delegation')

    delegates = FinalApprovalDelegate.objects.select_related('delegated_user', 'granted_by').order_by('-is_active', '-updated_at')
    # کاربران مالی قابل تفویض
    finance_users = User.objects.filter(
        profile__role__in=['finance', 'finance_manager'], is_active=True
    ).select_related('profile').exclude(
        id__in=delegates.values_list('delegated_user_id', flat=True)
    )
    return render(request, 'payments/final_approval_delegation.html', {
        'delegates': delegates,
        'finance_users': finance_users,
        'is_staff_user': True,
    })


@login_required
@require_POST
def delegate_final_approval(request, payment_id):
    """تفویض اختیار تأیید نهایی به کاربر دیگر توسط مدیر مالی."""
    redirect_target = _safe_next_url(request, default=reverse('submit'))
    role = _user_role(request.user)
    if not _can_delegate_final_approval(role, request.user.is_superuser):
        messages.error(request, 'فقط مدیر مالی می‌تواند تفویض اختیار کند.')
        return redirect(redirect_target)

    payment = get_object_or_404(PaymentRecord, id=payment_id)
    delegate_user_id = request.POST.get('delegate_to') or ''

    if delegate_user_id == 'clear':
        payment.final_approval_delegated_to = None
        payment.save(update_fields=['final_approval_delegated_to'])
        messages.success(request, f'تفویض اختیار تأیید نهایی سند #{payment_id} لغو شد.')
    elif delegate_user_id.isdigit():
        delegate_user = get_object_or_404(User, id=int(delegate_user_id), is_active=True)
        payment.final_approval_delegated_to = delegate_user
        payment.save(update_fields=['final_approval_delegated_to'])
        # اطلاع‌رسانی به کاربر تفویض‌شده
        customer_name = f"{payment.first_name} {payment.last_name}".strip() or (payment.user.username if payment.user else f'#{payment_id}')
        _notify_users(
            [delegate_user],
            '📋 تفویض اختیار تأیید نهایی',
            f'اختیار تأیید نهایی سند #{payment_id} مشتری {customer_name} به شما تفویض شد.',
            reverse('submit'), category=UserNotification.CATEGORY_SYSTEM, actor=request.user,
        )
        _log_activity(payment, request.user, PaymentActivityLog.ACTION_STATUS_CHANGED,
                      note=f'تفویض اختیار تأیید نهایی به {_display_name(delegate_user)}')
        messages.success(request, f'اختیار تأیید نهایی سند #{payment_id} به {delegate_user.get_full_name() or delegate_user.username} تفویض شد.')
    else:
        messages.error(request, 'کاربر نامعتبر است.')

    return redirect(redirect_target)


@login_required
@require_POST
def finance_unified_action(request, payment_id):
    """عملیات یکپارچه مالی — ثبت مالی یا عودت به بازرگانی."""
    redirect_target = _safe_next_url(request, default=request.META.get('HTTP_REFERER') or reverse('submit'))
    action_type = (request.POST.get('finance_action') or '').strip()
    note = (request.POST.get('note') or '').strip()

    if action_type == 'finance_register':
        payment = get_object_or_404(PaymentRecord, id=payment_id)
        role = _user_role(request.user)
        if not _can_finance_register(role, payment, request.user.is_superuser):
            messages.error(request, 'مجاز به ثبت مالی این سند نیستید.')
            return redirect(redirect_target)
        payment.finance_status = PaymentRecord.FINANCE_STATUS_APPROVED
        payment.finance_registered_at = timezone.now()
        payment.finance_registered_by = request.user
        _sync_pending_final_flag(payment)
        payment.save(update_fields=[
            'finance_status', 'finance_registered_at', 'finance_registered_by',
            'pending_final_approval', 'pending_final_approval_since',
        ])
        _log_activity(payment, request.user, PaymentActivityLog.ACTION_FINANCE_REGISTERED, note=note)
        if payment.pending_final_approval:
            customer_name = f"{payment.first_name} {payment.last_name}".strip() or (payment.user.username if payment.user else f'#{payment_id}')
            _notify_users(
                list(_staff_notification_users({'finance_manager'})),
                '✅ سند آماده تأیید نهایی',
                f'سند #{payment_id} مشتری {customer_name} هم ثبت بازرگانی و هم ثبت مالی دارد.',
                reverse('pending_final_approval'), category=UserNotification.CATEGORY_SYSTEM, actor=request.user,
            )
        messages.success(request, f'ثبت مالی سند #{payment_id} انجام شد.')

    elif action_type == 'return_to_commercial':
        payment = get_object_or_404(PaymentRecord, id=payment_id)
        role = _user_role(request.user)
        dept = _department_role(role)
        if dept != 'finance' and not request.user.is_superuser:
            messages.error(request, 'فقط واحد مالی می‌تواند سند را عودت دهد.')
            return redirect(redirect_target)
        if not note:
            messages.error(request, 'برای عودت به بازرگانی، ثبت توضیح الزامی است.')
            return redirect(redirect_target)
        old_status = payment.status
        payment.status = PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL
        payment.finance_status = None
        payment.finance_registered_at = None
        payment.finance_registered_by = None
        payment.pending_final_approval = False
        payment.pending_final_approval_since = None
        payment.save(update_fields=[
            'status', 'finance_status', 'finance_registered_at', 'finance_registered_by',
            'pending_final_approval', 'pending_final_approval_since',
        ])
        _log_activity(payment, request.user, PaymentActivityLog.ACTION_STATUS_CHANGED,
                      from_status=old_status, to_status=payment.status,
                      note=note)
        messages.warning(request, f'سند #{payment_id} به بازرگانی عودت داده شد.')

    else:
        messages.error(request, 'عملیات نامعتبر است.')

    return redirect(redirect_target)


@login_required
@require_POST
def finance_register_payment(request, payment_id):
    """ثبت مالی — فلگ مستقل مالی. هر زمانی قابل انجام است (مگر رد/تأیید نهایی)."""
    redirect_target = _safe_next_url(request, default=request.META.get('HTTP_REFERER') or reverse('submit'))
    role = _user_role(request.user)
    if not _can_finance_register(role, PaymentRecord.objects.get(id=payment_id), request.user.is_superuser):
        messages.error(request, 'شما مجاز به ثبت مالی این سند نیستید.')
        return redirect(redirect_target)

    payment = get_object_or_404(PaymentRecord, id=payment_id)
    note = (request.POST.get('note') or '').strip()

    payment.finance_status = PaymentRecord.FINANCE_STATUS_APPROVED
    payment.finance_registered_at = timezone.now()
    payment.finance_registered_by = request.user
    _sync_pending_final_flag(payment)
    payment.save(update_fields=[
        'finance_status', 'finance_registered_at', 'finance_registered_by',
        'pending_final_approval', 'pending_final_approval_since',
    ])

    _log_activity(payment, request.user, PaymentActivityLog.ACTION_FINANCE_REGISTERED,
                  note=note or '')

    if payment.pending_final_approval:
        customer_name = f"{payment.first_name} {payment.last_name}".strip() or (payment.user.username if payment.user else f'#{payment_id}')
        _notify_users(
            list(_staff_notification_users({'finance_manager'})),
            '✅ سند آماده تأیید نهایی',
            f'سند #{payment_id} مشتری {customer_name} هم ثبت بازرگانی و هم ثبت مالی دارد و آماده تأیید نهایی است.',
            reverse('pending_final_approval'),
            category=UserNotification.CATEGORY_SYSTEM,
            actor=request.user,
        )

    messages.success(request, f'ثبت مالی سند #{payment_id} با موفقیت انجام شد.')
    return redirect(redirect_target)


@login_required
@require_POST
def finance_final_approve(request, payment_id):
    """تأیید نهایی — فقط مدیر مالی، فقط وقتی هر دو فلگ تکمیل شده."""
    redirect_target = _safe_next_url(request, default=request.META.get('HTTP_REFERER') or reverse('submit'))
    payment = get_object_or_404(PaymentRecord, id=payment_id)
    role = _user_role(request.user)

    if not _can_final_approve(role, payment, request.user.is_superuser):
        messages.error(request, 'تأیید نهایی فقط توسط مدیر مالی و در صورت تکمیل هر دو فلگ مجاز است.')
        return redirect(redirect_target)

    note = (request.POST.get('note') or '').strip()
    old_status = payment.status
    payment.status = PaymentRecord.STATUS_FINAL_APPROVED
    payment.pending_final_approval = False
    payment.pending_final_approval_since = None
    payment.save(update_fields=['status', 'pending_final_approval', 'pending_final_approval_since'])

    _log_activity(payment, request.user, PaymentActivityLog.ACTION_FINAL_APPROVED,
                  from_status=old_status, to_status=payment.status,
                  note=note or '')

    customer_name = f"{payment.first_name} {payment.last_name}".strip() or (payment.user.username if payment.user else f'#{payment_id}')
    _notify_users(
        [payment.user] if payment.user else [],
        'تأیید نهایی سند',
        f'سند #{payment_id} مشتری {customer_name} توسط مدیر مالی تأیید نهایی شد.',
        reverse('submit'),
        category=UserNotification.CATEGORY_SYSTEM,
        actor=request.user,
    )

    messages.success(request, f'سند #{payment_id} با موفقیت تأیید نهایی شد.')
    return redirect(redirect_target)


@login_required
@require_POST
def staff_update_status(request, payment_id):
    redirect_target = _safe_next_url(request, default=request.META.get('HTTP_REFERER') or '')
    if not redirect_target:
        redirect_target = 'submit'

    if not _is_staff_user(request.user):
        messages.error(request, 'شما دسترسی بررسی اسناد را ندارید.')
        return redirect(redirect_target)

    payment = get_object_or_404(PaymentRecord, id=payment_id)
    staff_role = _user_role(request.user)
    can_act = _can_staff_act_on_payment(staff_role, payment, is_system_admin=request.user.is_superuser)
    if not can_act:
        dept = _department_role(staff_role)
        if dept in {'commercial', 'sales'}:
            fresh_logs = list(payment.activity_logs.select_related('actor', 'actor__profile').all())
            can_act = _commercial_can_revise(payment, logs=fresh_logs)
    if not can_act:
        messages.error(request, 'در وضعیت فعلی، امکان تغییر این سند برای شما وجود ندارد.')
        return redirect(redirect_target)

    form = StaffStatusUpdateForm(request.POST)
    if not form.is_valid():
        for field_errors in form.errors.values():
            for error in field_errors:
                messages.error(request, error)
        return redirect(redirect_target)

    target_status = form.cleaned_data['status']
    allowed_statuses = {value for value, _ in _staff_status_choices_for_role(staff_role)}
    if not request.user.is_superuser and target_status not in allowed_statuses:
        messages.error(request, 'این تغییر وضعیت برای نقش شما مجاز نیست.')
        return redirect(redirect_target)

    note = (form.cleaned_data['note'] or '').strip()
    if target_status == PaymentRecord.STATUS_INCOMPLETE and not note:
        messages.error(request, 'برای وضعیت «ناقص»، ثبت توضیح الزامی است.')
        return redirect(redirect_target)

    from_status = payment.status
    payment.status = target_status
    payment.last_staff_note = note
    payment.rejection_reason = form.cleaned_data['rejection_reason'] if target_status == PaymentRecord.STATUS_REJECTED else ''
    department_role = _department_role(staff_role)

    # Finance can hard-lock records on terminal decisions.
    if request.user.is_superuser:
        payment.is_locked = False
    elif department_role == 'finance' and target_status in {
        PaymentRecord.STATUS_FINAL_APPROVED,
        PaymentRecord.STATUS_REJECTED,
        PaymentRecord.STATUS_INCOMPLETE,
    }:
        payment.is_locked = True

    selected_counterparty = form.cleaned_data['counterparty']
    if selected_counterparty and department_role in {'commercial', 'staff'}:
        payment.counterparty = selected_counterparty

    update_fields = ['status', 'last_staff_note', 'rejection_reason', 'counterparty', 'is_locked']

    # سند رد شده: pending_final_approval پاک می‌شود تا از صف تأیید خارج شود
    if target_status == PaymentRecord.STATUS_REJECTED:
        payment.pending_final_approval = False
        payment.pending_final_approval_since = None
        update_fields += ['pending_final_approval', 'pending_final_approval_since']

    # هر بار که سند از حالت ناقص خارج می‌شود، ثبت مالی قبلی ابطال می‌شود
    REACTIVATE_FROM_INCOMPLETE = {
        PaymentRecord.STATUS_PENDING,
        PaymentRecord.STATUS_COMMERCIAL_REVIEW,
        PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL,
    }
    if (request.user.is_superuser and from_status == PaymentRecord.STATUS_INCOMPLETE
            and target_status in REACTIVATE_FROM_INCOMPLETE):
        payment.finance_status = None
        payment.finance_registered_at = None
        payment.finance_registered_by = None
        payment.pending_final_approval = False
        payment.pending_final_approval_since = None
        update_fields += [
            'finance_status', 'finance_registered_at', 'finance_registered_by',
            'pending_final_approval', 'pending_final_approval_since',
        ]

    payment.save(update_fields=update_fields)

    _log_activity(
        payment,
        request.user,
        PaymentActivityLog.ACTION_STATUS_CHANGED,
        from_status=from_status,
        to_status=payment.status,
        note=payment.last_staff_note,
    )
    _notify_payment_status_changed(payment, request.user, from_status, payment.status)

    messages.success(request, 'وضعیت سند با موفقیت ثبت شد.')
    return redirect(redirect_target)


@login_required
def edit_payment(request, payment_id):
    payment = get_object_or_404(PaymentRecord, id=payment_id)
    return_url = _safe_next_url(request)

    if _is_staff_user(request.user):
        return HttpResponseForbidden('کاربران واحدها امکان ویرایش سند مشتری را ندارند.')

    if payment.user_id != request.user.id:
        return HttpResponseForbidden('فقط امکان ویرایش اسناد ثبت شده توسط خودتان وجود دارد.')

    if not _can_customer_edit_payment(payment):
        messages.error(request, 'این سند توسط واحدهای سازمانی ثبت یا بررسی شده است و قابل ویرایش نیست.')
        return redirect(return_url or 'submit')

    profile = None
    try:
        profile = request.user.profile
    except UserProfile.DoesNotExist:
        profile = None

    initial_data = _account_initial_data(request.user, profile, payment=payment)

    if request.method == 'POST':
        form = PaymentRecordForm(request.POST, request.FILES, instance=payment, initial=initial_data)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.user = request.user
            payment.first_name = initial_data['first_name']
            payment.last_name = initial_data['last_name']
            payment.organization = initial_data['organization']
            payment.city = initial_data['city']
            payment.phone = initial_data['phone']
            from_status = payment.status
            payment.status = PaymentRecord.STATUS_PENDING
            payment.is_locked = False
            # وقتی مشتری رفع نقص می‌کند، ثبت مالی قبلی ابطال می‌شود تا مالی مجدداً بررسی کند
            payment.finance_status = None
            payment.finance_registered_at = None
            payment.finance_registered_by = None
            payment.pending_final_approval = False
            payment.pending_final_approval_since = None
            payment.created_at = timezone.now()
            payment.last_edited_at = timezone.now()
            payment.save()
            _save_receipts(payment, form)
            _log_activity(
                payment, request.user, PaymentActivityLog.ACTION_EDITED,
                from_status=from_status, to_status=payment.status,
                note='رفع نقص توسط مشتری - ثبت مالی ابطال شد',
            )
            _notify_payment_edited(payment, request.user, title='ویرایش فیش توسط مشتری')
            messages.success(request, 'سند با موفقیت ویرایش شد و برای بررسی مجدد در صف قرار گرفت.')
            return redirect(return_url or 'submit')
    else:
        form = PaymentRecordForm(instance=payment, initial=initial_data)

    return render(request, 'payments/edit_payment.html', {
        'form': form,
        'payment': payment,
        'source_profiles': _source_profiles_for_user(request.user),
        'destination_profiles': _destination_profiles_for_user(request.user),
        'customer_info': initial_data,
        'customer_debt': _customer_debt_summary(request.user),
        'return_url': return_url,
    })


@login_required
def payment_timeline(request, payment_id):
    payment = get_object_or_404(
        PaymentRecord.objects.select_related('user', 'counterparty').prefetch_related('receipts'),
        id=payment_id,
    )
    is_staff_user = _is_staff_user(request.user)
    if not is_staff_user and payment.user_id != request.user.id:
        return HttpResponseForbidden('فقط امکان مشاهده تاریخچه اسناد خودتان وجود دارد.')

    _mark_notifications_read_for_url(request.user, request.path)
    if not is_staff_user and payment.customer_seen_at is None:
        payment.customer_seen_at = timezone.now()
        payment.save(update_fields=['customer_seen_at'])
    _log_activity(payment, request.user, PaymentActivityLog.ACTION_VIEWED, note='مشاهده تاریخچه')
    raw_logs = payment.activity_logs.select_related('actor', 'actor__profile').all()

    if is_staff_user:
        # کارکنان: جزئیات کامل — مدیر سیستم از نمایش حذف می‌شود
        logs = []
        for log in raw_logs:
            if _is_superuser_actor(log):
                continue
            logs.append({
                'text':        _log_text(log),
                'note':        log.note,
                'jalali_time': _format_jalali_datetime(log.created_at),
                'action':      log.action,
                'actor_id':    log.actor_id,
                'actor_name':  _display_name(log.actor),
                'actor_role':  _role_title(log.actor),
                'from_status': dict(PaymentRecord.STATUS_CHOICES).get(log.from_status, log.from_status or ''),
                'to_status':   dict(PaymentRecord.STATUS_CHOICES).get(log.to_status, log.to_status or ''),
                'is_customer_note': log.action == PaymentActivityLog.ACTION_CUSTOMER_NOTE,
            })
        logs = _group_consecutive_views(logs)
    else:
        # مشتری: وضعیت‌های کلیدی بدون جزئیات کارکنان
        logs = [
            {
                'text':             row['text'],
                'note':             row['note'],
                'jalali_time':      row['time'],
                'is_customer_note': row.get('is_customer_note', False),
                'icon':             row.get('icon', '🔄'),
            }
            for row in _customer_visible_logs(raw_logs)
        ]

    can_add_note = (
        not is_staff_user
        and payment.user_id == request.user.id
        and payment.status != PaymentRecord.STATUS_FINAL_APPROVED
    )
    return render(request, 'payments/timeline.html', {
        'payment': payment,
        'logs': logs,
        'is_staff_user': is_staff_user,
        'can_add_note': can_add_note,
        'return_url': _safe_next_url(request),
        'return_label': _return_link_label(request, 'بازگشت'),
    })


@login_required
@require_POST
def add_payment_note(request, payment_id):
    payment = get_object_or_404(PaymentRecord, id=payment_id)
    return_url = _safe_next_url(request)
    timeline_url = reverse('payment_timeline', args=[payment_id])
    if return_url:
        timeline_url += f'?next={return_url}'

    if payment.user_id != request.user.id:
        return HttpResponseForbidden('فقط صاحب فیش می‌تواند توضیح اضافه کند.')

    if payment.status == PaymentRecord.STATUS_FINAL_APPROVED:
        messages.error(request, 'فیش تایید نهایی شده و امکان افزودن توضیح وجود ندارد.')
        return redirect(timeline_url)

    note_text = (request.POST.get('note') or '').strip()
    if not note_text:
        messages.error(request, 'متن توضیح نمی‌تواند خالی باشد.')
    elif len(note_text) > 1000:
        messages.error(request, 'توضیح حداکثر ۱۰۰۰ کاراکتر می‌تواند باشد.')
    else:
        _log_activity(payment, request.user, PaymentActivityLog.ACTION_CUSTOMER_NOTE, note=note_text)
        _notify_payment_customer_note(payment, request.user, note_text)
        messages.success(request, 'توضیح شما با موفقیت در روال فیش ثبت شد.')

    return redirect(timeline_url)


@login_required
def counterparties_manage(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden('شما دسترسی مدیریت طرف حساب را ندارید.')

    if request.method == 'POST':
        form = CounterpartyForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('counterparties_manage')
    else:
        form = CounterpartyForm()

    counterparties = Counterparty.objects.all()
    page_obj = _paginate_queryset(request, counterparties, per_page=10, page_param='page')
    page_base_query = _build_query_string(request, remove_keys=['page'])
    return render(request, 'payments/counterparties.html', {
        'form': form,
        'counterparties': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'export_dataset': 'counterparties',
        'export_fields': COUNTERPARTY_EXPORT_FIELDS,
    })


@login_required
def counterparty_manage_list(request):
    """لیست و مدیریت طرف حساب‌ها — قابل دسترس برای مدیران."""
    if not _can_manage_users(request.user) and not request.user.is_superuser:
        return HttpResponseForbidden('دسترسی ممنوع.')

    q = (request.GET.get('q') or '').strip()
    status_f = (request.GET.get('status') or '').strip()
    qs = Counterparty.objects.select_related('user').prefetch_related('bank_accounts').order_by('name')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(phone__icontains=q))
    if status_f:
        qs = qs.filter(status=status_f)

    page_obj = _paginate_queryset(request, qs, per_page=15)
    page_base_query = _build_query_string(request, remove_keys=['page'])
    return render(request, 'payments/counterparty_manage.html', {
        'counterparties': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'filters': {'q': q, 'status': status_f},
        'status_choices': Counterparty.STATUS_CHOICES,
        'editing': None,
    })


@login_required
def counterparty_manage_edit(request, cp_id=None):
    """ایجاد یا ویرایش طرف حساب با حساب‌های بانکی."""
    if not _can_manage_users(request.user) and not request.user.is_superuser:
        return HttpResponseForbidden('دسترسی ممنوع.')

    cp_instance = get_object_or_404(Counterparty, id=cp_id) if cp_id else None

    if request.method == 'POST':
        form = CounterpartyManagementForm(request.POST, instance=cp_instance)
        bank_formset = CounterpartyBankAccountFormSet(request.POST, instance=cp_instance or Counterparty())
        if form.is_valid():
            cp = form.save(commit=False)
            bank_fs = CounterpartyBankAccountFormSet(request.POST, instance=cp)
            if bank_fs.is_valid():
                cp.save()
                bank_fs.instance = cp
                bank_fs.save()
                action = 'ایجاد طرف حساب' if not cp_id else 'ویرایش طرف حساب'
                _log_system_activity(request.user, None, 'counterparty_action', f'{action}: {cp.name}')
                messages.success(request, f'طرف حساب «{cp.name}» با موفقیت ذخیره شد.')
                return redirect('counterparty_manage_list')
        else:
            bank_formset = CounterpartyBankAccountFormSet(request.POST, instance=cp_instance or Counterparty())
    else:
        form = CounterpartyManagementForm(instance=cp_instance)
        bank_formset = CounterpartyBankAccountFormSet(instance=cp_instance or Counterparty())

    return render(request, 'payments/counterparty_manage.html', {
        'counterparties': Counterparty.objects.select_related('user').order_by('name')[:50],
        'page_obj': None,
        'filters': {},
        'status_choices': Counterparty.STATUS_CHOICES,
        'editing': cp_instance,
        'form': form,
        'bank_formset': bank_formset,
    })


def _payment_detail_changes_note(before, after, form):
    lines = []
    for field_name in form.fields:
        old_value = before.get(field_name)
        new_value = getattr(after, field_name)
        old_text = '' if old_value is None else str(old_value)
        new_text = '' if new_value is None else str(new_value)
        if old_text != new_text:
            label = form.fields[field_name].label or field_name
            lines.append(f'{label}: «{old_text or "-"}» به «{new_text or "-"}»')
    return '\n'.join(lines)


@login_required
def system_logo_settings(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden('فقط مدیر سیستم امکان تغییر لوگوی سامانه را دارد.')
    settings_obj = SystemSettings.load()
    if request.method == 'POST':
        if request.POST.get('form_name') == 'menu_settings':
            menu_form = SystemMenuSettingsForm(request.POST, instance=settings_obj)
            logo_form = SystemLogoSettingsForm(instance=settings_obj)
            if menu_form.is_valid():
                menu_form.save()
                messages.success(request, 'تنظیمات منو ذخیره شد.')
                return redirect('system_logo_settings')
            messages.error(request, 'ذخیره تنظیمات منو انجام نشد. لطفا خطاها را بررسی کنید.')
        else:
            logo_form = SystemLogoSettingsForm(request.POST, request.FILES, instance=settings_obj)
            menu_form = SystemMenuSettingsForm(instance=settings_obj)
            if logo_form.is_valid():
                logo_form.save()
                messages.success(request, 'تنظیمات لوگوی سامانه ذخیره شد.')
                return redirect('system_logo_settings')
            messages.error(request, 'ذخیره لوگو انجام نشد. لطفا خطاها را بررسی کنید.')
    else:
        logo_form = SystemLogoSettingsForm(instance=settings_obj)
        menu_form = SystemMenuSettingsForm(instance=settings_obj)
    return render(request, 'payments/system_logo_settings.html', {
        'form': logo_form,
        'menu_form': menu_form,
        'settings_obj': settings_obj,
    })


@login_required
def reconciliation_center(request):
    if not _can_access_reconciliation(request.user):
        return HttpResponseForbidden('شما دسترسی مغایرت‌گیری ندارید.')

    threads_qs = _reconciliation_threads_for_user(request.user)
    is_customer_user = _user_role(request.user) == 'customer'

    customer_options = []
    if not is_customer_user:
        customer_ids = threads_qs.order_by().values_list('customer_id', flat=True).distinct()
        customer_options = [
            (customer.id, customer.profile.display_name)
            for customer in User.objects.filter(id__in=customer_ids).select_related('profile').order_by('first_name', 'last_name', 'username')
        ]

    thread_tab = request.GET.get('tab') if request.GET.get('tab') in {'all', 'by_customer'} else 'all'
    selected_customer_id = request.GET.get('customer', '').strip()
    if thread_tab == 'by_customer' and not is_customer_user:
        if selected_customer_id:
            threads_qs = threads_qs.filter(customer_id=selected_customer_id)
        else:
            threads_qs = threads_qs.none()
    else:
        thread_tab = 'all'
        selected_customer_id = ''

    active_thread = None
    selected_id = request.GET.get('thread')
    if selected_id:
        active_thread = get_object_or_404(threads_qs, id=selected_id)
    else:
        active_thread = threads_qs.first()

    thread_form = ReconciliationThreadForm(user=request.user)
    message_form = ReconciliationMessageForm()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create_thread':
            thread_form = ReconciliationThreadForm(request.POST, user=request.user)
            if thread_form.is_valid():
                thread = thread_form.save(commit=False)
                if _user_role(request.user) == 'customer':
                    thread.customer = request.user
                thread.created_by = request.user
                thread.save()
                thread_form.save_m2m()
                # سازنده کارشناس را خودکار به گفتگو اضافه می‌کند تا دسترسی داشته باشد
                if _is_staff_user(request.user) and not thread.staff_participants.filter(id=request.user.id).exists():
                    thread.staff_participants.add(request.user)
                messages.success(request, 'گفتگوی مغایرت‌گیری ایجاد شد.')
                return redirect(f"{reverse('reconciliation_center')}?thread={thread.id}")
        elif action == 'send_message':
            thread = get_object_or_404(ReconciliationThread, id=request.POST.get('thread_id'))
            if not _can_access_reconciliation_thread(request.user, thread):
                return HttpResponseForbidden('شما عضو این گفتگو نیستید.')
            if thread.status == ReconciliationThread.STATUS_CLOSED:
                messages.error(request, 'این گفتگو بسته شده است.')
                return redirect(f"{reverse('reconciliation_center')}?thread={thread.id}")
            message_form = ReconciliationMessageForm(request.POST, request.FILES)
            if message_form.is_valid():
                message = message_form.save(commit=False)
                message.thread = thread
                message.sender = request.user
                uploaded = message_form.cleaned_data.get('attachment')
                if uploaded:
                    message.attachment_name = uploaded.name
                message.save()
                thread.updated_at = timezone.now()
                thread.save(update_fields=['updated_at'])
                # باطل کردن cache شمارش پیام‌های نخوانده برای سایر اعضای thread
                from django.core.cache import cache as _cache
                from .context_processors import recon_unread_cache_key
                for _p in thread.staff_participants.exclude(id=request.user.id):
                    _cache.delete(recon_unread_cache_key(_p.id))
                if thread.customer_id and thread.customer_id != request.user.id:
                    _cache.delete(recon_unread_cache_key(thread.customer_id))
                return redirect(f"{reverse('reconciliation_center')}?thread={thread.id}")
            active_thread = thread
        elif action in {'close_thread', 'open_thread'}:
            thread = get_object_or_404(threads_qs, id=request.POST.get('thread_id'))
            if _user_role(request.user) == 'customer' and not request.user.is_superuser:
                return HttpResponseForbidden('بستن یا بازکردن گفتگو فقط برای کارشناسان مجاز است.')
            thread.status = ReconciliationThread.STATUS_CLOSED if action == 'close_thread' else ReconciliationThread.STATUS_OPEN
            thread.save(update_fields=['status', 'updated_at'])
            return redirect(f"{reverse('reconciliation_center')}?thread={thread.id}")

    filtered_threads_qs, thread_filters = _apply_reconciliation_filters(threads_qs, request)
    thread_rows = list(filtered_threads_qs[:80])
    for thread in thread_rows:
        thread.document_url = _reconciliation_document_url(thread.document_type, thread.document_id, thread.id)
        thread.last_message = thread.messages.last()

    active_messages = []
    if active_thread:
        _mark_reconciliation_thread_read(active_thread, request.user)
        active_thread.document_url = _reconciliation_document_url(active_thread.document_type, active_thread.document_id, active_thread.id)
        active_messages = list(active_thread.messages.select_related('sender', 'sender__profile'))
        for message in active_messages:
            message.document_url = _reconciliation_document_url(message.document_type, message.document_id, active_thread.id)

    return render(request, 'payments/reconciliation_center.html', {
        'threads': thread_rows,
        'thread_filters': thread_filters,
        'document_type_choices': ReconciliationThread.DOCUMENT_CHOICES,
        'status_choices': ReconciliationThread.STATUS_CHOICES,
        'active_thread': active_thread,
        'active_messages': active_messages,
        'thread_form': thread_form,
        'message_form': message_form,
        'is_customer_user': is_customer_user,
        'can_manage_thread_state': request.user.is_superuser or (_user_role(request.user) != 'customer' and _can_access_reconciliation(request.user)),
        'thread_tab': thread_tab,
        'customer_options': customer_options,
        'selected_customer_id': selected_customer_id,
    })


@login_required
def reconciliation_poll(request):
    """به‌روزرسانی لحظه‌ای: پیام‌های جدید گفتگوی فعال + فهرست گفتگوها بدون نیاز به رفرش صفحه."""
    if not _can_access_reconciliation(request.user):
        return HttpResponseForbidden('شما دسترسی مغایرت‌گیری ندارید.')

    threads_qs = _reconciliation_threads_for_user(request.user)
    is_customer_user = _user_role(request.user) == 'customer'

    thread_tab = request.GET.get('tab') if request.GET.get('tab') in {'all', 'by_customer'} else 'all'
    selected_customer_id = request.GET.get('customer', '').strip()
    if thread_tab == 'by_customer' and not is_customer_user:
        if selected_customer_id:
            threads_qs = threads_qs.filter(customer_id=selected_customer_id)
        else:
            threads_qs = threads_qs.none()
    else:
        thread_tab = 'all'
        selected_customer_id = ''

    active_thread = None
    selected_id = request.GET.get('thread')
    if selected_id:
        active_thread = threads_qs.filter(id=selected_id).first()

    filtered_threads_qs, thread_filters = _apply_reconciliation_filters(threads_qs, request)
    thread_rows = list(filtered_threads_qs[:80])
    for thread in thread_rows:
        thread.document_url = _reconciliation_document_url(thread.document_type, thread.document_id, thread.id)
        thread.last_message = thread.messages.last()

    messages_html = ''
    last_message_id = 0
    if active_thread:
        after_id = int(request.GET.get('after') or 0)
        _mark_reconciliation_thread_read(active_thread, request.user)
        new_messages = list(
            active_thread.messages
            .select_related('sender', 'sender__profile')
            .filter(id__gt=after_id)
        )
        for message in new_messages:
            message.document_url = _reconciliation_document_url(message.document_type, message.document_id, active_thread.id)
        if new_messages:
            last_message_id = new_messages[-1].id
            messages_html = ''.join(
                render_to_string('payments/partials/_reconciliation_message.html', {'message': message}, request=request)
                for message in new_messages
            )
        else:
            last_message_id = after_id

    threads_html = render_to_string('payments/partials/_reconciliation_thread_list.html', {
        'threads': thread_rows,
        'active_thread': active_thread,
        'thread_tab': thread_tab,
        'selected_customer_id': selected_customer_id,
        'thread_filters': thread_filters,
    }, request=request)

    return JsonResponse({
        'messages_html': messages_html,
        'last_message_id': last_message_id,
        'threads_html': threads_html,
    })


_THREAD_PREDEFINED_TITLES = {
    'status_change':    'درخواست تغییر وضعیت',
    'payment_review':   'بررسی فیش واریزی',
    'discrepancy':      'مغایرت واریز',
    'final_approval':   'درخواست تأیید نهایی',
}


@login_required
def payment_start_thread(request):
    """ایجاد یا بازیابی گفتگوی مغایرت‌گیری برای یک سند پرداخت."""
    if request.method != 'POST':
        return HttpResponseForbidden()

    if not _can_access_reconciliation(request.user):
        messages.error(request, 'دسترسی به مغایرت‌گیری ندارید.')
        return redirect(request.POST.get('next') or reverse('submit'))

    payment_id = request.POST.get('payment_id', '').strip()
    payment = get_object_or_404(
        PaymentRecord.objects.select_related('user', 'user__profile'),
        id=payment_id,
    )

    is_staff = _is_staff_user(request.user)
    is_owner = payment.user_id and payment.user_id == request.user.id
    if not is_staff and not is_owner:
        messages.error(request, 'دسترسی ندارید.')
        return redirect(request.POST.get('next') or reverse('submit'))

    if not payment.user:
        messages.error(request, 'این سند فاقد حساب کاربری مشتری است.')
        return redirect(request.POST.get('next') or reverse('submit'))

    title_choice = request.POST.get('title_choice', 'status_change')
    title_custom = request.POST.get('title_custom', '').strip()
    if title_choice == 'custom':
        base_title = title_custom or 'درخواست تغییر وضعیت'
    else:
        base_title = _THREAD_PREDEFINED_TITLES.get(title_choice, 'درخواست تغییر وضعیت')

    from .templatetags.payment_dates import thousand_sep as _tsep, jalali_date as _jdate
    amount_str = _tsep(payment.amount) if payment.amount else '—'
    pay_date_str = _jdate(payment.pay_date) if payment.pay_date else '—'
    auto_body = (
        f"لطفا وضعیت سند را تغییر دهید.\n\n"
        f"جزئیات سند:\n"
        f"• شماره سند: #{payment.id}\n"
        f"• مبلغ: {amount_str} ریال\n"
        f"• تاریخ واریز: {pay_date_str}\n"
        f"• کد پیگیری: {payment.tracking_code or '—'}\n"
        f"• وضعیت فعلی: {payment.get_status_display()}"
    )

    from django.core.cache import cache as _cache
    from .context_processors import recon_unread_cache_key

    def _invalidate_thread_caches(thread_obj):
        for _p in thread_obj.staff_participants.exclude(id=request.user.id):
            _cache.delete(recon_unread_cache_key(_p.id))
        if thread_obj.customer_id and thread_obj.customer_id != request.user.id:
            _cache.delete(recon_unread_cache_key(thread_obj.customer_id))

    existing = ReconciliationThread.objects.filter(
        document_type=ReconciliationThread.DOC_PAYMENT,
        document_id=payment.id,
    ).order_by('-updated_at').first()

    if existing:
        if existing.status == ReconciliationThread.STATUS_CLOSED:
            existing.status = ReconciliationThread.STATUS_OPEN
            existing.save(update_fields=['status', 'updated_at'])
        ReconciliationMessage.objects.create(
            thread=existing,
            sender=request.user,
            body=auto_body,
        )
        existing.updated_at = timezone.now()
        existing.save(update_fields=['updated_at'])
        _invalidate_thread_caches(existing)
        messages.success(request, 'گفتگوی موجود بروزرسانی و پیام ارسال شد.')
        return redirect(f"{reverse('reconciliation_center')}?thread={existing.id}")

    thread_title = f"{base_title} — سند #{payment.id}"
    staff_qs = list(User.objects.filter(
        is_active=True,
        profile__role__in=['finance', 'commercial', 'sales'],
    ).select_related('profile'))

    thread = ReconciliationThread(
        title=thread_title,
        customer=payment.user,
        created_by=request.user,
        document_type=ReconciliationThread.DOC_PAYMENT,
        document_id=payment.id,
        status=ReconciliationThread.STATUS_OPEN,
    )
    thread.save()
    thread.staff_participants.set(staff_qs)
    if is_staff and not thread.staff_participants.filter(id=request.user.id).exists():
        thread.staff_participants.add(request.user)

    ReconciliationMessage.objects.create(
        thread=thread,
        sender=request.user,
        body=auto_body,
    )
    _invalidate_thread_caches(thread)
    messages.success(request, 'گفتگوی جدید ایجاد شد.')
    return redirect(f"{reverse('reconciliation_center')}?thread={thread.id}")


@login_required
def staff_edit_payment_details(request, payment_id):
    if not _can_edit_payment_details(request.user):
        return HttpResponseForbidden('شما دسترسی تکمیل اطلاعات فیش‌ها را ندارید.')

    payment = get_object_or_404(
        PaymentRecord.objects.select_related('user', 'counterparty').prefetch_related('receipts'),
        id=payment_id,
    )
    return_url = _safe_next_url(request, default=request.META.get('HTTP_REFERER') or reverse('submit'))
    editable_statuses = {
        PaymentRecord.STATUS_PENDING,
        PaymentRecord.STATUS_COMMERCIAL_REVIEW,
        PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL,
        PaymentRecord.STATUS_APPROVED,
        PaymentRecord.STATUS_INCOMPLETE,
    }
    if not request.user.is_superuser and payment.status not in editable_statuses:
        return HttpResponseForbidden('این سند در وضعیت فعلی قابل تکمیل اطلاعات نیست.')

    if request.method == 'POST':
        form = StaffPaymentDetailsForm(request.POST, instance=payment)
        before = {field_name: getattr(payment, field_name) for field_name in form.fields}
        if form.is_valid():
            payment = form.save(commit=False)
            change_note = _payment_detail_changes_note(before, payment, form)
            if change_note:
                payment.last_edited_at = timezone.now()
                payment.save(update_fields=list(form.fields.keys()) + ['last_edited_at'])
                _log_activity(
                    payment,
                    request.user,
                    PaymentActivityLog.ACTION_EDITED,
                    from_status=payment.status,
                    to_status=payment.status,
                    note=change_note,
                )
                _notify_payment_edited(payment, request.user, title='تکمیل اطلاعات فیش')
                messages.success(request, 'اطلاعات فیش ثبت و در تاریخچه سند لاگ شد.')
            else:
                messages.info(request, 'تغییری برای ثبت وجود نداشت.')
            return redirect(return_url)
    else:
        form = StaffPaymentDetailsForm(instance=payment, initial={'amount': payment.amount})

    return render(request, 'payments/staff_edit_payment_details.html', {
        'form': form,
        'payment': payment,
        'return_url': return_url,
    })


@login_required
def counterparty_edit(request, counterparty_id):
    if not request.user.is_superuser:
        return HttpResponseForbidden('شما دسترسی مدیریت طرف حساب را ندارید.')

    counterparty = get_object_or_404(Counterparty, id=counterparty_id)

    if request.method == 'POST':
        form = CounterpartyForm(request.POST, instance=counterparty)
        if form.is_valid():
            form.save()
            return redirect('counterparties_manage')
    else:
        form = CounterpartyForm(instance=counterparty)

    return render(request, 'payments/counterparty_edit.html', {'form': form, 'counterparty': counterparty})


@login_required
def users_manage(request):
    can_manage_user_accounts = _can_manage_users(request.user)
    can_review_profile_changes = _can_review_profile_change(request.user)
    if not (can_manage_user_accounts or can_review_profile_changes):
        return HttpResponseForbidden('شما دسترسی مدیریت کاربران را ندارید.')

    password_suggestion = _suggest_five_digit_password()
    filters = {
        'q': (request.GET.get('q') or '').strip(),
        'role': (request.GET.get('role') or '').strip(),
        'status': (request.GET.get('status') or '').strip(),
    }
    if request.method == 'POST' and not can_manage_user_accounts:
        return HttpResponseForbidden('شما دسترسی ایجاد یا ویرایش کاربران را ندارید.')

    if request.method == 'POST':
        form = UserAccountManagementForm(request.POST, request.FILES, password_suggestion=password_suggestion)
        if form.is_valid():
            created_user = form.save()
            _log_system_activity(
                request.user,
                created_user,
                SystemActivityLog.ACTION_USER_CREATED,
                'کاربر جدید ایجاد شد.',
            )
            messages.success(request, 'کاربر جدید با موفقیت ایجاد شد.')
            return redirect('users_manage')
    else:
        form = UserAccountManagementForm(
            initial={'password': password_suggestion, 'force_password_change': True, 'is_active': True},
            password_suggestion=password_suggestion,
        )

    pending_profile_changes = (
        ProfileChangeRequest.objects
        .filter(status=ProfileChangeRequest.STATUS_PENDING)
        .select_related('user', 'user__profile', 'requested_by')
        .order_by('-created_at', '-id')
    )
    users_source = _managed_users(query=filters['q'], role=filters['role'], status=filters['status']) if can_manage_user_accounts else User.objects.none()
    users_page = _paginate_queryset(
        request,
        users_source,
        per_page=10,
        page_param='page',
    )
    page_base_query = _build_query_string(request, remove_keys=['page'])
    return render(request, 'payments/users_manage.html', {
        'form': form,
        'users': users_page,
        'page_obj': users_page,
        'page_base_query': page_base_query,
        'password_suggestion': password_suggestion,
        'editing_user': None,
        'filters': filters,
        'export_dataset': 'users',
        'export_fields': USER_EXPORT_FIELDS,
        'pending_profile_changes': pending_profile_changes,
        'can_manage_user_accounts': can_manage_user_accounts,
        'can_review_profile_changes': can_review_profile_changes,
    })


@login_required
def user_edit(request, user_id):
    if not _can_manage_users(request.user):
        return HttpResponseForbidden('شما دسترسی مدیریت کاربران را ندارید.')

    managed_user = get_object_or_404(User.objects.select_related('profile'), id=user_id)
    password_suggestion = _suggest_five_digit_password()
    filters = {
        'q': (request.GET.get('q') or '').strip(),
        'role': (request.GET.get('role') or '').strip(),
        'status': (request.GET.get('status') or '').strip(),
    }

    if request.method == 'POST':
        form = UserAccountManagementForm(
            request.POST,
            request.FILES,
            instance=managed_user,
            password_suggestion=password_suggestion,
        )
        if form.is_valid():
            form.save()
            _log_system_activity(
                request.user,
                managed_user,
                SystemActivityLog.ACTION_USER_UPDATED,
                'اطلاعات کاربر ویرایش شد.',
            )
            messages.success(request, 'اطلاعات کاربر با موفقیت بروزرسانی شد.')
            return redirect('users_manage')
    else:
        form = UserAccountManagementForm(instance=managed_user, password_suggestion=password_suggestion)

    users_page = _paginate_queryset(
        request,
        _managed_users(query=filters['q'], role=filters['role'], status=filters['status']),
        per_page=10,
        page_param='page',
    )
    page_base_query = _build_query_string(request, remove_keys=['page'])
    return render(request, 'payments/users_manage.html', {
        'form': form,
        'users': users_page,
        'page_obj': users_page,
        'page_base_query': page_base_query,
        'password_suggestion': password_suggestion,
        'editing_user': managed_user,
        'filters': filters,
        'export_dataset': 'users',
        'export_fields': USER_EXPORT_FIELDS,
        'can_manage_user_accounts': True,
        'can_review_profile_changes': _can_review_profile_change(request.user),
        'pending_profile_changes': [],
    })


@login_required
def access_management(request):
    if not _can_manage_access(request.user):
        return HttpResponseForbidden('شما دسترسی مدیریت دسترسی‌ها را ندارید.')

    filters = {
        'q': (request.GET.get('q') or '').strip(),
        'role': (request.GET.get('role') or '').strip(),
    }
    users_source = _access_manageable_users(request.user, query=filters['q'], role=filters['role'])
    users_page = _paginate_queryset(request, users_source, per_page=10, page_param='page')
    page_base_query = _build_query_string(request, remove_keys=['page'])

    if request.user.is_superuser:
        role_choices = UserAccessManagementForm.ROLE_CHOICES
    else:
        dept_role = ACCESS_DEPARTMENT_MANAGER_ROLES.get(_user_role(request.user))
        role_choices = [c for c in UserAccessManagementForm.ROLE_CHOICES if c[0] == dept_role]

    return render(request, 'payments/access_management.html', {
        'users': users_page,
        'page_obj': users_page,
        'page_base_query': page_base_query,
        'filters': filters,
        'role_choices': role_choices,
        'allow_role_change': request.user.is_superuser,
    })


@login_required
def access_management_edit(request, user_id):
    target = get_object_or_404(User.objects.select_related('profile'), id=user_id)
    if not _can_manage_access_for_target(request.user, target):
        return HttpResponseForbidden('شما دسترسی مدیریت این کاربر را ندارید.')

    allow_role_change = request.user.is_superuser

    if request.method == 'POST':
        form = UserAccessManagementForm(request.POST, target=target, allow_role_change=allow_role_change)
        if form.is_valid():
            form.save()
            _log_system_activity(
                request.user,
                target,
                SystemActivityLog.ACTION_USER_UPDATED,
                'دسترسی‌های کاربر بروزرسانی شد.',
            )
            messages.success(request, 'دسترسی‌های کاربر با موفقیت بروزرسانی شد.')
            return redirect('access_management')
    else:
        form = UserAccessManagementForm(target=target, allow_role_change=allow_role_change)

    return render(request, 'payments/access_management_edit.html', {
        'form': form,
        'target': target,
    })


@login_required
def invoices_dashboard(request):
    is_staff_user = _is_staff_user(request.user)
    can_upload_invoices = _can_upload_invoices(request.user)

    if request.method == 'POST':
        if not can_upload_invoices:
            return HttpResponseForbidden('شما دسترسی بارگذاری فاکتور مشتری را ندارید.')
        form = InvoiceUploadForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            invoice = form.save(commit=False)
            invoice.uploaded_by = request.user
            invoice.amount = form.cleaned_data.get('amount')
            invoice.save()
            _notify_invoice_created(invoice, request.user)
            messages.success(request, 'فاکتور با موفقیت برای مشتری ثبت شد.')
            return redirect('invoices_dashboard')
    else:
        form = InvoiceUploadForm(user=request.user)

    can_view_invoices = _can_view_invoices(request.user)
    records = _invoice_records_for_user(request.user)
    records, filters = _apply_invoice_filters(records, request, is_staff_user=is_staff_user)
    page_obj = _paginate_queryset(request, records, per_page=10, page_param='page')
    page_base_query = _build_query_string(request, remove_keys=['page'])
    user_display_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username

    return render(request, 'payments/invoices.html', {
        'form': form,
        'records': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'filters': filters,
        'is_staff_user': is_staff_user,
        'can_upload_invoices': can_upload_invoices,
        'can_view_invoices': can_view_invoices,
        'can_delete_documents': _can_delete_customer_documents(request.user),
        'user_display_name': user_display_name,
        'customer_rows': _invoice_customer_rows() if can_upload_invoices else [],
        'export_dataset': 'invoices',
        'export_fields': INVOICE_EXPORT_FIELDS,
    })


@login_required
@require_POST
def invoice_parse_preview(request):
    if not _can_upload_invoices(request.user):
        return JsonResponse({'ok': False, 'message': 'شما دسترسی خواندن اطلاعات فاکتور را ندارید.'}, status=403)

    uploaded = request.FILES.get('attachment')
    if not uploaded:
        return JsonResponse({'ok': False, 'message': 'ابتدا فایل فاکتور را انتخاب کنید.'}, status=400)
    if os.path.splitext(uploaded.name or '')[1].lower() not in InvoiceUploadForm.ALLOWED_EXTENSIONS:
        return JsonResponse({'ok': False, 'message': 'خواندن خودکار فقط برای فایل PDF یا تصویر فاکتور فعال است.'}, status=400)
    max_size = UploadSettings.load().invoice_max_upload_size_bytes
    if uploaded.size and uploaded.size > max_size:
        return JsonResponse({'ok': False, 'message': 'حجم فایل بیشتر از حد مجاز فاکتور است.'}, status=400)

    job = create_preview_extraction_job(uploaded, requested_by=request.user)
    process_invoice_extraction_job(job.id)
    job.refresh_from_db()
    return JsonResponse(_invoice_extraction_payload(job))


@login_required
def invoice_parse_status(request, job_id):
    job = get_object_or_404(InvoiceExtractionJob, id=job_id)
    if job.requested_by_id != request.user.id and not request.user.is_superuser:
        return JsonResponse({'ok': False, 'message': 'شما دسترسی مشاهده این پردازش را ندارید.'}, status=403)
    return JsonResponse(_invoice_extraction_payload(job))


def _invoice_extraction_payload(job):
    result = job.result_json or {}
    fields = flatten_fields(result)
    return {
        **result,
        'job_id': job.id,
        'status': job.status,
        'fields': fields,
        'raw_text_preview': result.get('raw_text_preview', ''),
        'warnings': job.warnings,
        'message': result.get('message') or job.error_message or 'پردازش انجام شد.',
        'ok': bool(fields),
    }


@login_required
def price_lists_dashboard(request):
    is_staff_user = _is_staff_user(request.user)
    can_upload = _can_upload_price_lists(request.user)

    if request.method == 'POST':
        if not can_upload:
            return HttpResponseForbidden('شما دسترسی بارگذاری لیست قیمت را ندارید.')
        form = PriceListUploadForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            batch_id = uuid.uuid4()
            created_price_lists = []
            for uploaded_file in form.cleaned_data['files']:
                file_bytes = uploaded_file.read()
                uploaded_file.seek(0)
                for customer in form.cleaned_data['customers']:
                    price_list = PriceList(
                        customer=customer,
                        uploaded_by=request.user,
                        title=form.cleaned_data.get('title') or '',
                        note=form.cleaned_data.get('note') or '',
                        batch_id=batch_id,
                    )
                    price_list.file.save(uploaded_file.name, ContentFile(file_bytes), save=False)
                    price_list.save()
                    created_price_lists.append(price_list)
            notified_customers = sorted({price_list.customer for price_list in created_price_lists}, key=lambda user: user.id)
            for customer in notified_customers:
                first_file = next(price_list for price_list in created_price_lists if price_list.customer_id == customer.id)
                _notify_users(
                    [customer],
                    'لیست قیمت جدید',
                    f'{len(form.cleaned_data["files"])} فایل لیست قیمت جدید برای شما ثبت شد.',
                    reverse('price_list_file', args=[first_file.id]),
                    category=UserNotification.CATEGORY_SYSTEM,
                    actor=request.user,
                )
            messages.success(
                request,
                f'{len(form.cleaned_data["files"])} فایل لیست قیمت برای {len(notified_customers)} مشتری ثبت شد.'
            )
            return redirect('price_lists')
    else:
        form = PriceListUploadForm(user=request.user)

    records = _price_lists_for_user(request.user)
    customer_filter = (request.GET.get('customer') or '').strip()
    city_filter = (request.GET.get('city') or '').strip()
    province_filter = (request.GET.get('province') or '').strip()
    organization_filter = (request.GET.get('organization') or '').strip()
    if is_staff_user and customer_filter:
        records = records.filter(
            Q(customer__first_name__icontains=customer_filter) |
            Q(customer__last_name__icontains=customer_filter) |
            Q(customer__username__icontains=customer_filter) |
            Q(customer__profile__organization__icontains=customer_filter)
        )
    if is_staff_user and city_filter:
        records = records.filter(customer__profile__city__icontains=city_filter)
    if is_staff_user and province_filter:
        records = records.filter(customer__profile__province__icontains=province_filter)
    if is_staff_user and organization_filter:
        records = records.filter(customer__profile__organization__icontains=organization_filter)
    page_obj = _paginate_queryset(request, records, per_page=10, page_param='page')
    page_base_query = _build_query_string(request, remove_keys=['page'])
    return render(request, 'payments/price_lists.html', {
        'form': form,
        'records': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'is_staff_user': is_staff_user,
        'can_upload_price_lists': can_upload,
        'can_delete_documents': _can_delete_customer_documents(request.user),
        'filters': {
            'customer': customer_filter,
            'city': city_filter,
            'province': province_filter,
            'organization': organization_filter,
        },
        'user_display_name': f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username,
    })


@login_required
def proformas_dashboard(request):
    is_staff_user = _is_staff_user(request.user)
    can_issue = _can_issue_proformas(request.user)

    if request.method == 'POST':
        if not can_issue:
            return HttpResponseForbidden('شما دسترسی صدور پیش فاکتور را ندارید.')
        form = ProformaInvoiceForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            created_proformas = []
            for uploaded_file in form.cleaned_data['files']:
                file_bytes = uploaded_file.read()
                uploaded_file.seek(0)
                for customer in form.cleaned_data['customers']:
                    proforma = ProformaInvoice(
                        customer=customer,
                        issued_by=request.user,
                        title=form.cleaned_data.get('title') or '',
                        valid_until=form.cleaned_data['valid_until'],
                        note=form.cleaned_data.get('note') or '',
                    )
                    proforma.file.save(uploaded_file.name, ContentFile(file_bytes), save=False)
                    proforma.save()
                    created_proformas.append(proforma)
            notified_customers = sorted({proforma.customer for proforma in created_proformas}, key=lambda user: user.id)
            for customer in notified_customers:
                first_proforma = next(proforma for proforma in created_proformas if proforma.customer_id == customer.id)
                _notify_users(
                    [customer],
                    'پیش فاکتور جدید',
                    f'{len(form.cleaned_data["files"])} پیش فاکتور جدید برای شما صادر شد.',
                    reverse('proforma_detail', args=[first_proforma.id]),
                    category=UserNotification.CATEGORY_SYSTEM,
                    actor=request.user,
                )
            messages.success(
                request,
                f'{len(form.cleaned_data["files"])} فایل پیش فاکتور برای {len(notified_customers)} مشتری صادر شد.'
            )
            return redirect('proformas')
    else:
        form = ProformaInvoiceForm(user=request.user)

    records = _proformas_for_user(request.user)
    customer_filter = (request.GET.get('customer') or '').strip()
    if is_staff_user and customer_filter:
        records = records.filter(
            Q(customer__first_name__icontains=customer_filter) |
            Q(customer__last_name__icontains=customer_filter) |
            Q(customer__username__icontains=customer_filter) |
            Q(customer__profile__organization__icontains=customer_filter)
        )
    page_obj = _paginate_queryset(request, records, per_page=10, page_param='page')
    page_base_query = _build_query_string(request, remove_keys=['page'])
    return render(request, 'payments/proformas.html', {
        'form': form,
        'records': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'is_staff_user': is_staff_user,
        'can_issue_proformas': can_issue,
        'can_delete_documents': _can_delete_customer_documents(request.user),
        'filters': {'customer': customer_filter},
        'today_jalali': _today_jalali_date(),
        'user_display_name': f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username,
    })


@login_required
def proforma_detail(request, proforma_id):
    proforma = get_object_or_404(
        ProformaInvoice.objects.select_related('customer', 'customer__profile', 'issued_by').prefetch_related('logs', 'logs__actor'),
        id=proforma_id,
    )
    if not _can_access_proforma(request.user, proforma):
        return HttpResponseForbidden('فقط امکان مشاهده پیش فاکتورهای خودتان وجود دارد.')

    _mark_notifications_read_for_url(request.user, request.path)
    is_staff_user = _is_staff_user(request.user)
    if not is_staff_user:
        if proforma.customer_seen_at is None:
            proforma.customer_seen_at = timezone.now()
            proforma.save(update_fields=['customer_seen_at'])
        _log_proforma(proforma, request.user, ProformaInvoiceLog.ACTION_VIEWED)

    today = _today_jalali_date()
    can_approve = (
        not is_staff_user
        and proforma.customer_id == request.user.id
        and not proforma.is_approved
        and proforma.valid_until >= today
    )
    if request.method == 'POST':
        if request.POST.get('action') != 'approve' or not can_approve:
            return HttpResponseForbidden('امکان تایید این پیش فاکتور وجود ندارد.')
        proforma.status = ProformaInvoice.STATUS_APPROVED
        proforma.approved_at = timezone.now()
        proforma.save(update_fields=['status', 'approved_at'])
        _log_proforma(proforma, request.user, ProformaInvoiceLog.ACTION_APPROVED)

        # اطلاع‌رسانی به صادرکننده پیش‌فاکتور
        notify_targets = []
        if proforma.issued_by_id:
            notify_targets.append(proforma.issued_by)

        # اگر پیش‌فاکتور به سفارشی متصل است
        if proforma.order_id:
            order_ref = proforma.order
            # لاگ تایید در تاریخچه سفارش
            CustomerOrderLog.objects.create(
                order=order_ref,
                actor=request.user,
                action=CustomerOrderLog.ACTION_PROFORMA_APPROVED,
                note=f'پیش فاکتور «{proforma.title or "بدون عنوان"}» توسط مشتری تایید شد.',
            )
            # اطلاع‌رسانی به کارشناس فروش سفارش (اگر متفاوت از صادرکننده است)
            if order_ref.sales_expert_id and order_ref.sales_expert_id != proforma.issued_by_id:
                notify_targets.append(order_ref.sales_expert)

        _notify_users(
            notify_targets,
            'تایید پیش فاکتور',
            f'پیش فاکتور «{proforma.title or proforma.id}» توسط مشتری تایید شد.',
            reverse('proforma_detail', args=[proforma.id]),
            category=UserNotification.CATEGORY_SYSTEM,
            actor=request.user,
        )
        messages.success(request, 'پیش فاکتور با موفقیت تایید شد.')
        return redirect('proforma_detail', proforma_id=proforma.id)

    logs = [
        {
            'action': log.get_action_display(),
            'actor': log.actor.get_full_name() or log.actor.username if log.actor else 'سیستم',
            'time': _format_jalali_datetime(log.created_at),
            'note': log.note,
        }
        for log in proforma.logs.all()
    ] if is_staff_user else []
    return render(request, 'payments/proforma_detail.html', {
        'proforma': proforma,
        'is_staff_user': is_staff_user,
        'can_approve': can_approve,
        'can_delete_documents': _can_delete_customer_documents(request.user),
        'is_expired': proforma.valid_until < today,
        'logs': logs,
        'return_url': _safe_next_url(request, default=reverse('proformas')),
        'return_label': _return_link_label(request, 'بازگشت'),
    })


@login_required
def invoice_detail(request, invoice_id):
    invoice = get_object_or_404(
        InvoiceRecord.objects.select_related('customer', 'customer__profile', 'uploaded_by'),
        id=invoice_id,
    )
    is_staff_user = _is_staff_user(request.user)
    just_marked_seen = False
    return_url = _safe_next_url(request)
    return_label = _return_link_label(request, 'بازگشت به فاکتورها')

    # Staff needs permission to view invoices
    if is_staff_user and not _can_view_invoices(request.user):
        return HttpResponseForbidden('شما دسترسی مشاهده فاکتور را ندارید.')

    if is_staff_user and not _can_staff_access_customer(request.user, invoice.customer_id):
        return HttpResponseForbidden('امکان مشاهده فاکتور این مشتری برای شما وجود ندارد.')

    # Customers can only see their own invoices
    if not is_staff_user and invoice.customer_id != request.user.id:
        return HttpResponseForbidden('فقط امکان مشاهده فاکتورهای خودتان وجود دارد.')

    _mark_notifications_read_for_url(request.user, request.path)
    if not is_staff_user and invoice.customer_seen_at is None:
        invoice.customer_seen_at = timezone.now()
        invoice.save(update_fields=['customer_seen_at'])
        just_marked_seen = True

    if request.method == 'POST':
        if is_staff_user:
            return HttpResponseForbidden('ثبت یادداشت فقط برای مشتری فعال است.')
        previous_note = (invoice.customer_note or '').strip()
        form = InvoiceCustomerNoteForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            new_note = (form.cleaned_data.get('customer_note') or '').strip()
            if new_note and new_note != previous_note:
                _notify_invoice_customer_note(invoice, request.user)
            messages.success(request, 'یادداشت شما ذخیره شد.')
            redirect_url = request.path
            if return_url:
                redirect_url = f"{redirect_url}?{urlencode({'next': return_url})}"
            return redirect(redirect_url)
    else:
        form = InvoiceCustomerNoteForm(instance=invoice)

    customer_profile = getattr(invoice.customer, 'profile', None)
    return render(request, 'payments/invoice_detail.html', {
        'invoice': invoice,
        'form': form,
        'is_staff_user': is_staff_user,
        'can_delete_documents': _can_delete_customer_documents(request.user) and _can_view_invoices(request.user),
        'customer_profile': customer_profile,
        'just_marked_seen': just_marked_seen,
        'return_url': return_url,
        'return_label': return_label,
    })


@login_required
def invoice_file(request, invoice_id):
    invoice = get_object_or_404(InvoiceRecord, id=invoice_id)
    if not _can_staff_access_customer(request.user, invoice.customer_id):
        return HttpResponseForbidden('امکان حذف فاکتور این مشتری برای شما وجود ندارد.')
    if not _can_access_invoice(request.user, invoice):
        return HttpResponseForbidden('فقط امکان مشاهده فایل فاکتورهای خودتان وجود دارد.')
    _mark_notifications_read_for_url(request.user, request.path)
    return _file_response(invoice.attachment, as_attachment=request.GET.get('download') == '1')


@login_required
def price_list_file(request, price_list_id):
    price_list = get_object_or_404(PriceList.objects.select_related('customer'), id=price_list_id)
    if _is_staff_user(request.user):
        if not _can_staff_access_customer(request.user, price_list.customer_id):
            return HttpResponseForbidden('امکان مشاهده لیست قیمت این مشتری برای شما وجود ندارد.')
        return _file_response(price_list.file, as_attachment=request.GET.get('download') == '1')

    latest = PriceList.objects.filter(customer=request.user).order_by('-created_at', '-id').first()
    if not latest or price_list.customer_id != request.user.id or latest.batch_id != price_list.batch_id:
        return HttpResponseForbidden('فقط امکان مشاهده آخرین لیست قیمت خودتان وجود دارد.')
    _mark_notifications_read_for_url(request.user, request.path)
    if price_list.customer_seen_at is None:
        price_list.customer_seen_at = timezone.now()
        price_list.save(update_fields=['customer_seen_at'])
    return _file_response(price_list.file, as_attachment=request.GET.get('download') == '1')


@login_required
def proforma_file(request, proforma_id):
    proforma = get_object_or_404(ProformaInvoice.objects.select_related('customer'), id=proforma_id)
    if not _can_access_proforma(request.user, proforma):
        return HttpResponseForbidden('فقط امکان مشاهده فایل پیش فاکتورهای خودتان وجود دارد.')
    _mark_notifications_read_for_url(request.user, request.path)
    if not _is_staff_user(request.user):
        if proforma.customer_seen_at is None:
            proforma.customer_seen_at = timezone.now()
            proforma.save(update_fields=['customer_seen_at'])
        _log_proforma(proforma, request.user, ProformaInvoiceLog.ACTION_FILE_VIEWED)
    return _file_response(proforma.file, as_attachment=request.GET.get('download') == '1')


def _document_delete_redirect(request, fallback):
    next_url = request.POST.get('next') or request.GET.get('next') or request.META.get('HTTP_REFERER') or reverse(fallback)
    if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return reverse(fallback)
    return next_url


@login_required
@require_POST
def invoice_delete(request, invoice_id):
    if not _can_delete_customer_documents(request.user) or not _can_view_invoices(request.user):
        return HttpResponseForbidden('شما دسترسی حذف فاکتور را ندارید.')
    invoice = get_object_or_404(InvoiceRecord, id=invoice_id)
    _delete_file_field(invoice.attachment)
    invoice.delete()
    messages.success(request, 'فاکتور حذف شد.')
    return redirect(_document_delete_redirect(request, 'invoices_dashboard'))


@login_required
@require_POST
def price_list_delete(request, price_list_id):
    if not _can_delete_customer_documents(request.user):
        return HttpResponseForbidden('شما دسترسی حذف لیست قیمت را ندارید.')
    price_list = get_object_or_404(PriceList, id=price_list_id)
    if not _can_staff_access_customer(request.user, price_list.customer_id):
        return HttpResponseForbidden('امکان حذف لیست قیمت این مشتری برای شما وجود ندارد.')
    _delete_file_field(price_list.file)
    price_list.delete()
    messages.success(request, 'لیست قیمت حذف شد.')
    return redirect(_document_delete_redirect(request, 'price_lists'))


@login_required
@require_POST
def proforma_delete(request, proforma_id):
    if not _can_delete_customer_documents(request.user):
        return HttpResponseForbidden('شما دسترسی حذف پیش فاکتور را ندارید.')
    proforma = get_object_or_404(ProformaInvoice, id=proforma_id)
    if not _can_staff_access_customer(request.user, proforma.customer_id):
        return HttpResponseForbidden('امکان حذف پیش فاکتور این مشتری برای شما وجود ندارد.')
    _delete_file_field(proforma.file)
    proforma.delete()
    messages.success(request, 'پیش فاکتور حذف شد.')
    return redirect(_document_delete_redirect(request, 'proformas'))


@login_required
def orders_dashboard(request):
    if not _can_view_orders(request.user):
        return HttpResponseForbidden('شما دسترسی مشاهده سفارش ها را ندارید.')

    is_staff_user = _is_staff_user(request.user)
    order_form = None
    item_formset = None
    can_create_order = False

    if not is_staff_user:
        try:
            can_create_order = not request.user.profile.suspended
        except UserProfile.DoesNotExist:
            can_create_order = False
        if request.method == 'POST' and not can_create_order:
            return HttpResponseForbidden('حساب شما غیرفعال است و امکان ثبت سفارش جدید وجود ندارد.')
        assigned_sales_expert = None
        try:
            assigned_sales_expert = request.user.sales_assignment.sales_user
        except (CustomerSalesAssignment.DoesNotExist, AttributeError):
            pass

        if request.method == 'POST':
            order_form = CustomerOrderForm(request.POST)
            item_formset = CustomerOrderItemFormSet(request.POST)
            if order_form.is_valid() and item_formset.is_valid():
                order = order_form.save(commit=False)
                order.customer = request.user
                order.requested_sales_expert = assigned_sales_expert
                order.sales_expert = assigned_sales_expert
                order.save()
                item_formset.instance = order
                item_formset.save()
                CustomerOrderLog.objects.create(order=order, actor=request.user, action=CustomerOrderLog.ACTION_CREATED, to_status=order.status, note=order.customer_note)
                notify_users = [assigned_sales_expert] if assigned_sales_expert else list(_staff_notification_users({'sales', 'commercial'}))
                _notify_users(
                    notify_users,
                    'سفارش جدید مشتری',
                    f'سفارش {order.order_number} توسط {request.user.get_full_name() or request.user.username} ثبت شد.',
                    reverse('order_detail', args=[order.id]),
                    category=UserNotification.CATEGORY_SYSTEM,
                    actor=request.user,
                )
                messages.success(request, 'سفارش شما با موفقیت ثبت شد.')
                return redirect('orders')
        else:
            if can_create_order:
                order_form = CustomerOrderForm()
                item_formset = CustomerOrderItemFormSet()

    records = _orders_for_user(request.user)
    records, filters = _apply_order_filters(records, request, is_staff_user=is_staff_user)
    status_summary = [
        {'status': status, 'label': label, 'count': records.filter(status=status).count()}
        for status, label in CustomerOrder.STATUS_CHOICES
    ]
    page_obj = _paginate_queryset(request, records, per_page=10, page_param='page')
    page_base_query = _build_query_string(request, remove_keys=['page'])

    assigned_sales_expert_ctx = None
    if not is_staff_user:
        try:
            assigned_sales_expert_ctx = request.user.sales_assignment.sales_user
        except (CustomerSalesAssignment.DoesNotExist, AttributeError):
            pass

    return render(request, 'payments/orders.html', {
        'order_form': order_form,
        'item_formset': item_formset,
        'records': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'filters': filters,
        'status_choices': CustomerOrder.STATUS_CHOICES,
        'status_summary': status_summary,
        'is_staff_user': is_staff_user,
        'can_create_order': can_create_order,
        'can_manage_orders': _can_manage_orders(request.user),
        'assigned_sales_expert': assigned_sales_expert_ctx,
        'export_dataset': 'orders',
        'export_fields': ORDER_EXPORT_FIELDS,
    })


@login_required
def sales_expert_dashboard(request):
    role = _user_role(request.user)
    if role not in {'sales', 'sales_manager'} and not request.user.is_superuser:
        return HttpResponseForbidden('این بخش فقط برای کارشناسان و مدیر فروش فعال است.')

    is_manager = request.user.is_superuser or role == 'sales_manager'

    # مشتریان تخصیص‌یافته به این کارشناس (یا همه برای مدیر)
    if is_manager:
        assignments = (
            CustomerSalesAssignment.objects
            .select_related('customer', 'customer__profile', 'sales_user')
            .filter(sales_user__isnull=False)
            .order_by('sales_user__first_name', 'customer__username')
        )
    else:
        assignments = (
            CustomerSalesAssignment.objects
            .select_related('customer', 'customer__profile')
            .filter(sales_user=request.user)
            .order_by('customer__username')
        )

    assigned_ids = [a.customer_id for a in assignments]
    open_statuses = [CustomerOrder.STATUS_SUBMITTED, CustomerOrder.STATUS_REVIEWING, CustomerOrder.STATUS_PROFORMA_SENT]

    open_orders_qs = (
        CustomerOrder.objects
        .filter(customer_id__in=assigned_ids, status__in=open_statuses)
        .select_related('customer', 'customer__profile', 'sales_expert')
        .order_by('-created_at')
    )

    # آمار سریع
    open_order_counts = {
        row['customer']: row['cnt']
        for row in CustomerOrder.objects
            .filter(customer_id__in=assigned_ids, status__in=open_statuses)
            .values('customer')
            .annotate(cnt=Count('id'))
    }
    last_order_dates = {
        row['customer']: row['last']
        for row in CustomerOrder.objects
            .filter(customer_id__in=assigned_ids)
            .values('customer')
            .annotate(last=Max('created_at'))
    }

    # ترکیب داده مشتریان
    customer_rows = []
    for a in assignments:
        uid = a.customer_id
        customer_rows.append({
            'customer': a.customer,
            'profile': a.customer.profile if hasattr(a.customer, 'profile') else None,
            'sales_user': a.sales_user if is_manager else None,
            'open_orders': open_order_counts.get(uid, 0),
            'last_order': last_order_dates.get(uid),
        })
    customer_rows.sort(key=lambda r: -(r['open_orders']))

    return render(request, 'payments/sales_expert_dashboard.html', {
        'is_manager': is_manager,
        'assigned_count': len(assigned_ids),
        'open_orders_count': sum(open_order_counts.values()),
        'open_orders': open_orders_qs[:15],
        'customer_rows': customer_rows,
        'is_staff_user': True,
    })


@login_required
def sales_assignments_dashboard(request):
    if not _can_manage_sales_assignments(request.user):
        return HttpResponseForbidden('این بخش فقط برای مدیر فروش و مدیر سیستم فعال است.')

    if request.method == 'POST':
        form = SalesAssignmentBulkForm(request.POST)
        if form.is_valid():
            customers = form.cleaned_data['customers']
            sales_user = form.cleaned_data['sales_user']
            note = form.cleaned_data.get('note') or ''

            # جمع‌آوری کارشناسان قدیمی برای اطلاع‌رسانی
            old_experts_map = {
                ca.customer_id: ca.sales_user
                for ca in CustomerSalesAssignment.objects
                    .filter(customer__in=customers, sales_user__isnull=False)
                    .select_related('sales_user')
            }

            for customer in customers:
                CustomerSalesAssignment.objects.update_or_create(
                    customer=customer,
                    defaults={'sales_user': sales_user, 'assigned_by': request.user, 'note': note},
                )
            if form.cleaned_data.get('transfer_open_orders'):
                updated_orders = (
                    CustomerOrder.objects
                    .filter(customer__in=customers)
                    .exclude(status__in=[CustomerOrder.STATUS_COMPLETED, CustomerOrder.STATUS_CANCELLED])
                )
                for order in updated_orders:
                    old_sales_id = order.sales_expert_id
                    order.sales_expert = sales_user
                    order.save(update_fields=['sales_expert', 'updated_at'])
                    if old_sales_id != sales_user.id:
                        CustomerOrderLog.objects.create(
                            order=order,
                            actor=request.user,
                            action=CustomerOrderLog.ACTION_ASSIGNED,
                            note=f'تفویض توسط مدیر فروش به {sales_user.get_full_name() or sales_user.username}',
                        )

            # اطلاع‌رسانی به کارشناس جدید
            _notify_users(
                [sales_user],
                'تخصیص مشتریان',
                f'{len(customers)} مشتری به شما تخصیص داده شد.',
                reverse('sales_expert_dashboard'),
                category=UserNotification.CATEGORY_SYSTEM,
                actor=request.user,
            )

            # اطلاع‌رسانی به کارشناسان قدیمی که مشتریانشان منتقل شد
            displaced_experts = {}
            for customer in customers:
                old_exp = old_experts_map.get(customer.id)
                if old_exp and old_exp.id != sales_user.id:
                    displaced_experts.setdefault(old_exp.id, {'user': old_exp, 'count': 0})
                    displaced_experts[old_exp.id]['count'] += 1
            for entry in displaced_experts.values():
                _notify_users(
                    [entry['user']],
                    'انتقال مشتریان',
                    f'{entry["count"]} مشتری از لیست شما به {sales_user.get_full_name() or sales_user.username} منتقل شد.',
                    reverse('sales_expert_dashboard'),
                    category=UserNotification.CATEGORY_SYSTEM,
                    actor=request.user,
                )

            messages.success(request, f'{len(customers)} مشتری به {sales_user.get_full_name() or sales_user.username} تخصیص داده شد.')
            return redirect('sales_assignments')
    else:
        form = SalesAssignmentBulkForm()

    rows, filters = _sales_assignment_rows(request)
    page_obj = _paginate_queryset(request, rows, per_page=15, page_param='page')
    page_base_query = _build_query_string(request, remove_keys=['page'])
    assigned_count = sum(1 for row in rows if row['sales_user'])

    # آمار تجمیعی به تفکیک کارشناس فروش
    expert_stats_map = {}
    for row in rows:
        su = row['sales_user']
        if not su:
            continue
        if su.id not in expert_stats_map:
            expert_stats_map[su.id] = {'user': su, 'customers': 0, 'open_orders': 0}
        expert_stats_map[su.id]['customers'] += 1
        expert_stats_map[su.id]['open_orders'] += row['open_orders']
    expert_stats = sorted(expert_stats_map.values(), key=lambda x: -x['customers'])

    return render(request, 'payments/sales_assignments.html', {
        'form': form,
        'records': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'filters': filters,
        'assigned_count': assigned_count,
        'unassigned_count': len(rows) - assigned_count,
        'expert_stats': expert_stats,
        'export_dataset': 'sales_assignments',
        'export_fields': SALES_ASSIGNMENT_EXPORT_FIELDS,
    })


@login_required
def order_detail(request, order_id):
    order = get_object_or_404(
        CustomerOrder.objects.select_related('customer', 'customer__profile', 'sales_expert', 'requested_sales_expert').prefetch_related('items', 'logs', 'logs__actor', 'proformas'),
        id=order_id,
    )
    if not _orders_for_user(request.user).filter(id=order.id).exists():
        return HttpResponseForbidden('امکان مشاهده این سفارش برای شما وجود ندارد.')

    _mark_notifications_read_for_url(request.user, request.path)
    can_manage = _can_manage_orders(request.user)
    status_form = StaffOrderUpdateForm(instance=order) if can_manage else None
    proforma_form = OrderProformaUploadForm() if can_manage else None

    if request.method == 'POST':
        if not can_manage:
            return HttpResponseForbidden('امکان تغییر سفارش برای نقش شما فعال نیست.')
        action = request.POST.get('action')
        if action == 'update_status':
            old_status = order.status
            old_sales_id = order.sales_expert_id
            status_form = StaffOrderUpdateForm(request.POST, instance=order)
            if status_form.is_valid():
                updated = status_form.save()
                if updated.sales_expert_id:
                    CustomerSalesAssignment.objects.update_or_create(
                        customer=updated.customer,
                        defaults={'sales_user': updated.sales_expert, 'assigned_by': request.user, 'note': 'تخصیص از صفحه سفارش'},
                    )
                if old_status != updated.status:
                    CustomerOrderLog.objects.create(order=updated, actor=request.user, action=CustomerOrderLog.ACTION_STATUS_CHANGED, from_status=old_status, to_status=updated.status, note=updated.staff_note)
                    _notify_users([updated.customer], 'تغییر وضعیت سفارش', f'وضعیت سفارش {updated.order_number} به «{updated.get_status_display()}» تغییر کرد.', reverse('order_detail', args=[updated.id]), category=UserNotification.CATEGORY_SYSTEM, actor=request.user)
                if old_sales_id != updated.sales_expert_id:
                    assignee = updated.sales_expert.get_full_name() or updated.sales_expert.username if updated.sales_expert else '-'
                    CustomerOrderLog.objects.create(order=updated, actor=request.user, action=CustomerOrderLog.ACTION_ASSIGNED, note=f'تخصیص به {assignee}')
                    if updated.sales_expert_id:
                        _notify_users([updated.sales_expert], 'تخصیص سفارش', f'سفارش {updated.order_number} به شما تخصیص داده شد.', reverse('order_detail', args=[updated.id]), category=UserNotification.CATEGORY_SYSTEM, actor=request.user)
                messages.success(request, 'سفارش بروزرسانی شد.')
                return redirect('order_detail', order_id=updated.id)
        elif action == 'issue_proforma':
            proforma_form = OrderProformaUploadForm(request.POST, request.FILES)
            if proforma_form.is_valid():
                created = []
                for uploaded_file in proforma_form.cleaned_data['files']:
                    proforma = ProformaInvoice(
                        customer=order.customer,
                        order=order,
                        issued_by=request.user,
                        title=proforma_form.cleaned_data.get('title') or order.title or order.order_number,
                        valid_until=proforma_form.cleaned_data['valid_until'],
                        note=proforma_form.cleaned_data.get('note') or '',
                    )
                    proforma.file.save(uploaded_file.name, uploaded_file, save=False)
                    proforma.save()
                    created.append(proforma)
                previous_status = order.status
                order.status = CustomerOrder.STATUS_PROFORMA_SENT
                if not order.sales_expert_id:
                    order.sales_expert = request.user
                order.save(update_fields=['status', 'sales_expert', 'updated_at'])
                CustomerOrderLog.objects.create(order=order, actor=request.user, action=CustomerOrderLog.ACTION_PROFORMA_CREATED, from_status=previous_status, to_status=order.status, note=f'{len(created)} پیش فاکتور صادر شد.')
                _notify_users([order.customer], 'پیش فاکتور سفارش صادر شد', f'{len(created)} پیش فاکتور برای سفارش {order.order_number} صادر شد.', reverse('order_detail', args=[order.id]), category=UserNotification.CATEGORY_SYSTEM, actor=request.user)
                messages.success(request, 'پیش فاکتور سفارش صادر و به مشتری اطلاع رسانی شد.')
                return redirect('order_detail', order_id=order.id)

    is_staff = _is_staff_user(request.user)
    proformas = list(order.proformas.all())
    has_approved_proforma = any(p.is_approved for p in proformas)
    all_proformas_approved = bool(proformas) and all(p.is_approved for p in proformas)
    pending_proformas = [p for p in proformas if not p.is_approved]
    today = _today_jalali_date()
    expired_proformas = [p for p in proformas if not p.is_approved and p.valid_until < today]

    # پیشنهاد تکمیل سفارش برای کارمند وقتی پیش‌فاکتور تایید شده اما سفارش هنوز تکمیل نشده
    suggest_completion = (
        is_staff and can_manage
        and has_approved_proforma
        and order.status not in [CustomerOrder.STATUS_COMPLETED, CustomerOrder.STATUS_CANCELLED]
    )

    return render(request, 'payments/order_detail.html', {
        'order': order,
        'status_form': status_form,
        'proforma_form': proforma_form,
        'is_staff_user': is_staff,
        'can_manage_orders': can_manage,
        'return_url': _safe_next_url(request, default=reverse('orders')),
        'return_label': _return_link_label(request, 'بازگشت'),
        'logs': order.logs.all() if is_staff else order.logs.exclude(action=CustomerOrderLog.ACTION_ASSIGNED),
        'proformas': proformas,
        'has_approved_proforma': has_approved_proforma,
        'all_proformas_approved': all_proformas_approved,
        'pending_proformas_count': len(pending_proformas),
        'expired_proformas_count': len(expired_proformas),
        'suggest_completion': suggest_completion,
        'today': today,
    })


@login_required
def receipt_file(request, receipt_id):
    receipt = get_object_or_404(PaymentReceipt.objects.select_related('payment'), id=receipt_id)
    if not _can_access_payment(request.user, receipt.payment):
        return HttpResponseForbidden('فقط امکان مشاهده فایل فیش‌های خودتان وجود دارد.')
    _log_activity(receipt.payment, request.user, PaymentActivityLog.ACTION_VIEWED, note='مشاهده فایل فیش')
    return _file_response(receipt.image)


@login_required
@require_POST
def rotate_receipt(request, receipt_id):
    if not _is_staff_user(request.user):
        return JsonResponse({'error': 'دسترسی ممنوع'}, status=403)

    receipt = get_object_or_404(PaymentReceipt.objects.select_related('payment'), id=receipt_id)
    if not _can_access_payment(request.user, receipt.payment):
        return JsonResponse({'error': 'دسترسی ممنوع'}, status=403)

    try:
        degrees = float(request.POST.get('degrees', 0))
    except (ValueError, TypeError):
        return JsonResponse({'error': 'مقدار درجه نامعتبر است'}, status=400)

    if degrees == 0:
        return JsonResponse({'success': True})
    if not (-360 < degrees < 360):
        return JsonResponse({'error': 'درجه چرخش باید بین ‎-359 تا 359 باشد'}, status=400)

    ext = os.path.splitext(receipt.image.name)[1].lower()
    if ext not in ('.jpg', '.jpeg', '.png', '.webp', '.bmp'):
        return JsonResponse({'error': 'این نوع فایل قابل چرخش نیست'}, status=400)

    try:
        from PIL import Image, ImageOps
        path = receipt.image.path
        img = Image.open(path)
        img = ImageOps.exif_transpose(img)

        if ext in ('.jpg', '.jpeg') and img.mode in ('RGBA', 'P', 'LA'):
            img = img.convert('RGB')

        # degrees مثبت = راست (CW) از دید کاربر → در Pillow (CCW) باید نگیت شود
        pil_angle = -degrees
        fill = (255, 255, 255, 0) if img.mode == 'RGBA' else (255, 255, 255) if img.mode in ('RGB', 'L') else None
        img = img.rotate(pil_angle, expand=True, fillcolor=fill)

        fmt = 'JPEG' if ext in ('.jpg', '.jpeg') else (img.format or 'PNG')
        save_kwargs = {'quality': 90, 'optimize': True, 'exif': b''} if fmt == 'JPEG' else {}
        img.save(path, format=fmt, **save_kwargs)

        direction_label = 'راست' if degrees > 0 else 'چپ'
        _log_activity(receipt.payment, request.user, PaymentActivityLog.ACTION_EDITED,
                      note=f'چرخش تصویر {abs(degrees):.1f}° {direction_label}')
        return JsonResponse({'success': True})

    except Exception:
        logger.exception('Failed to rotate receipt %s', receipt_id)
        return JsonResponse({'error': 'خطا در چرخش تصویر'}, status=500)


@login_required
def legacy_payment_receipt_file(request, payment_id):
    payment = get_object_or_404(PaymentRecord, id=payment_id)
    if not _can_access_payment(request.user, payment):
        return HttpResponseForbidden('فقط امکان مشاهده فایل فیش‌های خودتان وجود دارد.')
    _log_activity(payment, request.user, PaymentActivityLog.ACTION_VIEWED, note='مشاهده فایل فیش')
    return _file_response(payment.receipt_image)


@login_required
def product_catalog_search(request):
    q = (request.GET.get('q') or '').strip()
    if len(q) < 1:
        return JsonResponse({'results': []})
    results = list(
        ProductCatalog.objects
        .filter(is_active=True)
        .filter(Q(product_name__icontains=q) | Q(product_code__icontains=q))
        .values('product_name', 'product_code', 'unit', 'coefficient')[:20]
    )
    for r in results:
        r['coefficient'] = str(r['coefficient']) if r['coefficient'] is not None else ''
    return JsonResponse({'results': results})


@login_required
def import_product_catalog(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden('فقط مدیران سیستم امکان ایمپورت دارند.')
    if request.method != 'POST':
        return redirect('admin:payments_productcatalog_changelist')

    file = request.FILES.get('file')
    if not file:
        messages.error(request, 'فایلی انتخاب نشده است.')
        return redirect('admin:payments_productcatalog_changelist')

    try:
        import openpyxl
        from decimal import Decimal, InvalidOperation

        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messages.error(request, 'فایل خالی است.')
            return redirect('admin:payments_productcatalog_changelist')

        headers = [str(c or '').strip() for c in rows[0]]

        def _col(candidates):
            for name in candidates:
                for i, h in enumerate(headers):
                    if h.lower() == name.lower():
                        return i
            return None

        name_col  = _col(['نام کالا', 'product_name', 'نام'])
        code_col  = _col(['کد کالا', 'product_code', 'کد'])
        unit_col  = _col(['واحد', 'unit'])
        coeff_col = _col(['ضریب', 'coefficient', 'ضريب'])

        if name_col is None:
            messages.error(request, 'ستون «نام کالا» در فایل یافت نشد.')
            return redirect('admin:payments_productcatalog_changelist')

        added = 0
        for row in rows[1:]:
            def cell(idx):
                return str(row[idx] or '').strip() if idx is not None and idx < len(row) else ''

            name = cell(name_col)
            if not name:
                continue
            coeff = None
            if coeff_col is not None and coeff_col < len(row) and row[coeff_col] is not None:
                try:
                    coeff = Decimal(str(row[coeff_col]))
                except InvalidOperation:
                    pass

            ProductCatalog.objects.create(
                product_name=name,
                product_code=cell(code_col),
                unit=cell(unit_col),
                coefficient=coeff,
            )
            added += 1

        messages.success(request, f'{added} کالا با موفقیت به کاتالوگ اضافه شد.')

    except Exception:
        logger.exception('Product catalog import failed')
        messages.error(request, 'خطا در پردازش فایل اکسل.')

    return redirect('admin:payments_productcatalog_changelist')


@login_required
def counterparty_dashboard(request):
    cp = _get_user_counterparty(request.user)
    if not cp:
        return HttpResponseForbidden('این بخش فقط برای طرف حساب‌های ثبت‌شده فعال است.')

    payments = (
        PaymentRecord.objects
        .filter(counterparty=cp)
        .select_related('user', 'user__profile')
        .prefetch_related('receipts')
        .order_by('-created_at', '-id')
    )

    # فیلترها
    status_filter = (request.GET.get('status') or '').strip()
    approved_filter = (request.GET.get('approved') or '').strip()
    if status_filter:
        payments = payments.filter(status=status_filter)
    if approved_filter == 'approved':
        payments = payments.filter(counterparty_status=PaymentRecord.CP_STATUS_APPROVED)
    elif approved_filter == 'returned':
        payments = payments.filter(counterparty_status=PaymentRecord.CP_STATUS_RETURNED)
    elif approved_filter == 'rejected':
        payments = payments.filter(counterparty_status=PaymentRecord.CP_STATUS_REJECTED)
    elif approved_filter == 'pending':
        payments = payments.filter(counterparty_status__isnull=True)

    # خروجی اکسل
    if request.GET.get('export') == 'excel':
        _log_counterparty_action(cp, request.user, 'دریافت خروجی اکسل')
        fields = [
            _field('id', 'شماره فیش', lambda p: p.id),
            _field('customer', 'مشتری', lambda p: f"{p.first_name} {p.last_name}".strip() or (p.user.get_full_name() if p.user else '')),
            _field('organization', 'مجموعه', lambda p: p.organization),
            _field('amount', 'مبلغ (ریال)', lambda p: p.amount),
            _field('pay_date', 'تاریخ واریز', lambda p: _format_jalali_date(p.pay_date)),
            _field('tracking_code', 'کد پیگیری', lambda p: p.tracking_code or ''),
            _field('payer_full_name', 'نام واریز کننده', lambda p: p.payer_full_name),
            _field('payer_account_number', 'شماره حساب واریز کننده', lambda p: p.payer_account_number),
            _field('payer_bank_name', 'بانک واریز کننده', lambda p: p.payer_bank_name),
            _field('beneficiary_account_number', 'شماره حساب مقصد', lambda p: p.beneficiary_account_number),
            _field('status', 'وضعیت', lambda p: p.get_status_display()),
            _field('cp_approved', 'تایید طرف حساب', lambda p: _format_jalali_datetime(p.counterparty_decided_at) if p.counterparty_decided_at and p.counterparty_status == PaymentRecord.CP_STATUS_APPROVED else 'تایید نشده'),
            _field('created_at', 'تاریخ ثبت', lambda p: _format_jalali_datetime(p.created_at)),
        ]
        return _export_response(
            _timestamped_excel_filename('counterparty_payments.xlsx'),
            'فیش‌های طرف حساب',
            fields,
            list(payments),
        )

    page_obj = _paginate_queryset(request, payments, per_page=20)
    page_base_query = _build_query_string(request, remove_keys=['page'])

    total = payments.count()
    approved_count = payments.filter(counterparty_status=PaymentRecord.CP_STATUS_APPROVED).count()
    returned_count = payments.filter(counterparty_status=PaymentRecord.CP_STATUS_RETURNED).count()
    rejected_count = payments.filter(counterparty_status=PaymentRecord.CP_STATUS_REJECTED).count()

    # ثبت مشاهده داشبورد
    _log_counterparty_action(cp, request.user, 'مشاهده داشبورد')

    return render(request, 'payments/counterparty_dashboard.html', {
        'cp': cp,
        'records': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'total': total,
        'approved_count': approved_count,
        'returned_count': returned_count,
        'rejected_count': rejected_count,
        'pending_count': total - approved_count - returned_count - rejected_count,
        'status_choices': PaymentRecord.STATUS_CHOICES,
        'filters': {'status': status_filter, 'approved': approved_filter},
        'can_operate': cp.can_operate,
        'cp_status': cp.status,
    })


def _cp_check_action(request, payment_id):
    """بررسی دسترسی طرف حساب برای اقدام روی یک فیش."""
    cp = _get_user_counterparty(request.user)
    if not cp:
        return None, None, HttpResponseForbidden('دسترسی ممنوع.')
    if not cp.can_operate:
        return None, None, HttpResponseForbidden('حساب شما غیرفعال است.')
    payment = get_object_or_404(
        PaymentRecord.objects.select_related('counterparty', 'user'),
        id=payment_id, counterparty=cp,
    )
    if payment.counterparty_decided:
        messages.error(request, 'تصمیم قبلی برای این فیش ثبت شده و قابل تغییر نیست.')
        return None, None, redirect('counterparty_dashboard')
    return cp, payment, None


@login_required
@require_POST
def counterparty_approve_payment(request, payment_id):
    """طرف حساب فیش را تایید می‌کند — توضیح اختیاری."""
    cp, payment, err = _cp_check_action(request, payment_id)
    if err:
        return err

    note = (request.POST.get('note') or '').strip()
    now = timezone.now()

    payment.counterparty_status = PaymentRecord.CP_STATUS_APPROVED
    payment.counterparty_note = note
    payment.counterparty_decided_at = now
    payment.counterparty_decided_by = request.user
    _sync_pending_final_flag(payment)
    payment.save(update_fields=[
        'counterparty_status', 'counterparty_note',
        'counterparty_decided_at', 'counterparty_decided_by',
        'pending_final_approval', 'pending_final_approval_since',
    ])

    _log_activity(payment, request.user, PaymentActivityLog.ACTION_CP_APPROVED,
                  note=note or f'تایید توسط طرف حساب: {cp.name}')
    _log_counterparty_action(cp, request.user, f'تایید فیش #{payment_id}')

    _notify_users(
        list(_staff_notification_users({'commercial', 'commercial_manager'})),
        '✅ تایید فیش توسط طرف حساب',
        f'فیش #{payment_id} توسط «{cp.name}» تایید شد.' + (f' توضیح: {note}' if note else ''),
        reverse('payment_timeline', args=[payment.id]), category=UserNotification.CATEGORY_PAYMENT, actor=request.user,
    )
    messages.success(request, f'✅ فیش #{payment_id} با موفقیت تایید شد.')
    return redirect('counterparty_dashboard')


@login_required
@require_POST
def counterparty_return_payment_cp(request, payment_id):
    """طرف حساب فیش را به دلیل نقص عودت می‌دهد — توضیح اجباری."""
    cp, payment, err = _cp_check_action(request, payment_id)
    if err:
        return err

    note = (request.POST.get('note') or '').strip()
    if not note:
        messages.error(request, 'ثبت توضیح برای عودت اجباری است.')
        return redirect('counterparty_dashboard')

    old_status = payment.status
    payment.counterparty_status = PaymentRecord.CP_STATUS_RETURNED
    payment.counterparty_note = note
    payment.counterparty_decided_at = timezone.now()
    payment.counterparty_decided_by = request.user
    payment.status = PaymentRecord.STATUS_COMMERCIAL_REVIEW  # برگشت به صف بازرگانی
    payment.save(update_fields=[
        'counterparty_status', 'counterparty_note',
        'counterparty_decided_at', 'counterparty_decided_by', 'status',
    ])

    _log_activity(payment, request.user, PaymentActivityLog.ACTION_CP_RETURNED,
                  from_status=old_status, to_status=payment.status,
                  note=f'عودت توسط طرف حساب: {note}')
    _log_counterparty_action(cp, request.user, f'عودت فیش #{payment_id}: {note}')

    _notify_users(
        list(_staff_notification_users({'commercial', 'commercial_manager'})),
        '⚠ عودت فیش از طرف حساب',
        f'فیش #{payment_id} توسط «{cp.name}» عودت داده شد. دلیل: {note}',
        reverse('payment_timeline', args=[payment.id]), category=UserNotification.CATEGORY_PAYMENT, actor=request.user,
    )
    messages.warning(request, f'⚠ فیش #{payment_id} به بازرگانی عودت داده شد.')
    return redirect('counterparty_dashboard')


@login_required
@require_POST
def counterparty_reject_payment_cp(request, payment_id):
    """طرف حساب فیش را رد/ابطال می‌کند — توضیح اجباری."""
    cp, payment, err = _cp_check_action(request, payment_id)
    if err:
        return err

    note = (request.POST.get('note') or '').strip()
    if not note:
        messages.error(request, 'ثبت توضیح برای رد/ابطال اجباری است.')
        return redirect('counterparty_dashboard')

    payment.counterparty_status = PaymentRecord.CP_STATUS_REJECTED
    payment.counterparty_note = note
    payment.counterparty_decided_at = timezone.now()
    payment.counterparty_decided_by = request.user
    payment.save(update_fields=[
        'counterparty_status', 'counterparty_note',
        'counterparty_decided_at', 'counterparty_decided_by',
    ])

    _log_activity(payment, request.user, PaymentActivityLog.ACTION_CP_REJECTED,
                  note=f'رد/ابطال توسط طرف حساب: {note}')
    _log_counterparty_action(cp, request.user, f'رد فیش #{payment_id}: {note}')

    _notify_users(
        list(_staff_notification_users({'commercial', 'commercial_manager'})),
        '🚫 رد فیش توسط طرف حساب',
        f'فیش #{payment_id} توسط «{cp.name}» رد/ابطال شد. دلیل: {note}',
        reverse('payment_timeline', args=[payment.id]), category=UserNotification.CATEGORY_PAYMENT, actor=request.user,
    )
    messages.error(request, f'🚫 فیش #{payment_id} رد/ابطال شد.')
    return redirect('counterparty_dashboard')


@login_required
@require_POST
def counterparty_return_payment(request, payment_id):
    """بازرگانی فیش را از حالت تایید طرف حساب خارج می‌کند."""
    if not _is_staff_user(request.user) or _user_role(request.user) not in {'commercial', 'commercial_manager'} and not request.user.is_superuser:
        return HttpResponseForbidden('فقط بازرگانی یا مدیر سیستم می‌تواند این عملیات را انجام دهد.')

    payment = get_object_or_404(PaymentRecord, id=payment_id)

    if not payment.is_counterparty_approved:
        messages.error(request, 'این فیش تایید طرف حساب ندارد.')
        return redirect(_safe_next_url(request) or 'submit')

    note = (request.POST.get('note') or '').strip()
    payment.counterparty_status = None
    payment.counterparty_note = None
    payment.counterparty_decided_at = None
    payment.counterparty_decided_by = None
    payment.pending_final_approval = False
    payment.pending_final_approval_since = None
    payment.save(update_fields=[
        'counterparty_status', 'counterparty_note',
        'counterparty_decided_at', 'counterparty_decided_by',
        'pending_final_approval', 'pending_final_approval_since',
    ])

    _log_activity(payment, request.user, PaymentActivityLog.ACTION_CP_RETURNED,
                  note=note or 'بازگشت از تایید طرف حساب توسط بازرگانی')

    messages.success(request, f'فیش #{payment_id} به صف بررسی طرف حساب بازگشت داده شد.')
    return redirect(_safe_next_url(request) or 'submit')


def _log_counterparty_action(cp, user, description):
    """ثبت عملیات طرف حساب در لاگ سیستم."""
    SystemActivityLog.objects.create(
        actor=user,
        target_user=None,
        action='counterparty_action',
        description=f'طرف حساب «{cp.name}»: {description}',
    )


def login_ad_image(request, ad_id):
    ad = get_object_or_404(LoginAdvertisement, id=ad_id, is_visible=True)
    today = timezone.localdate()
    if ad.start_date > today or ad.end_date < today:
        raise Http404
    return _file_response(ad.image)


@login_required
def bank_names_autocomplete(request):
    """
    API endpoint for bank names autocomplete.
    Returns list of unique bank names matching the search query.
    """
    query = (request.GET.get('q') or '').strip()
    field_type = (request.GET.get('type') or 'payer').strip()  # 'payer' or 'beneficiary'
    
    if len(query) < 1:
        return JsonResponse({'results': []})
    
    if field_type == 'beneficiary':
        bank_names = (
            PaymentRecord.objects
            .filter(beneficiary_bank_name__icontains=query)
            .values_list('beneficiary_bank_name', flat=True)
            .distinct()
            .order_by('beneficiary_bank_name')[:20]
        )
    else:  # payer
        bank_names = (
            PaymentRecord.objects
            .filter(payer_bank_name__icontains=query)
            .values_list('payer_bank_name', flat=True)
            .distinct()
            .order_by('payer_bank_name')[:20]
        )
    
    results = [{'text': name, 'id': name} for name in bank_names if name]
    return JsonResponse({'results': results})


@login_required
def export_records(request):
    return export_data(request, 'payments')


@login_required
def export_data(request, dataset):
    if dataset in {'payments', 'payment_history'}:
        is_staff_user = _is_staff_user(request.user)
        records = _history_payment_records_for_user(request.user) if dataset == 'payment_history' else _active_payment_records_for_user(request.user)
        records, _ = _apply_record_filters(records, request, is_staff_user=is_staff_user)
        records = _export_scope_records(request, records, page_param='page')
        fields = _selected_export_fields(request, PAYMENT_EXPORT_FIELDS)
        return _export_response(f'{dataset}.xlsx', 'Payments', fields, records)

    if dataset == 'customer_payments':
        if not _is_staff_user(request.user):
            return HttpResponseForbidden('خروجی برای نقش کاربری شما فعال نیست.')
        customer = get_object_or_404(User, id=request.GET.get('customer_id'))
        records = (
            PaymentRecord.objects
            .filter(user=customer)
            .select_related('counterparty', 'user')
            .order_by('-created_at')
        )
        records = _export_scope_records(request, records, page_param='payments_page')
        fields = _selected_export_fields(request, PAYMENT_EXPORT_FIELDS)
        return _export_response('customer_payments.xlsx', 'Payments', fields, records)

    if dataset in {'invoices', 'customer_invoices'}:
        if dataset == 'customer_invoices':
            if not _is_staff_user(request.user):
                return HttpResponseForbidden('خروجی برای نقش کاربری شما فعال نیست.')
            customer = get_object_or_404(User, id=request.GET.get('customer_id'))
            records = InvoiceRecord.objects.filter(customer=customer).select_related('customer', 'customer__profile', 'uploaded_by')
            invoice_number_filter = (request.GET.get('invoice_number') or '').strip()
            if invoice_number_filter:
                records = records.filter(invoice_number__icontains=invoice_number_filter)
            records = _export_scope_records(request, records, page_param='invoice_page')
        else:
            records = _invoice_records_for_user(request.user)
            records, _ = _apply_invoice_filters(records, request, is_staff_user=_is_staff_user(request.user))
            records = _export_scope_records(request, records, page_param='page')
        fields = _selected_export_fields(request, INVOICE_EXPORT_FIELDS)
        return _export_response(f'{dataset}.xlsx', 'Invoices', fields, records)

    if dataset == 'customers':
        if not _is_staff_user(request.user):
            return HttpResponseForbidden('خروجی برای نقش کاربری شما فعال نیست.')
        records, _ = _customer_list_rows(request)
        records = _export_scope_records(request, records, page_param='page')
        fields = _selected_export_fields(request, CUSTOMER_EXPORT_FIELDS)
        return _export_response('customers.xlsx', 'Customers', fields, records)

    if dataset == 'users':
        if not _can_manage_users(request.user):
            return HttpResponseForbidden('خروجی برای نقش کاربری شما فعال نیست.')
        records = _managed_users(
            query=(request.GET.get('q') or '').strip(),
            role=(request.GET.get('role') or '').strip(),
            status=(request.GET.get('status') or '').strip(),
        )
        records = _export_scope_records(request, records, page_param='page')
        fields = _selected_export_fields(request, USER_EXPORT_FIELDS)
        return _export_response('users.xlsx', 'Users', fields, records)

    if dataset == 'counterparties':
        if not request.user.is_superuser:
            return HttpResponseForbidden('خروجی برای نقش کاربری شما فعال نیست.')
        fields = _selected_export_fields(request, COUNTERPARTY_EXPORT_FIELDS)
        records = _export_scope_records(request, Counterparty.objects.all(), page_param='page')
        return _export_response('counterparties.xlsx', 'Counterparties', fields, records)

    if dataset == 'daily_plans':
        if not _can_view_daily_payments(request.user):
            return HttpResponseForbidden('خروجی برای نقش کاربری شما فعال نیست.')
        period = _daily_payment_period(request)
        fields = _selected_export_fields(request, DAILY_PLAN_EXPORT_FIELDS)
        records = _export_scope_records(request, _daily_plans_for_period(period['start_date'], period['end_date']), page_param='page')
        return _export_response('daily_payment_plans.xlsx', 'Daily Plans', fields, records)

    if dataset == 'daily_assignments':
        if not _can_view_daily_payments(request.user):
            return HttpResponseForbidden('خروجی برای نقش کاربری شما فعال نیست.')
        plan = get_object_or_404(DailyPaymentPlan, id=request.GET.get('plan_id'))
        fields = _selected_export_fields(request, DAILY_ASSIGNMENT_EXPORT_FIELDS)
        records = _export_scope_records(request, _daily_assignments_for_plan(plan), page_param='page')
        return _export_response('daily_assignments.xlsx', 'Assignments', fields, records)

    if dataset == 'customer_daily_assignments':
        if _is_staff_user(request.user):
            return HttpResponseForbidden('این خروجی برای پنل مشتری است.')
        fields = _selected_export_fields(request, DAILY_ASSIGNMENT_EXPORT_FIELDS)
        records = _customer_daily_assignments_for_user(request.user, request=request)
        status_filter = (request.GET.get('status') or '').strip()
        if status_filter:
            records = [
                assignment for assignment in records
                if (
                    (status_filter == 'none' and assignment.report['paid_amount'] <= 0) or
                    (status_filter == 'partial' and 0 < assignment.report['paid_amount'] < assignment.expected_amount) or
                    (status_filter == 'complete' and assignment.expected_amount > 0 and assignment.report['paid_amount'] >= assignment.expected_amount)
                )
            ]
        records = _export_scope_records(request, records, page_param='page')
        return _export_response('my_daily_payment_assignments.xlsx', 'Daily Assignments', fields, records)

    if dataset == 'orders':
        if not _can_view_orders(request.user):
            return HttpResponseForbidden('خروجی سفارش برای نقش کاربری شما فعال نیست.')
        records = _orders_for_user(request.user)
        records, _ = _apply_order_filters(records, request, is_staff_user=_is_staff_user(request.user))
        records = _export_scope_records(request, records, page_param='page')
        fields = _selected_export_fields(request, ORDER_EXPORT_FIELDS)
        return _export_response('orders.xlsx', 'Orders', fields, records)

    if dataset == 'sales_assignments':
        if not _can_manage_sales_assignments(request.user):
            return HttpResponseForbidden('خروجی تخصیص مشتریان برای نقش شما فعال نیست.')
        records, _ = _sales_assignment_rows(request)
        records = _export_scope_records(request, records, page_param='page')
        fields = _selected_export_fields(request, SALES_ASSIGNMENT_EXPORT_FIELDS)
        return _export_response('sales_assignments.xlsx', 'Sales Assignments', fields, records)

    raise Http404


@login_required
def customer_detail(request, user_id):
    """
    Dedicated view for staff to see all documents (payments and invoices) for a specific customer.
    """
    is_staff_user = _is_staff_user(request.user)
    if not is_staff_user:
        return HttpResponseForbidden('این بخش فقط برای کاربران واحدها قابل دسترسی است.')
    return_url = _safe_next_url(request)

    # Get the customer user
    customer_user = get_object_or_404(User.objects.select_related('profile'), id=user_id)
    customer_profile = getattr(customer_user, 'profile', None)

    # Get all payments for this customer
    payments = (
        PaymentRecord.objects
        .filter(user=customer_user)
        .select_related('counterparty', 'user')
        .prefetch_related('receipts', 'activity_logs', 'activity_logs__actor')
        .order_by('-created_at')
    )
    payments = _enrich_records(
        payments,
        staff_role=_user_role(request.user),
        is_system_admin=request.user.is_superuser,
        can_edit_payment_details=_can_edit_payment_details(request.user),
    )

    # Get all invoices for this customer
    invoices = (
        InvoiceRecord.objects
        .filter(customer=customer_user)
        .select_related('customer', 'customer__profile', 'uploaded_by')
        .order_by('-created_at')
    )
    can_view_invoices = _can_view_invoices(request.user)
    if not can_view_invoices:
        invoices = InvoiceRecord.objects.none()

    invoice_number_filter = (request.GET.get('invoice_number') or '').strip()
    if invoice_number_filter and can_view_invoices:
        invoices = invoices.filter(invoice_number__icontains=invoice_number_filter)

    # Calculate summary
    total_payments = len(payments)
    total_invoices = invoices.count()
    total_amount = sum(p.amount for p in payments)
    invoice_total_amount = invoices.aggregate(total=Sum('amount'))['total'] or 0

    payments_page_obj = _paginate_queryset(request, payments, per_page=10, page_param='payments_page')
    invoices_page_obj = _paginate_queryset(request, invoices, per_page=10, page_param='invoice_page')
    payments_page_base_query = _build_query_string(request, remove_keys=['payments_page'])
    invoices_page_base_query = _build_query_string(request, remove_keys=['invoice_page'])

    # Status breakdown for payments
    status_counts = {}
    for payment in payments:
        status = payment.get_status_display()
        status_counts[status] = status_counts.get(status, 0) + 1

    user_display_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username

    return render(request, 'payments/customer_detail.html', {
        'customer_user': customer_user,
        'customer_profile': customer_profile,
        'payments': payments_page_obj,
        'invoices': invoices_page_obj,
        'payments_page_obj': payments_page_obj,
        'invoices_page_obj': invoices_page_obj,
        'payments_page_base_query': payments_page_base_query,
        'invoices_page_base_query': invoices_page_base_query,
        'total_payments': total_payments,
        'total_invoices': total_invoices,
        'total_amount': total_amount,
        'invoice_total_amount': invoice_total_amount,
        'customer_debt': _customer_debt_summary(customer_user),
        'status_counts': status_counts,
        'is_staff_user': is_staff_user,
        'staff_user_role': _user_role(request.user),
        'can_view_invoices': can_view_invoices,
        'filters': {
            'invoice_number': invoice_number_filter,
        },
        'user_display_name': user_display_name,
        'return_url': return_url,
        'payment_export_fields': PAYMENT_EXPORT_FIELDS,
        'invoice_export_fields': INVOICE_EXPORT_FIELDS,
        'customer_export_params': {'customer_id': customer_user.id},
    })


@login_required
def customers_list(request):
    """
    List all customers with their document counts for staff users.
    """
    is_staff_user = _is_staff_user(request.user)
    if not is_staff_user:
        return HttpResponseForbidden('این بخش فقط برای کاربران واحدها قابل دسترسی است.')

    # فقط مشتریان — طرف حساب‌ها مستثنی هستند
    customers = (
        UserProfile.objects
        .filter(role='customer')
        .exclude(user__counterparty_account__isnull=False)
        .select_related('user')
        .order_by('user__username')
    )
    if _user_role(request.user) == 'sales' and not request.user.is_superuser:
        customers = customers.filter(user__sales_assignment__sales_user=request.user)
    filters = {
        'q':               (request.GET.get('q') or '').strip(),
        'status':          (request.GET.get('status') or '').strip(),
        'accounting_code': (request.GET.get('accounting_code') or '').strip(),
    }
    if filters['q']:
        customers = customers.filter(
            Q(user__username__icontains=filters['q']) |
            Q(user__first_name__icontains=filters['q']) |
            Q(user__last_name__icontains=filters['q']) |
            Q(first_name__icontains=filters['q']) |
            Q(last_name__icontains=filters['q']) |
            Q(organization__icontains=filters['q']) |
            Q(city__icontains=filters['q']) |
            Q(province__icontains=filters['q']) |
            Q(phone__icontains=filters['q']) |
            Q(mobile__icontains=filters['q']) |
            Q(representative_name__icontains=filters['q']) |
            Q(representative_mobile__icontains=filters['q']) |
            Q(accounting_code__icontains=filters['q'])
        )
    if filters['accounting_code']:
        customers = customers.filter(accounting_code__icontains=filters['accounting_code'])
    if filters['status'] == 'active':
        customers = customers.filter(suspended=False, user__is_active=True)
    elif filters['status'] == 'suspended':
        customers = customers.filter(suspended=True)
    elif filters['status'] == 'inactive':
        customers = customers.filter(user__is_active=False)

    payment_stats = {
        row['user']: row
        for row in (
            PaymentRecord.objects
            .filter(user__profile__role='customer')
            .values('user')
            .annotate(
                payment_count=Count('id'),
                total_amount=Sum('amount'),
                latest_payment_date=Max('created_at'),
            )
        )
    }
    confirmed_payment_totals = {
        row['user']: row['total'] or 0
        for row in (
            PaymentRecord.objects
            .filter(
                user__profile__role='customer',
                status__in=[
                    PaymentRecord.STATUS_APPROVED,
                    PaymentRecord.STATUS_FINAL_APPROVED,
                ],
            )
            .values('user')
            .annotate(total=Sum('amount'))
        )
    }
    review_payment_totals = {
        row['user']: row['total'] or 0
        for row in (
            PaymentRecord.objects
            .filter(user__profile__role='customer')
            .exclude(status=PaymentRecord.STATUS_REJECTED)
            .values('user')
            .annotate(total=Sum('amount'))
        )
    }
    invoice_counts = {
        row['customer']: row['invoice_count']
        for row in (
            InvoiceRecord.objects
            .filter(customer__profile__role='customer')
            .values('customer')
            .annotate(invoice_count=Count('id'))
        )
    }
    invoice_totals = {
        row['customer']: row['total'] or 0
        for row in (
            InvoiceRecord.objects
            .filter(customer__profile__role='customer')
            .values('customer')
            .annotate(total=Sum('amount'))
        )
    }

    # Calculate counts for each customer
    customer_data = []
    for profile in customers:
        stats = payment_stats.get(profile.user_id, {})
        invoice_total = invoice_totals.get(profile.user_id, 0)
        confirmed_payment_total = confirmed_payment_totals.get(profile.user_id, 0)
        review_payment_total = review_payment_totals.get(profile.user_id, 0)

        customer_data.append({
            'profile': profile,
            'user': profile.user,
            'payment_count': stats.get('payment_count') or 0,
            'invoice_count': invoice_counts.get(profile.user_id, 0),
            'total_amount': stats.get('total_amount') or 0,
            'invoice_total': invoice_total,
            'confirmed_debt': invoice_total - confirmed_payment_total,
            'review_debt': invoice_total - review_payment_total,
            'latest_payment_date': stats.get('latest_payment_date'),
        })

    user_display_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
    page_obj = _paginate_queryset(request, customer_data, per_page=10, page_param='page')
    page_base_query = _build_query_string(request, remove_keys=['page'])

    return render(request, 'payments/customers_list.html', {
        'customer_data': page_obj,
        'page_obj': page_obj,
        'page_base_query': page_base_query,
        'is_staff_user': is_staff_user,
        'can_manage_users': _can_manage_users(request.user),
        'user_display_name': user_display_name,
        'filters': filters,
        'export_dataset': 'customers',
        'export_fields': CUSTOMER_EXPORT_FIELDS,
        'can_import_accounting_codes': _can_import_customer_accounting_codes(request.user),
    })


_CUSTOMER_ACCOUNTING_IMPORT_SESSION_KEY = 'customer_accounting_code_import_preview'


def _normalize_customer_match_text(value):
    text = str(value or '').strip()
    text = text.translate(str.maketrans({
        'ي': 'ی',
        'ك': 'ک',
        'ۀ': 'ه',
        'ة': 'ه',
        'ؤ': 'و',
        'إ': 'ا',
        'أ': 'ا',
        'آ': 'ا',
    }))
    text = re.sub(r'[\u064b-\u065f\u0670]', '', text)
    text = re.sub(r'[^\w\s\u0600-\u06ff]', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip().lower()
    return text


def _split_accounting_title(title):
    raw = str(title or '').strip()
    city = ''
    match = re.search(r'\(([^()]+)\)\s*$', raw)
    if match:
        city = match.group(1).strip()
        raw = raw[:match.start()].strip()
    return raw, city


def _customer_match_candidates(profile):
    user = profile.user
    values = [
        f'{profile.first_name} {profile.last_name}',
        f'{user.first_name} {user.last_name}',
        profile.display_name,
        profile.organization,
        profile.representative_name,
        user.username,
        profile.phone,
        profile.mobile,
    ]
    return [v.strip() for v in values if str(v or '').strip()]


def _customer_match_index(profile):
    return {
        'profile': profile,
        'city_norm': _normalize_customer_match_text(profile.city),
        'candidates': [
            (candidate, _normalize_customer_match_text(candidate))
            for candidate in _customer_match_candidates(profile)
        ],
    }


def _score_accounting_customer_match(title_norm, city_norm, customer_index):
    profile_city_norm = customer_index['city_norm']
    best_score = 0
    best_field = ''
    for candidate, candidate_norm in customer_index['candidates']:
        if not candidate_norm or not title_norm:
            continue
        if title_norm == candidate_norm:
            score = 100
        elif title_norm in candidate_norm or candidate_norm in title_norm:
            score = 88
        else:
            title_tokens = set(title_norm.split())
            candidate_tokens = set(candidate_norm.split())
            overlap = len(title_tokens & candidate_tokens) / max(len(title_tokens | candidate_tokens), 1)
            ratio = SequenceMatcher(None, title_norm, candidate_norm).ratio()
            score = int(max(overlap * 100, ratio * 92))
        if city_norm and profile_city_norm and city_norm == profile_city_norm:
            score = min(100, score + 5)
        if score > best_score:
            best_score = score
            best_field = candidate
    return best_score, best_field


def _build_accounting_code_import_preview(file_obj):
    import openpyxl

    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError('فایل خالی است.')

    headers = [str(c or '').strip() for c in rows[0]]

    def col(candidates, fallback):
        normalized_headers = [_normalize_customer_match_text(h) for h in headers]
        for name in candidates:
            needle = _normalize_customer_match_text(name)
            for idx, header in enumerate(normalized_headers):
                if header == needle:
                    return idx
        return fallback if fallback < len(headers) else None

    code_col = col(['کد', 'کد تفضیلی', 'accounting_code', 'code'], 1)
    title_col = col(['عنوان', 'نام', 'نام مشتری', 'customer', 'name', 'title'], 2)
    id_col = col(['ID', 'شناسه'], 0)
    if code_col is None or title_col is None:
        raise ValueError('ستون‌های «کد» و «عنوان» در فایل یافت نشد.')

    customers = [
        _customer_match_index(profile)
        for profile in (
            UserProfile.objects
            .filter(role='customer')
            .exclude(user__counterparty_account__isnull=False)
            .select_related('user')
        )
    ]
    exact_lookup = {}
    token_lookup = {}
    for idx, customer_index in enumerate(customers):
        for _, candidate_norm in customer_index['candidates']:
            if not candidate_norm:
                continue
            exact_lookup.setdefault(candidate_norm, set()).add(idx)
            for token in candidate_norm.split():
                token_lookup.setdefault(token, set()).add(idx)

    preview = []
    summary = {'total': 0, 'matched': 0, 'review': 0, 'unmatched': 0, 'ambiguous': 0}

    for row_number, row in enumerate(rows[1:], start=2):
        def cell(idx):
            if idx is None or idx >= len(row) or row[idx] is None:
                return ''
            value = row[idx]
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value).strip()

        code = cell(code_col)
        title_raw = cell(title_col)
        external_id = cell(id_col)
        if not code and not title_raw:
            continue
        title, city = _split_accounting_title(title_raw)
        title_norm = _normalize_customer_match_text(title)
        city_norm = _normalize_customer_match_text(city)
        scored = []
        candidate_indexes = set(exact_lookup.get(title_norm, set()))
        if not candidate_indexes:
            for token in title_norm.split():
                candidate_indexes.update(token_lookup.get(token, set()))
        for idx in candidate_indexes:
            customer_index = customers[idx]
            score, field = _score_accounting_customer_match(title_norm, city_norm, customer_index)
            if score:
                scored.append((score, field, customer_index['profile']))
        scored.sort(key=lambda item: item[0], reverse=True)

        best = scored[0] if scored else None
        second = scored[1] if len(scored) > 1 else None
        status = 'unmatched'
        status_label = 'بدون تطبیق'
        is_safe = False
        ambiguous = False
        profile = None
        matched_field = ''
        score = 0

        if best:
            score, matched_field, profile = best
            ambiguous = bool(second and second[0] >= score - 3)
            if score < 75:
                status = 'unmatched'
                status_label = 'بدون تطبیق'
                profile = None
                matched_field = ''
            elif ambiguous:
                status = 'ambiguous'
                status_label = 'چند تطبیق نزدیک'
            elif score >= 92:
                status = 'matched'
                status_label = 'تطبیق مطمئن'
                is_safe = True
            elif score >= 75:
                status = 'review'
                status_label = 'نیازمند بررسی'
            else:
                status = 'unmatched'
                status_label = 'بدون تطبیق'

        summary['total'] += 1
        summary[status] += 1
        preview.append({
            'row_number': row_number,
            'external_id': external_id,
            'code': code,
            'title_raw': title_raw,
            'title': title,
            'city': city,
            'status': status,
            'status_label': status_label,
            'is_safe': is_safe,
            'score': score,
            'matched_field': matched_field,
            'profile_id': profile.id if profile else None,
            'user_id': profile.user_id if profile else None,
            'customer_name': profile.display_name if profile else '',
            'customer_org': profile.organization if profile else '',
            'customer_city': profile.city if profile else '',
            'old_code': profile.accounting_code if profile else '',
        })

    return preview, summary


@login_required
def import_customer_accounting_codes(request):
    if not _can_import_customer_accounting_codes(request.user):
        return HttpResponseForbidden('شما دسترسی ورود کد تفضیلی مشتریان را ندارید.')

    preview = request.session.get(_CUSTOMER_ACCOUNTING_IMPORT_SESSION_KEY, [])
    summary = {
        'total': len(preview),
        'matched': sum(1 for row in preview if row.get('status') == 'matched'),
        'review': sum(1 for row in preview if row.get('status') == 'review'),
        'unmatched': sum(1 for row in preview if row.get('status') == 'unmatched'),
        'ambiguous': sum(1 for row in preview if row.get('status') == 'ambiguous'),
    }

    if request.method == 'POST' and request.POST.get('action') == 'preview':
        file = request.FILES.get('file')
        if not file:
            messages.error(request, 'فایل اکسل را انتخاب کنید.')
            return redirect('import_customer_accounting_codes')
        try:
            preview, summary = _build_accounting_code_import_preview(file)
            request.session[_CUSTOMER_ACCOUNTING_IMPORT_SESSION_KEY] = preview
            request.session.modified = True
            messages.success(request, f'{summary["total"]} ردیف خوانده شد؛ {summary["matched"]} مورد تطبیق مطمئن دارد.')
        except Exception as exc:
            logger.exception('Customer accounting code import preview failed')
            messages.error(request, f'خطا در خواندن فایل: {exc}')
            return redirect('import_customer_accounting_codes')

    elif request.method == 'POST' and request.POST.get('action') == 'apply':
        selected = {int(idx) for idx in request.POST.getlist('apply_rows') if str(idx).isdigit()}
        if not preview:
            messages.error(request, 'ابتدا فایل را بارگذاری و پیش‌نمایش را بررسی کنید.')
            return redirect('import_customer_accounting_codes')
        updated = 0
        skipped = 0
        with transaction.atomic():
            for idx, row in enumerate(preview):
                if idx not in selected:
                    skipped += 1
                    continue
                profile_id = row.get('profile_id')
                code = (row.get('code') or '').strip()
                if not profile_id or not code:
                    skipped += 1
                    continue
                profile = UserProfile.objects.select_for_update().get(id=profile_id, role='customer')
                if profile.accounting_code != code:
                    profile.accounting_code = code
                    profile.save(update_fields=['accounting_code'])
                    updated += 1
                else:
                    skipped += 1
        request.session.pop(_CUSTOMER_ACCOUNTING_IMPORT_SESSION_KEY, None)
        messages.success(request, f'{updated} کد تفضیلی ثبت/به‌روزرسانی شد. {skipped} ردیف بدون تغییر ماند.')
        return redirect('customers_list')

    return render(request, 'payments/import_customer_accounting_codes.html', {
        'preview': preview,
        'summary': summary,
        'user_display_name': f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username,
    })


@login_required
@require_POST
def reset_user_password(request, user_id):
    """
    Reset a user's password to a temporary numeric password and force them to change it.
    Only accessible to users with management permissions.
    """
    if not _can_manage_users(request.user):
        return HttpResponseForbidden('شما دسترسی ریست رمز عبور را ندارید.')

    target_user = get_object_or_404(User, id=user_id)
    temp_password = _suggest_five_digit_password()

    from django.http import JsonResponse

    email_sent = False
    email_note = 'برای این کاربر ایمیل ثبت نشده است؛ رمز جدید فقط به مدیر سیستم نمایش داده شد.'
    if target_user.email:
        try:
            email_sent, email_error = _send_temporary_password_email(target_user, temp_password)
            if email_sent:
                email_note = f'رمز جدید به ایمیل {target_user.email} ارسال شد.'
            else:
                email_note = email_error
        except Exception:
            email_note = 'ارسال ایمیل انجام نشد؛ رمز جدید فقط به مدیر سیستم نمایش داده شد.'

    target_user.set_password(temp_password)
    target_user.save()

    # Invalidate all sessions for the target user
    for session in Session.objects.all():
        session_data = session.get_decoded()
        if session_data.get('_auth_user_id') == str(target_user.id):
            session.delete()

    profile = getattr(target_user, 'profile', None)
    if profile:
        profile.force_password_change = True
        profile.save(update_fields=['force_password_change'])

    _log_system_activity(
        request.user,
        target_user,
        SystemActivityLog.ACTION_PASSWORD_RESET,
        f'رمز عبور کاربر ریست شد. {email_note} کاربر ملزم به تغییر رمز در ورود بعدی شد.',
    )

    message = 'رمز عبور کاربر با موفقیت ریست شد. این رمز فقط همین لحظه نمایش داده می‌شود.'
    if email_sent:
        message = 'رمز عبور کاربر با موفقیت ریست شد و به ایمیل کاربر ارسال شد. این رمز فقط همین لحظه نمایش داده می‌شود.'
    elif not target_user.email:
        message = 'رمز عبور کاربر با موفقیت ریست شد. برای این کاربر ایمیل ثبت نشده است؛ رمز را همین حالا به کاربر اطلاع دهید.'
    else:
        message = 'رمز عبور کاربر با موفقیت ریست شد. ارسال ایمیل انجام نشد؛ رمز را همین حالا به کاربر اطلاع دهید.'

    return JsonResponse({
        'success': True,
        'temp_password': temp_password,
        'message': message,
    })


# ─── SMS OTP MFA ─────────────────────────────────────────────────────────────

def _sms_mfa_is_active():
    """آیا SMS MFA در کل سیستم فعال است؟"""
    try:
        from .models import SystemSettings
        cfg = SystemSettings.load()
        return cfg.sms_provider != 'disabled' and bool(cfg.sms_api_key)
    except Exception:
        return False


@login_required
def sms_mfa_setup(request):
    """صفحه فعال/غیرفعال کردن SMS MFA توسط خود کاربر."""
    if not _sms_mfa_is_active():
        messages.info(request, 'ارسال پیامک در حال حاضر توسط مدیر سیستم غیرفعال است.')
        return redirect('submit')

    profile = getattr(request.user, 'profile', None)
    if not profile:
        messages.error(request, 'پروفایل کاربری یافت نشد.')
        return redirect('submit')

    if not profile.sms_number:
        messages.error(request, 'برای فعال‌سازی ورود پیامکی ابتدا باید شماره موبایل خود را در پروفایل ثبت کنید.')
        return redirect('profile_edit')

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'enable':
            profile.sms_mfa_enabled = True
            profile.save(update_fields=['sms_mfa_enabled'])
            messages.success(request, 'ورود دو مرحله‌ای با پیامک فعال شد.')
        elif action == 'disable':
            profile.sms_mfa_enabled = False
            profile.save(update_fields=['sms_mfa_enabled'])
            messages.success(request, 'ورود دو مرحله‌ای با پیامک غیرفعال شد.')
        return redirect('sms_mfa_setup')

    return render(request, 'payments/sms_mfa_setup.html', {
        'profile': profile,
        'sms_active': _sms_mfa_is_active(),
        'masked_phone': _mask_phone(profile.sms_number),
    })


def _mask_phone(phone):
    """09xxxxxxx89 → 09***x89"""
    if not phone or len(phone) < 6:
        return phone
    return phone[:3] + '****' + phone[-3:]


@login_required
def sms_otp_verify(request):
    """صفحه تأیید کد OTP پیامکی بعد از login."""
    if not request.session.get('sms_mfa_pending'):
        return redirect('submit')

    error = ''
    resent = False

    if request.method == 'POST':
        action = request.POST.get('action', 'verify')

        if action == 'resend':
            from .sms_service import send_otp
            otp = send_otp(request.user, purpose='mfa')
            if otp:
                request.session['sms_mfa_otp_key'] = otp.pk
                resent = True
                messages.success(request, 'کد جدید ارسال شد.')
            else:
                error = 'ارسال کد جدید ناموفق بود. لطفاً با مدیر سیستم تماس بگیرید.'

        else:
            submitted = request.POST.get('code', '').strip()
            from .sms_service import verify_otp
            ok, msg = verify_otp(request.user, submitted, purpose='mfa')
            if ok:
                request.session.pop('sms_mfa_pending', None)
                request.session.pop('sms_mfa_otp_key', None)
                next_url = request.session.pop('sms_mfa_next', '') or reverse('submit')
                return redirect(next_url)
            else:
                error = msg

    profile = getattr(request.user, 'profile', None)
    masked = _mask_phone(profile.sms_number if profile else '')
    return render(request, 'payments/sms_otp_verify.html', {
        'masked_phone': masked,
        'error': error,
        'resent': resent,
    })


@login_required
def sms_test_send(request):
    """ارسال پیامک آزمایشی — فقط ادمین."""
    if not request.user.is_superuser:
        return HttpResponseForbidden()
    if request.method == 'POST':
        phone = request.POST.get('phone', '').strip()
        msg = request.POST.get('message', 'پیامک آزمایشی از سامانه').strip()
        from .sms_service import send_sms
        ok, err = send_sms(phone, msg, purpose='test')
        if ok:
            messages.success(request, f'پیامک به {phone} ارسال شد.')
        else:
            messages.error(request, f'خطا: {err}')
        return redirect(request.META.get('HTTP_REFERER', '/admin/'))
    return HttpResponseForbidden()


# ─── تست خوانش فیش بانکی — فقط ادمین ──────────────────────────────────────

@login_required
def receipt_reader_test(request):
    """ابزار تست OCR فیش بانکی — فقط مدیر سیستم."""
    if not request.user.is_superuser:
        return HttpResponseForbidden('فقط مدیر سیستم به این صفحه دسترسی دارد.')

    result = None
    error  = None

    if request.method == 'POST' and request.FILES.get('receipt_file'):
        uploaded = request.FILES['receipt_file']
        import tempfile as _tmp, os as _os
        from .receipt_extraction import extract_receipt_file

        suffix = _os.path.splitext(uploaded.name or '')[1].lower() or '.tmp'
        tmp_path = None
        try:
            with _tmp.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                for chunk in uploaded.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            result = extract_receipt_file(tmp_path, original_name=uploaded.name)
        except Exception as exc:
            error = str(exc)
            logger.exception('Receipt reader test failed')
        finally:
            if tmp_path:
                try:
                    _os.unlink(tmp_path)
                except Exception:
                    pass

    return render(request, 'payments/receipt_reader_test.html', {
        'result':         result,
        'error':          error,
        'gemini_enabled':   bool(getattr(settings, 'GEMINI_API_KEY', '')),
        'claude_enabled':   bool(getattr(settings, 'ANTHROPIC_API_KEY', '')),
        'ocrspace_enabled': bool(getattr(settings, 'OCRSPACE_API_KEY', '')),
    })


# ─── پلتفرم درخواست نمایندگی ─────────────────────────────────────────────────

import hashlib as _hashlib
import random as _random
import string as _string


def _agency_generate_tracking():
    chars = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'
    return ''.join(_random.choices(chars, k=8))


def _agency_gen_otp():
    return ''.join(_random.choices(_string.digits, k=6))


def _agency_hash(code):
    return _hashlib.sha256(code.encode()).hexdigest()


def _agency_otp_send(phone, otp_code):
    from .sms_service import send_sms
    msg = f'کد تأیید درخواست نمایندگی: {otp_code}\nاعتبار: ۵ دقیقه'
    send_sms(phone, msg, purpose='agency_otp')


def _agency_normalize_phone(phone):
    p = (phone or '').strip().replace(' ', '').replace('-', '')
    if p.startswith('+98'):
        p = '0' + p[3:]
    if p.startswith('98') and len(p) == 12:
        p = '0' + p[2:]
    return p


def _agency_phone_valid(phone):
    p = _agency_normalize_phone(phone)
    return p.startswith('09') and len(p) == 11 and p.isdigit()


def _agency_can_apply(phone):
    """آیا این شماره درخواست فعال (non-final) دارد؟"""
    from .models import AgencyApplication
    return not AgencyApplication.objects.filter(
        phone=phone,
        status__in=[AgencyApplication.STATUS_PENDING, AgencyApplication.STATUS_REVIEWING, AgencyApplication.STATUS_INFO_NEEDED],
    ).exists()


def _agency_log(application, actor, action, note=''):
    from .models import AgencyApplicationLog
    AgencyApplicationLog.objects.create(application=application, actor=actor, action=action, note=note)


def _agency_notify_sms(phone, message):
    from .sms_service import send_sms
    send_sms(phone, message, purpose='agency_notify')


def _agency_create_user(application):
    """ایجاد کاربر جدید برای متقاضی تأییدشده — برگرداندن (user, password)."""
    from django.contrib.auth.models import User
    base = 'agent_' + _agency_normalize_phone(application.phone)[-8:]
    username = base
    counter = 1
    while User.objects.filter(username=username).exists():
        username = f'{base}_{counter}'
        counter += 1
    password = ''.join(_random.choices(_string.ascii_letters + _string.digits, k=10))
    user = User.objects.create_user(
        username=username,
        password=password,
        first_name=application.first_name,
        last_name=application.last_name,
        email=application.email or '',
    )
    try:
        from .models import UserProfile
        UserProfile.objects.create(
            user=user,
            role='sales',
            phone=application.phone,
            mobile=application.phone,
        )
    except Exception:
        pass
    return user, password


def agency_register_phone(request):
    """مرحله ۱: ورود شماره موبایل و ارسال کد تأیید."""
    if request.method == 'POST':
        phone = _agency_normalize_phone(request.POST.get('phone', ''))
        if not _agency_phone_valid(phone):
            return render(request, 'payments/agency_register.html', {'error': 'شماره موبایل معتبر نیست (مثال: ۰۹۱۲۳۴۵۶۷۸۹)'})

        # rate limit: max 3 OTP per session
        sent_count = request.session.get('agency_otp_count', 0)
        if sent_count >= 5:
            return render(request, 'payments/agency_register.html', {'error': 'تعداد درخواست کد تأیید از حد مجاز گذشته است. لطفاً بعداً تلاش کنید.'})

        if not _agency_can_apply(phone):
            return render(request, 'payments/agency_register.html', {
                'error': 'این شماره موبایل یک درخواست فعال دارد. برای پیگیری از بخش «استعلام وضعیت» استفاده کنید.',
                'track_url': reverse('agency_track'),
            })

        otp = _agency_gen_otp()
        request.session['agency_reg_phone'] = phone
        request.session['agency_reg_otp_hash'] = _agency_hash(otp)
        request.session['agency_reg_otp_expires'] = (timezone.now() + timezone.timedelta(minutes=5)).isoformat()
        request.session['agency_reg_verified'] = False
        request.session['agency_otp_count'] = sent_count + 1
        request.session.modified = True

        _agency_otp_send(phone, otp)
        return redirect(reverse('agency_verify'))

    return render(request, 'payments/agency_register.html', {})


def agency_register_verify(request):
    """مرحله ۲: تأیید کد OTP."""
    phone = request.session.get('agency_reg_phone')
    if not phone:
        return redirect(reverse('agency_register'))

    error = ''
    if request.method == 'POST':
        code = (request.POST.get('otp') or '').strip()
        stored_hash = request.session.get('agency_reg_otp_hash', '')
        expires_str = request.session.get('agency_reg_otp_expires', '')

        try:
            expires = timezone.datetime.fromisoformat(expires_str)
            if timezone.is_naive(expires):
                expires = timezone.make_aware(expires)
        except Exception:
            expires = timezone.now() - timezone.timedelta(minutes=1)

        if timezone.now() > expires:
            error = 'کد منقضی شده است. لطفاً کد جدید درخواست کنید.'
        elif not code or _agency_hash(code) != stored_hash:
            error = 'کد وارد شده اشتباه است.'
        else:
            request.session['agency_reg_verified'] = True
            request.session.modified = True
            return redirect(reverse('agency_apply'))

    return render(request, 'payments/agency_verify.html', {
        'phone': phone,
        'error': error,
    })


def agency_register_apply(request):
    """مرحله ۳: تکمیل فرم درخواست نمایندگی."""
    if not request.session.get('agency_reg_verified'):
        return redirect(reverse('agency_register'))

    phone = request.session.get('agency_reg_phone', '')
    error = ''
    form_data = {}

    if request.method == 'POST':
        form_data = request.POST
        required = ['first_name', 'last_name', 'national_id', 'province', 'city',
                    'home_address', 'business_address', 'activity_domain', 'services_offered']
        missing = [f for f in required if not request.POST.get(f, '').strip()]
        if missing:
            error = 'لطفاً تمامی فیلدهای ستاره‌دار را تکمیل کنید.'
        else:
            from .models import AgencyApplication
            tracking = _agency_generate_tracking()
            while AgencyApplication.objects.filter(tracking_code=tracking).exists():
                tracking = _agency_generate_tracking()

            app = AgencyApplication.objects.create(
                phone=phone,
                phone_verified=True,
                email=request.POST.get('email', '').strip(),
                first_name=request.POST.get('first_name', '').strip(),
                last_name=request.POST.get('last_name', '').strip(),
                national_id=request.POST.get('national_id', '').strip(),
                province=request.POST.get('province', '').strip(),
                city=request.POST.get('city', '').strip(),
                home_address=request.POST.get('home_address', '').strip(),
                business_address=request.POST.get('business_address', '').strip(),
                activity_domain=request.POST.get('activity_domain', '').strip(),
                services_offered=request.POST.get('services_offered', '').strip(),
                years_experience=int(request.POST.get('years_experience') or 0),
                has_business_license='has_business_license' in request.POST,
                referrer_name=request.POST.get('referrer_name', '').strip(),
                referrer_phone=_agency_normalize_phone(request.POST.get('referrer_phone', '')),
                motivation=request.POST.get('motivation', '').strip(),
                tracking_code=tracking,
                submitted_at=timezone.now(),
            )
            _agency_log(app, None, AgencyApplicationLog.ACTION_SUBMITTED, 'ثبت درخواست توسط متقاضی')

            # اعلان به مدیر فروش
            from .models import UserNotification
            _notify_users(
                list(_staff_notification_users({'sales_manager'})),
                '🤝 درخواست نمایندگی جدید',
                f'درخواست نمایندگی از {app.full_name} — {app.city} ثبت شد.',
                reverse('agency_application_detail', args=[app.pk]),
                category=UserNotification.CATEGORY_SYSTEM,
            )

            # پیامک تأیید به متقاضی
            _agency_notify_sms(phone, f'درخواست نمایندگی شما با موفقیت ثبت شد.\nکد پیگیری: {tracking}\nبرای استعلام وضعیت از این کد استفاده کنید.')

            # پاکسازی session
            for key in ('agency_reg_phone', 'agency_reg_otp_hash', 'agency_reg_otp_expires', 'agency_reg_verified', 'agency_otp_count'):
                request.session.pop(key, None)

            return redirect(reverse('agency_success', args=[tracking]))

    return render(request, 'payments/agency_apply.html', {
        'phone': phone,
        'error': error,
        'form_data': form_data,
    })


def agency_register_success(request, tracking_code):
    """صفحه تأیید ثبت موفق درخواست."""
    from .models import AgencyApplication
    app = get_object_or_404(AgencyApplication, tracking_code=tracking_code)
    return render(request, 'payments/agency_success.html', {'app': app})


def agency_application_track(request):
    """استعلام وضعیت درخواست نمایندگی با کد پیگیری یا شماره موبایل."""
    from .models import AgencyApplication
    result = None
    error = ''
    if request.method == 'POST':
        q = request.POST.get('query', '').strip()
        phone = _agency_normalize_phone(q)
        try:
            if q.upper().replace(' ', '') and len(q) <= 10 and not q.startswith('0'):
                result = AgencyApplication.objects.filter(tracking_code=q.upper()).first()
            elif _agency_phone_valid(phone):
                result = AgencyApplication.objects.filter(phone=phone).order_by('-created_at').first()
            if not result:
                error = 'درخواستی با این مشخصات یافت نشد.'
        except Exception:
            error = 'خطا در جستجو.'
    return render(request, 'payments/agency_track.html', {'result': result, 'error': error})


# ─── ویوهای کارشناس / مدیر ────────────────────────────────────────────────────

def _can_manage_agency(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    role = _user_role(user)
    return role in {'sales_manager', 'commercial_manager', 'sales'}


@login_required
def agency_applications_list(request):
    """لیست درخواست‌های نمایندگی — برای مدیر و کارشناس فروش."""
    if not _can_manage_agency(request.user):
        return HttpResponseForbidden('دسترسی ممنوع است.')

    from .models import AgencyApplication
    qs = AgencyApplication.objects.select_related('assigned_to', 'reviewed_by', 'created_user')

    status_f = request.GET.get('status', '')
    city_f   = request.GET.get('city', '')
    search_f = request.GET.get('q', '')

    if status_f:
        qs = qs.filter(status=status_f)
    if city_f:
        qs = qs.filter(city__icontains=city_f)
    if search_f:
        qs = qs.filter(
            Q(first_name__icontains=search_f) | Q(last_name__icontains=search_f) |
            Q(phone__icontains=search_f) | Q(tracking_code__icontains=search_f)
        )

    from django.core.paginator import Paginator
    paginator = Paginator(qs, 25)
    page = paginator.get_page(request.GET.get('page', 1))

    status_counts = [
        (val, label, AgencyApplication.objects.filter(status=val).count())
        for val, label in AgencyApplication.STATUS_CHOICES
    ]
    return render(request, 'payments/agency_applications.html', {
        'page': page,
        'status_choices': AgencyApplication.STATUS_CHOICES,
        'status_filter': status_f,
        'city_filter': city_f,
        'search': search_f,
        'status_counts': status_counts,
    })


@login_required
def agency_application_detail(request, app_id):
    """جزئیات یک درخواست نمایندگی."""
    if not _can_manage_agency(request.user):
        return HttpResponseForbidden('دسترسی ممنوع است.')

    from .models import AgencyApplication
    app = get_object_or_404(AgencyApplication.objects.select_related('assigned_to', 'reviewed_by', 'created_user').prefetch_related('logs__actor'), id=app_id)

    staff_users = list(User.objects.filter(
        profile__role__in=['sales', 'sales_manager', 'commercial_manager'],
        is_active=True,
    ).select_related('profile'))

    return render(request, 'payments/agency_application_detail.html', {
        'app': app,
        'staff_users': staff_users,
    })


@login_required
@require_POST
def agency_application_action(request, app_id):
    """تغییر وضعیت درخواست نمایندگی — تأیید / رد / شروع بررسی / درخواست اطلاعات."""
    if not _can_manage_agency(request.user):
        return HttpResponseForbidden('دسترسی ممنوع است.')

    from .models import AgencyApplication, AgencyApplicationLog, UserNotification
    app = get_object_or_404(AgencyApplication, id=app_id)
    action = request.POST.get('action', '')
    note   = (request.POST.get('note') or '').strip()

    if action == 'assign':
        uid = request.POST.get('assign_to')
        if uid:
            try:
                app.assigned_to = User.objects.get(id=int(uid))
                app.save(update_fields=['assigned_to', 'updated_at'])
                _agency_log(app, request.user, AgencyApplicationLog.ACTION_ASSIGNED, f'تخصیص به {app.assigned_to.get_full_name() or app.assigned_to.username}')
                messages.success(request, 'درخواست تخصیص داده شد.')
            except User.DoesNotExist:
                messages.error(request, 'کاربر یافت نشد.')
        return redirect(reverse('agency_application_detail', args=[app_id]))

    if action == 'reviewing':
        app.status = AgencyApplication.STATUS_REVIEWING
        app.reviewed_by = request.user
        app.save(update_fields=['status', 'reviewed_by', 'updated_at'])
        _agency_log(app, request.user, AgencyApplicationLog.ACTION_REVIEWING, note)
        _agency_notify_sms(app.phone, f'درخواست نمایندگی شما (کد: {app.tracking_code}) در دست بررسی کارشناسان ما است.')
        messages.success(request, 'وضعیت به «در حال بررسی» تغییر کرد.')

    elif action == 'info_needed':
        if not note:
            messages.error(request, 'لطفاً توضیح درخواست اطلاعات را وارد کنید.')
            return redirect(reverse('agency_application_detail', args=[app_id]))
        app.status = AgencyApplication.STATUS_INFO_NEEDED
        app.info_request_note = note
        app.save(update_fields=['status', 'info_request_note', 'updated_at'])
        _agency_log(app, request.user, AgencyApplicationLog.ACTION_INFO_NEEDED, note)
        _agency_notify_sms(app.phone, f'درخواست نمایندگی شما (کد: {app.tracking_code}) نیاز به اطلاعات تکمیلی دارد:\n{note}')
        messages.info(request, 'درخواست اطلاعات تکمیلی ثبت شد.')

    elif action == 'note':
        if note:
            app.staff_note = note
            app.save(update_fields=['staff_note', 'updated_at'])
            _agency_log(app, request.user, AgencyApplicationLog.ACTION_NOTE, note)
            messages.success(request, 'یادداشت ثبت شد.')

    elif action == 'reject':
        reason = note or (request.POST.get('rejection_reason') or '').strip()
        app.status = AgencyApplication.STATUS_REJECTED
        app.rejection_reason = reason
        app.reviewed_by = request.user
        app.reviewed_at = timezone.now()
        app.save(update_fields=['status', 'rejection_reason', 'reviewed_by', 'reviewed_at', 'updated_at'])
        _agency_log(app, request.user, AgencyApplicationLog.ACTION_REJECTED, reason)
        sms_text = f'درخواست نمایندگی شما (کد: {app.tracking_code}) پس از بررسی تأیید نشد.'
        if reason:
            sms_text += f'\nدلیل: {reason}'
        _agency_notify_sms(app.phone, sms_text)
        messages.warning(request, 'درخواست رد شد.')

    elif action == 'approve':
        if app.status == AgencyApplication.STATUS_APPROVED:
            messages.info(request, 'این درخواست قبلاً تأیید شده است.')
            return redirect(reverse('agency_application_detail', args=[app_id]))

        user, password = _agency_create_user(app)
        app.status = AgencyApplication.STATUS_APPROVED
        app.created_user = user
        app.reviewed_by = request.user
        app.reviewed_at = timezone.now()
        if note:
            app.staff_note = note
        app.save(update_fields=['status', 'created_user', 'reviewed_by', 'reviewed_at', 'staff_note', 'updated_at'])
        _agency_log(app, request.user, AgencyApplicationLog.ACTION_APPROVED, note)

        sms_text = (
            f'تبریک! درخواست نمایندگی شما (کد: {app.tracking_code}) تأیید شد.\n'
            f'اطلاعات ورود به سامانه:\n'
            f'نام کاربری: {user.username}\n'
            f'رمز عبور: {password}\n'
            f'پس از ورود، رمز عبور خود را تغییر دهید.'
        )
        _agency_notify_sms(app.phone, sms_text)

        _notify_users(
            [request.user],
            '✅ تأیید نمایندگی',
            f'درخواست {app.full_name} تأیید و کاربر {user.username} ایجاد شد.',
            reverse('agency_application_detail', args=[app.pk]),
            category=UserNotification.CATEGORY_SYSTEM,
            actor=request.user,
        )
        messages.success(request, f'درخواست تأیید شد. کاربر {user.username} ایجاد شد و اطلاعات ورود برای متقاضی پیامک شد.')

    else:
        messages.error(request, 'عملیات نامعتبر است.')

    return redirect(reverse('agency_application_detail', args=[app_id]))


# ═══════════════════════════════════════════════════════════════════════════════
#  گارانتی و خدمات پس از فروش
# ═══════════════════════════════════════════════════════════════════════════════

_WARRANTY_STAFF_ROLES = {'warranty', 'warranty_manager'}
_WARRANTY_ALPHABET    = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'

import datetime as _dt
import secrets as _secrets


def _is_warranty_staff(user):
    if user.is_superuser:
        return True
    return _user_role(user) in _WARRANTY_STAFF_ROLES


def _warranty_tracking_code():
    for _ in range(20):
        code = 'WR' + ''.join(_secrets.choice(_WARRANTY_ALPHABET) for _ in range(8))
        if not WarrantyClaim.objects.filter(tracking_code=code).exists():
            return code
    raise RuntimeError('warranty tracking code exhausted')


def _warranty_log(claim, actor, action, note='', visible_to_customer=True):
    WarrantyClaimLog.objects.create(
        claim=claim, actor=actor, action=action,
        note=note, is_visible_to_customer=visible_to_customer,
    )


def _warranty_sms(phone, text):
    try:
        from .sms_service import send_sms
        send_sms(phone, text, purpose='warranty_notify')
    except Exception:
        pass


def _warranty_due_date():
    return timezone.now() + _dt.timedelta(days=3)


def _warranty_notify_staff(claim, title, body):
    staff = User.objects.filter(profile__role__in=['warranty', 'warranty_manager'], is_active=True)
    if claim.assigned_to_id:
        staff = staff | User.objects.filter(pk=claim.assigned_to_id)
    _notify_users(
        list(staff.distinct()),
        title, body,
        reverse('warranty_staff_detail', args=[claim.pk]),
        category=UserNotification.CATEGORY_PAYMENT,
    )


def _save_warranty_files(request, claim, description='تصویر'):
    saved = 0
    for f in request.FILES.getlist('photos'):
        if f.size > 15 * 1024 * 1024:
            continue
        ext = os.path.splitext(f.name)[1].lower()
        if ext not in {'.jpg', '.jpeg', '.png', '.gif', '.webp', '.pdf', '.mp4', '.mov'}:
            continue
        WarrantyClaimFile.objects.create(
            claim=claim, file=f, description=description, uploaded_by=request.user,
        )
        saved += 1
        if saved >= 10:
            break
    if saved:
        _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_FILE,
                      note=f'{saved} فایل بارگذاری شد.', visible_to_customer=True)
    return saved


@login_required
def warranty_new(request):
    profile = getattr(request.user, 'profile', None)
    # Only 'customer' role users submit on their own behalf.
    # sales reps, counterparty agents, and warranty staff submit on behalf of the end-buyer.
    submitter_role = _user_role(request.user)
    is_direct_customer = (not request.user.is_superuser and submitter_role == 'customer')
    is_agent_submission = not is_direct_customer  # sales rep, counterparty, warranty staff, superuser

    def _duplicate_check(sn):
        return WarrantyClaim.objects.filter(
            serial_number=sn,
            status__in=[
                WarrantyClaim.STATUS_SUBMITTED, WarrantyClaim.STATUS_REVIEWING,
                WarrantyClaim.STATUS_INFO_NEEDED, WarrantyClaim.STATUS_APPROVED,
                WarrantyClaim.STATUS_IN_PROGRESS,
            ],
        ).first()

    if request.method == 'POST':
        data = request.POST
        errors = {}
        part_name    = data.get('part_name', '').strip()
        part_model   = data.get('part_model', '').strip()
        serial       = data.get('serial_number', '').strip()
        inv_no       = data.get('invoice_number', '').strip()
        defect       = data.get('defect_description', '').strip()
        c_name       = data.get('claimant_name', '').strip()
        c_phone      = data.get('claimant_phone', '').strip()
        c_email      = data.get('claimant_email', '').strip()
        purchase_str = data.get('purchase_date', '').strip()
        force_submit = data.get('force_submit') == '1'

        if not part_name:   errors['part_name']            = 'نام قطعه الزامی است.'
        if not serial:      errors['serial_number']        = 'شماره سریال الزامی است.'
        if not defect:      errors['defect_description']   = 'شرح خرابی الزامی است.'
        if not c_name:      errors['claimant_name']        = 'نام الزامی است.'
        if not c_phone:     errors['claimant_phone']       = 'شماره موبایل الزامی است.'

        purchase_date = None
        if not purchase_str:
            errors['purchase_date'] = 'تاریخ خرید الزامی است.'
        else:
            try:
                purchase_date = _dt.date.fromisoformat(purchase_str)
            except ValueError:
                errors['purchase_date'] = 'فرمت تاریخ معتبر نیست (YYYY-MM-DD).'

        duplicate_claim = _duplicate_check(serial) if serial else None

        if errors:
            return render(request, 'payments/warranty_new.html', {
                'errors': errors, 'form': data,
                'duplicate': duplicate_claim,
                'is_agent_submission': is_agent_submission,
            })

        if duplicate_claim and not force_submit:
            return render(request, 'payments/warranty_new.html', {
                'errors': {}, 'form': data,
                'duplicate': duplicate_claim, 'show_force': True,
                'is_agent_submission': is_agent_submission,
            })

        # For direct customers: link claim to their account.
        # For agents/staff: user=None (end-buyer may not have an account); track via submitted_by.
        claim = WarrantyClaim.objects.create(
            user=request.user if is_direct_customer else None,
            submitted_by=request.user,
            claimant_name=c_name, claimant_phone=c_phone, claimant_email=c_email,
            part_name=part_name, part_model=part_model,
            serial_number=serial, purchase_date=purchase_date, invoice_number=inv_no,
            defect_description=defect,
            tracking_code=_warranty_tracking_code(),
            due_date=_warranty_due_date(),
        )
        _save_warranty_files(request, claim, description='تصویر اولیه')
        _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_SUBMITTED, visible_to_customer=True)
        _warranty_notify_staff(
            claim, '🛡️ درخواست گارانتی جدید',
            f'درخواست گارانتی جدید برای «{part_name}» (سریال: {serial}) ثبت شد.',
        )
        _warranty_sms(
            c_phone,
            f'درخواست گارانتی شما ثبت شد.\nکد پیگیری: {claim.tracking_code}\n'
            'وضعیت را از طریق سامانه پیگیری کنید.',
        )
        messages.success(request, f'درخواست گارانتی با کد {claim.tracking_code} ثبت شد.')
        return redirect('warranty_claim_detail', claim_id=claim.pk)

    prefill = {}
    # Only pre-fill claimant info for direct customers submitting for themselves.
    if is_direct_customer and profile:
        prefill['claimant_name']  = f"{profile.first_name} {profile.last_name}".strip()
        prefill['claimant_phone'] = profile.mobile or profile.phone or ''
    return render(request, 'payments/warranty_new.html', {
        'form': prefill,
        'errors': {},
        'is_agent_submission': is_agent_submission,
    })


@login_required
def warranty_my_claims(request):
    # Show claims the user owns (direct customer) OR submitted on someone else's behalf (agent/staff).
    from django.db.models import Q as _Q
    claims = WarrantyClaim.objects.filter(
        _Q(user=request.user) | _Q(submitted_by=request.user)
    ).distinct().order_by('-created_at')
    status_f = request.GET.get('status', '')
    if status_f:
        claims = claims.filter(status=status_f)
    paginator = Paginator(claims, 20)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'payments/warranty_my_claims.html', {
        'page': page,
        'status_choices': WarrantyClaim.STATUS_CHOICES,
        'status_filter': status_f,
    })


@login_required
def warranty_claim_detail(request, claim_id):
    claim = get_object_or_404(WarrantyClaim, pk=claim_id)
    # Allow access to: the customer who owns the claim, the person who submitted it, and warranty staff.
    can_access = (
        claim.user_id == request.user.id or
        claim.submitted_by_id == request.user.id or
        _is_warranty_staff(request.user)
    )
    if not can_access:
        return HttpResponseForbidden('دسترسی ندارید.')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'reply' and claim.status == WarrantyClaim.STATUS_INFO_NEEDED:
            reply = request.POST.get('reply', '').strip()
            if not reply:
                messages.error(request, 'متن پاسخ را وارد کنید.')
            else:
                claim.customer_reply = reply
                claim.status = WarrantyClaim.STATUS_REVIEWING
                claim.save(update_fields=['customer_reply', 'status', 'updated_at'])
                _save_warranty_files(request, claim, description='فایل تکمیلی مشتری')
                _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_CUSTOMER_REPLY, note=reply)
                _warranty_notify_staff(
                    claim, '💬 پاسخ مشتری',
                    f'مشتری پاسخ اطلاعات تکمیلی برای گارانتی #{claim.id} را ارسال کرد.',
                )
                messages.success(request, 'پاسخ شما ثبت شد.')

        elif action == 'rate' and claim.status in {
            WarrantyClaim.STATUS_RESOLVED, WarrantyClaim.STATUS_CLOSED
        }:
            try:
                rating = max(1, min(5, int(request.POST.get('rating', '0'))))
            except ValueError:
                rating = None
            feedback = request.POST.get('feedback', '').strip()
            if rating:
                claim.customer_rating = rating
                claim.customer_feedback = feedback
                claim.save(update_fields=['customer_rating', 'customer_feedback', 'updated_at'])
                _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_RATED,
                              note=f'امتیاز: {rating}/5 — {feedback}')
                messages.success(request, 'امتیاز شما ثبت شد.')

        elif action == 'add_file':
            saved = _save_warranty_files(request, claim, description='فایل تکمیلی')
            if saved:
                messages.success(request, f'{saved} فایل بارگذاری شد.')
            else:
                messages.error(request, 'فایلی بارگذاری نشد.')

        return redirect('warranty_claim_detail', claim_id=claim.pk)

    return render(request, 'payments/warranty_claim_detail.html', {
        'claim': claim,
        'logs': claim.logs.filter(is_visible_to_customer=True),
        'files': claim.files.all(),
    })


def warranty_track(request):
    result = None
    error  = None
    if request.method == 'POST':
        query = request.POST.get('query', '').strip().upper()
        if not query:
            error = 'کد پیگیری را وارد کنید.'
        else:
            result = WarrantyClaim.objects.filter(tracking_code=query).first()
            if not result:
                error = f'درخواستی با کد «{query}» یافت نشد.'
    return render(request, 'payments/warranty_track.html', {'result': result, 'error': error})


@login_required
def warranty_staff_list(request):
    if not _is_warranty_staff(request.user):
        return HttpResponseForbidden('دسترسی ندارید.')

    qs = WarrantyClaim.objects.select_related('user', 'assigned_to').all()
    status_f   = request.GET.get('status', '')
    priority_f = request.GET.get('priority', '')
    search     = request.GET.get('q', '').strip()
    assigned_f = request.GET.get('assigned', '')
    overdue_f  = request.GET.get('overdue', '')

    if status_f:   qs = qs.filter(status=status_f)
    if priority_f: qs = qs.filter(priority=priority_f)
    if search:
        qs = qs.filter(
            Q(tracking_code__icontains=search) | Q(claimant_name__icontains=search) |
            Q(claimant_phone__icontains=search) | Q(part_name__icontains=search) |
            Q(serial_number__icontains=search)
        )
    if assigned_f == 'me':
        qs = qs.filter(assigned_to=request.user)
    elif assigned_f == 'unassigned':
        qs = qs.filter(assigned_to__isnull=True)
    if overdue_f == '1':
        qs = qs.filter(due_date__lt=timezone.now(), status__in=[
            WarrantyClaim.STATUS_SUBMITTED, WarrantyClaim.STATUS_REVIEWING,
            WarrantyClaim.STATUS_INFO_NEEDED, WarrantyClaim.STATUS_APPROVED,
            WarrantyClaim.STATUS_IN_PROGRESS,
        ])

    open_statuses = [
        WarrantyClaim.STATUS_SUBMITTED, WarrantyClaim.STATUS_REVIEWING,
        WarrantyClaim.STATUS_INFO_NEEDED, WarrantyClaim.STATUS_APPROVED,
        WarrantyClaim.STATUS_IN_PROGRESS,
    ]
    status_counts = [
        (val, label, WarrantyClaim.objects.filter(status=val).count())
        for val, label in WarrantyClaim.STATUS_CHOICES
    ]
    open_count    = WarrantyClaim.objects.filter(status__in=open_statuses).count()
    overdue_count = WarrantyClaim.objects.filter(
        due_date__lt=timezone.now(), status__in=open_statuses).count()

    paginator = Paginator(qs, 25)
    page = paginator.get_page(request.GET.get('page'))
    return render(request, 'payments/warranty_staff_list.html', {
        'page': page,
        'status_choices':   WarrantyClaim.STATUS_CHOICES,
        'priority_choices': WarrantyClaim.PRIORITY_CHOICES,
        'status_counts':    status_counts,
        'open_count':       open_count,
        'overdue_count':    overdue_count,
        'status_filter':    status_f,
        'priority_filter':  priority_f,
        'search':           search,
        'assigned_filter':  assigned_f,
        'overdue_filter':   overdue_f,
    })


@login_required
def warranty_staff_detail(request, claim_id):
    if not _is_warranty_staff(request.user):
        return HttpResponseForbidden('دسترسی ندارید.')
    claim = get_object_or_404(WarrantyClaim, pk=claim_id)
    warranty_staff_users = list(
        User.objects.filter(profile__role__in=['warranty', 'warranty_manager'], is_active=True)
        .select_related('profile')
    )
    return render(request, 'payments/warranty_staff_detail.html', {
        'claim': claim,
        'logs': claim.logs.all(),
        'files': claim.files.all(),
        'staff_users': warranty_staff_users,
        'resolution_choices': WarrantyClaim.RESOLUTION_CHOICES,
        'priority_choices':   WarrantyClaim.PRIORITY_CHOICES,
    })


@login_required
@require_POST
def warranty_staff_action(request, claim_id):
    if not _is_warranty_staff(request.user):
        return HttpResponseForbidden('دسترسی ندارید.')
    claim  = get_object_or_404(WarrantyClaim, pk=claim_id)
    action = request.POST.get('action', '').strip()
    note   = request.POST.get('note', '').strip()
    now    = timezone.now()

    if action == 'start_review':
        if claim.status == WarrantyClaim.STATUS_SUBMITTED:
            claim.status = WarrantyClaim.STATUS_REVIEWING
            claim.reviewed_at = now
            claim.save(update_fields=['status', 'reviewed_at', 'updated_at'])
            _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_REVIEWING)
            _warranty_sms(claim.claimant_phone,
                          f'درخواست گارانتی {claim.tracking_code} در دست بررسی قرار گرفت.')
            messages.success(request, 'بررسی شروع شد.')

    elif action == 'info_needed':
        if not note:
            messages.error(request, 'توضیح درخواست اطلاعات الزامی است.')
        else:
            claim.status = WarrantyClaim.STATUS_INFO_NEEDED
            claim.info_request_note = note
            claim.save(update_fields=['status', 'info_request_note', 'updated_at'])
            _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_INFO_NEEDED, note=note)
            _warranty_sms(claim.claimant_phone,
                          f'گارانتی {claim.tracking_code}: اطلاعات تکمیلی لازم است.\n{note}')
            messages.success(request, 'درخواست اطلاعات ارسال شد.')

    elif action == 'approve':
        claim.status = WarrantyClaim.STATUS_APPROVED
        claim.save(update_fields=['status', 'updated_at'])
        _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_APPROVED, note=note)
        _warranty_sms(claim.claimant_phone,
                      f'گارانتی {claim.tracking_code} تأیید شد. به زودی تماس خواهیم گرفت.')
        messages.success(request, 'گارانتی تأیید شد.')

    elif action == 'in_progress':
        claim.status = WarrantyClaim.STATUS_IN_PROGRESS
        claim.save(update_fields=['status', 'updated_at'])
        _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_IN_PROGRESS, note=note)
        _warranty_sms(claim.claimant_phone,
                      f'گارانتی {claim.tracking_code} در حال پردازش / تعمیر است.')
        messages.success(request, 'وضعیت «در حال پردازش» ثبت شد.')

    elif action == 'resolve':
        resolution_type = request.POST.get('resolution_type', '').strip()
        resolution_note = request.POST.get('resolution_note', '').strip()
        if not resolution_type:
            messages.error(request, 'نوع رفع مسئله را انتخاب کنید.')
        else:
            claim.status          = WarrantyClaim.STATUS_RESOLVED
            claim.resolution_type = resolution_type
            claim.resolution_note = resolution_note or note
            claim.resolved_at     = now
            claim.save(update_fields=['status', 'resolution_type', 'resolution_note',
                                      'resolved_at', 'updated_at'])
            res_label = dict(WarrantyClaim.RESOLUTION_CHOICES).get(resolution_type, resolution_type)
            _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_RESOLVED,
                          note=f'{res_label}: {resolution_note}'.strip())
            _warranty_sms(claim.claimant_phone,
                          f'گارانتی {claim.tracking_code} رفع شد ({res_label}). '
                          'لطفاً رضایت خود را در سامانه ثبت کنید.')
            messages.success(request, 'درخواست رفع شد.')

    elif action == 'reject':
        if not note:
            messages.error(request, 'دلیل رد الزامی است.')
        else:
            claim.status           = WarrantyClaim.STATUS_REJECTED
            claim.rejection_reason = note
            claim.save(update_fields=['status', 'rejection_reason', 'updated_at'])
            _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_REJECTED, note=note)
            _warranty_sms(claim.claimant_phone,
                          f'گارانتی {claim.tracking_code} رد شد.\nدلیل: {note}')
            messages.success(request, 'درخواست رد شد.')

    elif action == 'close':
        claim.status = WarrantyClaim.STATUS_CLOSED
        claim.save(update_fields=['status', 'updated_at'])
        _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_CLOSED,
                      note=note, visible_to_customer=False)
        messages.success(request, 'درخواست بسته شد.')

    elif action == 'assign':
        assign_to = request.POST.get('assign_to', '').strip()
        if assign_to:
            try:
                staff_user = User.objects.get(
                    pk=int(assign_to), profile__role__in=['warranty', 'warranty_manager'])
                claim.assigned_to = staff_user
                claim.save(update_fields=['assigned_to', 'updated_at'])
                _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_ASSIGNED,
                              note=f'تخصیص به {staff_user.get_full_name() or staff_user.username}',
                              visible_to_customer=False)
                messages.success(request, 'تخصیص انجام شد.')
            except (User.DoesNotExist, ValueError):
                messages.error(request, 'کارشناس معتبر نیست.')
        else:
            claim.assigned_to = None
            claim.save(update_fields=['assigned_to', 'updated_at'])
            messages.success(request, 'تخصیص برداشته شد.')

    elif action == 'priority':
        new_priority = request.POST.get('priority', '').strip()
        if new_priority in {v for v, _ in WarrantyClaim.PRIORITY_CHOICES}:
            old_label = claim.get_priority_display()
            claim.priority = new_priority
            claim.save(update_fields=['priority', 'updated_at'])
            _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_PRIORITY,
                          note=f'{old_label} → {claim.get_priority_display()}',
                          visible_to_customer=False)
            messages.success(request, 'اولویت تغییر کرد.')

    elif action == 'note':
        if note:
            visible = request.POST.get('note_visible', '1') == '1'
            _warranty_log(claim, request.user, WarrantyClaimLog.ACTION_NOTE,
                          note=note, visible_to_customer=visible)
            messages.success(request, 'یادداشت ثبت شد.')

    elif action == 'add_file':
        saved = _save_warranty_files(request, claim, description=note or 'فایل کارشناس')
        if saved:
            messages.success(request, f'{saved} فایل بارگذاری شد.')
        else:
            messages.error(request, 'فایلی بارگذاری نشد.')

    return redirect('warranty_staff_detail', claim_id=claim.pk)

