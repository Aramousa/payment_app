import jdatetime
import os
import uuid
from io import BytesIO
from unittest.mock import patch
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core import mail
from django.test import override_settings
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from .forms import DailyPaymentAssignmentForm, InvoiceUploadForm, PriceListUploadForm, ProformaInvoiceForm
from .models import CustomerOrder, CustomerSalesAssignment, DailyPaymentAssignment, DailyPaymentPlan, InvoiceExtractionJob, InvoiceRecord, PaymentActivityLog, PaymentReceipt, PaymentRecord, PriceList, ProfileChangeRequest, ProformaInvoice, ProformaInvoiceLog, SystemActivityLog, UserNotification
from .views import _staff_status_choices_for_role


class InvoiceFlowTests(TestCase):
    def setUp(self):
        self.commercial_user = User.objects.create_user(
            username='commercial1',
            password='pass1234',
            first_name='کاربر',
            last_name='بازرگانی',
        )
        self.commercial_user.profile.phone = '09120000001'
        self.commercial_user.profile.role = 'commercial'
        self.commercial_user.profile.can_upload_invoices = True
        self.commercial_user.profile.can_view_invoices = True
        self.commercial_user.profile.save()

        self.finance_user = User.objects.create_user(
            username='finance1',
            password='pass1234',
            first_name='Ú©Ø§Ø±Ø¨Ø±',
            last_name='Ù…Ø§Ù„ÛŒ',
        )
        self.finance_user.profile.role = 'finance'
        self.finance_user.profile.force_password_change = False
        self.finance_user.profile.save()

        self.data_entry_user = User.objects.create_user(
            username='dataentry1',
            password='pass1234',
            first_name='کاربر',
            last_name='ورود اطلاعات',
        )
        self.data_entry_user.profile.role = 'data_entry'
        self.data_entry_user.profile.can_edit_payment_details = True
        self.data_entry_user.profile.force_password_change = False
        self.data_entry_user.profile.save()

        self.sales_user = User.objects.create_user(
            username='sales1',
            password='pass1234',
            first_name='کاربر',
            last_name='فروش',
        )
        self.sales_user.profile.role = 'sales'
        self.sales_user.profile.force_password_change = False
        self.sales_user.profile.save()

        self.customer_user = User.objects.create_user(
            username='customer1',
            password='pass1234',
            first_name='علی',
            last_name='مشتری',
        )
        self.customer_profile = self.customer_user.profile
        self.customer_profile.phone = '09120000002'
        self.customer_profile.organization = 'شرکت آلفا'
        self.customer_profile.city = 'تهران'
        self.customer_profile.role = 'customer'
        self.customer_profile.force_password_change = False
        self.customer_profile.save()

        self.other_customer = User.objects.create_user(
            username='customer2',
            password='pass1234',
            first_name='رضا',
            last_name='دیگری',
        )
        self.other_customer.profile.phone = '09120000003'
        self.other_customer.profile.organization = 'شرکت بتا'
        self.other_customer.profile.city = 'اصفهان'
        self.other_customer.profile.role = 'customer'
        self.other_customer.profile.force_password_change = False
        self.other_customer.profile.save()

    def test_commercial_user_can_upload_invoice_for_customer(self):
        self.client.login(username='commercial1', password='pass1234')
        response = self.client.post(
            reverse('invoices_dashboard'),
            {
                'customer': str(self.customer_profile.id),
                'invoice_date': '1405/02/08',
                'amount': '2,500,000',
                'invoice_number': 'INV-1001',
                'reference_number': 'REF-1001',
                'customer_visible_note': 'توضیح برای مشتری',
                'internal_note': 'یادداشت داخلی',
                'confirm_assignment': 'on',
                'attachment': SimpleUploadedFile('invoice.pdf', b'%PDF-1.4 sample', content_type='application/pdf'),
            },
        )

        self.assertEqual(response.status_code, 302)
        invoice = InvoiceRecord.objects.get(reference_number='REF-1001')
        self.assertEqual(invoice.customer, self.customer_user)
        self.assertEqual(invoice.uploaded_by, self.commercial_user)
        self.assertEqual(invoice.amount, 2500000)
        self.assertEqual(invoice.customer_visible_note, 'توضیح برای مشتری')
        self.assertEqual(invoice.internal_note, 'یادداشت داخلی')

    def test_inactive_or_suspended_customers_are_excluded_from_operational_customer_choices(self):
        self.other_customer.is_active = False
        self.other_customer.save(update_fields=['is_active'])
        self.customer_profile.suspended = True
        self.customer_profile.save(update_fields=['suspended'])

        forms = [
            DailyPaymentAssignmentForm(),
            InvoiceUploadForm(),
            PriceListUploadForm(),
            ProformaInvoiceForm(),
        ]
        for form in forms:
            field_name = 'customer' if 'customer' in form.fields else 'customers'
            self.assertEqual(list(form.fields[field_name].queryset), [])

    def test_invoice_pdf_parse_preview_suggests_form_fields(self):
        self.client.login(username='commercial1', password='pass1234')
        extracted_text = (
            'شماره فاکتور: INV-2040\n'
            'تاریخ فاکتور: 1405/02/08\n'
            'مبلغ کل: 2,500,000 ریال\n'
            'شماره حواله: REF-2040\n'
        )

        with patch('payments.invoice_extraction.extract_pdf_text_or_ocr', return_value=(extracted_text, 'pymupdf_text', [])):
            response = self.client.post(
                reverse('invoice_parse_preview'),
                {'attachment': SimpleUploadedFile('invoice.pdf', b'%PDF-1.4 sample', content_type='application/pdf')},
            )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['ok'])
        self.assertEqual(payload['fields']['invoice_number'], 'INV-2040')
        self.assertEqual(payload['fields']['invoice_date'], '1405/02/08')
        self.assertEqual(payload['fields']['amount'], '2500000')
        self.assertEqual(payload['fields']['reference_number'], 'REF-2040')
        job = InvoiceExtractionJob.objects.get(requested_by=self.commercial_user, status=InvoiceExtractionJob.STATUS_DONE)
        self.assertEqual(job.text_source, 'pymupdf_text')
        status_response = self.client.get(reverse('invoice_parse_status', args=[job.id]))
        self.assertEqual(status_response.status_code, 200)
        self.assertEqual(status_response.json()['fields']['invoice_number'], 'INV-2040')

    def test_invoice_pdf_parse_preview_requires_upload_permission(self):
        self.client.login(username='customer1', password='pass1234')
        response = self.client.post(
            reverse('invoice_parse_preview'),
            {'attachment': SimpleUploadedFile('invoice.pdf', b'%PDF-1.4 sample', content_type='application/pdf')},
        )
        self.assertEqual(response.status_code, 403)

    def test_invoice_image_parse_preview_suggests_form_fields(self):
        self.client.login(username='commercial1', password='pass1234')
        extracted_text = (
            'شماره فاکتور: IMG-2040\n'
            'تاریخ صدور: 1405/03/01\n'
            'جمع کل: 3,750,000 ریال\n'
            'کد رهگیری: IMG-REF\n'
        )

        with patch('payments.invoice_extraction.extract_image_text', return_value=(extracted_text, 'paddleocr_image', [])):
            response = self.client.post(
                reverse('invoice_parse_preview'),
                {'attachment': SimpleUploadedFile('invoice.jpg', b'fake-image', content_type='image/jpeg')},
            )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['ok'])
        self.assertEqual(payload['fields']['invoice_number'], 'IMG-2040')
        self.assertEqual(payload['fields']['invoice_date'], '1405/03/01')
        self.assertEqual(payload['fields']['amount'], '3750000')
        self.assertEqual(payload['fields']['reference_number'], 'IMG-REF')

    def test_invoice_parse_preview_does_not_guess_unlabeled_amount(self):
        self.client.login(username='commercial1', password='pass1234')
        extracted_text = (
            'شماره فاکتور: INV-AMBIG\n'
            'تاریخ فاکتور: 1405/03/01\n'
            'ردیف 1 900,000\n'
            'ردیف 2 8,750,000\n'
        )

        with patch('payments.invoice_extraction.extract_pdf_text_or_ocr', return_value=(extracted_text, 'pymupdf_text', [])):
            response = self.client.post(
                reverse('invoice_parse_preview'),
                {'attachment': SimpleUploadedFile('invoice.pdf', b'%PDF-1.4 sample', content_type='application/pdf')},
            )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertNotIn('amount', payload['fields'])

    def test_invoice_parse_preview_suggests_unique_currency_amount(self):
        self.client.login(username='commercial1', password='pass1234')
        extracted_text = (
            'شماره فاکتور: INV-CURRENCY\n'
            'تاریخ فاکتور: 1405/03/01\n'
            '2,450,000 ریال\n'
        )

        with patch('payments.invoice_extraction.extract_pdf_text_or_ocr', return_value=(extracted_text, 'pymupdf_text', [])):
            response = self.client.post(
                reverse('invoice_parse_preview'),
                {'attachment': SimpleUploadedFile('invoice.pdf', b'%PDF-1.4 sample', content_type='application/pdf')},
            )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['fields']['amount'], '2450000')

    def test_customer_view_marks_invoice_seen_and_allows_note(self):
        invoice = InvoiceRecord.objects.create(
            customer=self.customer_user,
            uploaded_by=self.commercial_user,
            amount=900000,
            invoice_date=jdatetime.date(1405, 2, 8),
            reference_number='REF-2001',
            attachment=SimpleUploadedFile('invoice.pdf', b'%PDF-1.4 sample', content_type='application/pdf'),
            customer_visible_note='متن قابل مشاهده',
            internal_note='متن داخلی',
        )

        self.client.login(username='customer1', password='pass1234')
        detail_url = reverse('invoice_detail', args=[invoice.id])

        response = self.client.get(detail_url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'متن قابل مشاهده')
        self.assertContains(response, 'مشاهده فایل فاکتور')
        self.assertContains(response, reverse('invoice_file', args=[invoice.id]))
        self.assertNotContains(response, 'متن داخلی')
        invoice.refresh_from_db()
        self.assertIsNotNone(invoice.customer_seen_at)

        response = self.client.post(detail_url, {'customer_note': 'یادداشت تست'})
        self.assertEqual(response.status_code, 302)
        invoice.refresh_from_db()
        self.assertEqual(invoice.customer_note, 'یادداشت تست')
        self.assertIsNotNone(invoice.customer_note_updated_at)

    def test_customer_cannot_access_other_customer_invoice(self):
        invoice = InvoiceRecord.objects.create(
            customer=self.other_customer,
            uploaded_by=self.commercial_user,
            amount=500000,
            invoice_date=jdatetime.date(1405, 2, 8),
            reference_number='REF-3001',
            attachment=SimpleUploadedFile('invoice.pdf', b'%PDF-1.4 sample', content_type='application/pdf'),
        )

        self.client.login(username='customer1', password='pass1234')
        response = self.client.get(reverse('invoice_detail', args=[invoice.id]))
        self.assertEqual(response.status_code, 403)

    def test_customer_invoice_list_only_contains_own_invoices(self):
        own_invoice = InvoiceRecord.objects.create(
            customer=self.customer_user,
            uploaded_by=self.commercial_user,
            amount=700000,
            invoice_date=jdatetime.date(1405, 2, 8),
            invoice_number='INV-OWN',
            reference_number='REF-OWN',
            attachment=SimpleUploadedFile('own.pdf', b'%PDF-1.4 own', content_type='application/pdf'),
        )
        other_invoice = InvoiceRecord.objects.create(
            customer=self.other_customer,
            uploaded_by=self.commercial_user,
            amount=800000,
            invoice_date=jdatetime.date(1405, 2, 8),
            invoice_number='INV-OTHER',
            reference_number='REF-OTHER',
            attachment=SimpleUploadedFile('other.pdf', b'%PDF-1.4 other', content_type='application/pdf'),
        )

        self.client.login(username='customer1', password='pass1234')
        response = self.client.get(reverse('invoices_dashboard'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, own_invoice.invoice_number)
        self.assertNotContains(response, other_invoice.invoice_number)

    def test_customer_home_shows_operational_sections(self):
        invoice = InvoiceRecord.objects.create(
            customer=self.customer_user,
            uploaded_by=self.commercial_user,
            amount=700000,
            invoice_date=jdatetime.date(1405, 2, 8),
            invoice_number='HOME-INV',
            reference_number='HOME-REF',
            attachment=SimpleUploadedFile('home-invoice.pdf', b'%PDF-1.4 invoice', content_type='application/pdf'),
        )
        payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=500000,
            pay_date=jdatetime.date(1405, 2, 8),
            tracking_code='HOME-PAY',
        )
        price_list = PriceList.objects.create(
            customer=self.customer_user,
            uploaded_by=self.sales_user,
            title='HOME-PRICE',
            file=SimpleUploadedFile('home-price.pdf', b'%PDF-1.4 price', content_type='application/pdf'),
        )
        proforma = ProformaInvoice.objects.create(
            customer=self.customer_user,
            issued_by=self.commercial_user,
            title='HOME-PROFORMA',
            valid_until=jdatetime.date(1405, 12, 29),
            file=SimpleUploadedFile('home-proforma.pdf', b'%PDF-1.4 proforma', content_type='application/pdf'),
        )

        self.client.login(username='customer1', password='pass1234')
        response = self.client.get(reverse('submit'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'فاکتورهای من')
        self.assertContains(response, 'فیش‌های واریزی من')
        self.assertContains(response, 'لیست قیمت اختصاصی')
        self.assertContains(response, 'پیش‌فاکتورهای من')
        self.assertContains(response, invoice.invoice_number)
        self.assertContains(response, payment.tracking_code)
        self.assertContains(response, price_list.title)
        self.assertContains(response, proforma.title)
        self.assertContains(response, reverse('invoice_detail', args=[invoice.id]))
        self.assertContains(response, reverse('payment_timeline', args=[payment.id]))
        self.assertContains(response, reverse('payment_create'))
        self.assertContains(response, reverse('price_list_file', args=[price_list.id]))
        self.assertContains(response, 'dashboard-file-preview-link')
        self.assertContains(response, reverse('proforma_detail', args=[proforma.id]))

        response = self.client.get(reverse('payment_create'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="payment-form"')
        self.assertContains(response, 'name="receipt_images"')
        self.assertContains(response, 'name="amount"')
        self.assertContains(response, 'name="tracking_code"')
        self.assertContains(response, 'name="pay_date"')
        self.assertContains(response, payment.tracking_code)
        self.assertContains(response, reverse('payment_timeline', args=[payment.id]))

    def test_customer_cannot_access_other_customer_invoice_file(self):
        invoice = InvoiceRecord.objects.create(
            customer=self.other_customer,
            uploaded_by=self.commercial_user,
            amount=500000,
            invoice_date=jdatetime.date(1405, 2, 8),
            invoice_number='INV-FILE-OTHER',
            reference_number='REF-FILE-OTHER',
            attachment=SimpleUploadedFile('other-file.pdf', b'%PDF-1.4 other', content_type='application/pdf'),
        )

        self.client.login(username='customer1', password='pass1234')
        response = self.client.get(reverse('invoice_file', args=[invoice.id]))
        self.assertEqual(response.status_code, 403)

    def test_customer_payment_list_and_receipt_file_are_scoped_to_owner(self):
        own_payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='علی',
            last_name='مشتری',
            organization='شرکت آلفا',
            city='تهران',
            phone='09120000002',
            amount=100000,
            pay_date=jdatetime.date(1405, 2, 8),
            tracking_code='PAY-OWN',
        )
        other_payment = PaymentRecord.objects.create(
            user=self.other_customer,
            first_name='رضا',
            last_name='دیگری',
            organization='شرکت بتا',
            city='اصفهان',
            phone='09120000003',
            amount=200000,
            pay_date=jdatetime.date(1405, 2, 8),
            tracking_code='PAY-OTHER',
        )
        other_receipt = PaymentReceipt.objects.create(
            payment=other_payment,
            image=SimpleUploadedFile('other-receipt.pdf', b'%PDF-1.4 receipt', content_type='application/pdf'),
            file_hash='other-receipt-hash',
        )

        self.client.login(username='customer1', password='pass1234')
        response = self.client.get(reverse('submit'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, own_payment.tracking_code)
        self.assertNotContains(response, other_payment.tracking_code)

        response = self.client.get(reverse('payment_create'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, own_payment.tracking_code)
        self.assertNotContains(response, other_payment.tracking_code)

        response = self.client.get(reverse('payment_timeline', args=[other_payment.id]))
        self.assertEqual(response.status_code, 403)

        response = self.client.get(reverse('receipt_file', args=[other_receipt.id]))
        self.assertEqual(response.status_code, 403)

    def test_uploaded_receipt_filename_is_server_generated_ascii(self):
        payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=100000,
            pay_date=jdatetime.date(1405, 2, 8),
            tracking_code='UNIQUE-FILE',
        )
        receipt = PaymentReceipt.objects.create(
            payment=payment,
            image=SimpleUploadedFile('فیش واریزی.pdf', b'%PDF-1.4 receipt', content_type='application/pdf'),
            file_hash='unique-file-hash',
        )

        basename = os.path.basename(receipt.image.name)
        self.assertTrue(receipt.image.name.startswith('receipts/paymentreceipt_user'))
        self.assertTrue(basename.endswith('.pdf'))
        self.assertNotIn('فیش', basename)
        basename.encode('ascii')

    def test_payment_workflow_active_queue_history_and_logs(self):
        payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=100000,
            pay_date=jdatetime.date(1405, 2, 8),
            tracking_code='WF-APPROVE',
        )

        self.client.login(username='commercial1', password='pass1234')
        response = self.client.get(reverse('submit'))
        self.assertEqual(response.status_code, 200)
        payment.refresh_from_db()
        self.assertEqual(payment.status, PaymentRecord.STATUS_COMMERCIAL_REVIEW)
        self.assertTrue(
            PaymentActivityLog.objects.filter(
                payment=payment,
                actor=self.commercial_user,
                action=PaymentActivityLog.ACTION_STATUS_CHANGED,
                to_status=PaymentRecord.STATUS_COMMERCIAL_REVIEW,
            ).exists()
        )
        self.assertTrue(
            PaymentActivityLog.objects.filter(
                payment=payment,
                actor=self.commercial_user,
                action=PaymentActivityLog.ACTION_VIEWED,
            ).exists()
        )

        response = self.client.post(
            reverse('staff_update_status', args=[payment.id]),
            {'status': PaymentRecord.STATUS_APPROVED, 'note': '', 'next': reverse('submit')},
        )
        self.assertEqual(response.status_code, 302)
        payment.refresh_from_db()
        self.assertEqual(payment.status, PaymentRecord.STATUS_APPROVED)

        response = self.client.get(reverse('submit'))
        self.assertNotContains(response, 'WF-APPROVE')
        response = self.client.get(reverse('payment_history'))
        self.assertContains(response, 'WF-APPROVE')

        self.client.logout()
        self.client.login(username='finance1', password='pass1234')
        response = self.client.get(reverse('submit'))
        self.assertContains(response, 'WF-APPROVE')

        response = self.client.post(
            reverse('staff_update_status', args=[payment.id]),
            {'status': PaymentRecord.STATUS_FINAL_APPROVED, 'note': '', 'next': reverse('submit')},
        )
        self.assertEqual(response.status_code, 302)
        payment.refresh_from_db()
        self.assertEqual(payment.status, PaymentRecord.STATUS_FINAL_APPROVED)

        response = self.client.get(reverse('submit'))
        self.assertNotContains(response, 'WF-APPROVE')
        response = self.client.get(reverse('payment_history'))
        self.assertContains(response, 'WF-APPROVE')

    def test_staff_status_choices_for_generic_staff_role_are_not_empty(self):
        choices = _staff_status_choices_for_role('staff')
        self.assertTrue(len(choices) > 0)
        self.assertIn((PaymentRecord.STATUS_APPROVED, 'ثبت بازرگانی'), choices)

    def test_dashboard_notification_bell_is_visible_for_staff_roles(self):
        self.client.login(username='commercial1', password='pass1234')
        response = self.client.get(reverse('submit'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'notification-bell')

        self.client.logout()
        self.client.login(username='finance1', password='pass1234')
        response = self.client.get(reverse('submit'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'notification-bell')

    def test_status_changes_create_user_notifications_and_feed_marks_read(self):
        payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=100000,
            pay_date=jdatetime.date(1405, 2, 8),
            tracking_code='NOTIFY-INCOMPLETE',
            status=PaymentRecord.STATUS_COMMERCIAL_REVIEW,
        )

        self.client.login(username='commercial1', password='pass1234')
        response = self.client.post(
            reverse('staff_update_status', args=[payment.id]),
            {'status': PaymentRecord.STATUS_INCOMPLETE, 'note': 'Needs correction', 'next': reverse('submit')},
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            UserNotification.objects.filter(
                user=self.customer_user,
                category=UserNotification.CATEGORY_PAYMENT,
                title='تغییر وضعیت فیش',
                is_read=False,
            ).exists()
        )

        self.client.logout()
        self.client.login(username='customer1', password='pass1234')
        response = self.client.get(reverse('notifications_feed'))
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['unread_count'], 1)
        self.assertEqual(payload['items'][0]['title'], 'تغییر وضعیت فیش')

        response = self.client.post(reverse('notifications_mark_read'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(UserNotification.objects.filter(user=self.customer_user, is_read=False).count(), 0)

    def test_customer_sees_commercial_and_final_approval_as_distinct_statuses(self):
        commercial_payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=100000,
            pay_date=jdatetime.date(1405, 2, 8),
            tracking_code='CUSTOMER-COMMERCIAL',
            status=PaymentRecord.STATUS_APPROVED,
        )
        final_payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=200000,
            pay_date=jdatetime.date(1405, 2, 9),
            tracking_code='CUSTOMER-FINAL',
            status=PaymentRecord.STATUS_FINAL_APPROVED,
        )

        self.client.login(username='customer1', password='pass1234')
        response = self.client.get(reverse('payment_create'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, commercial_payment.tracking_code)
        self.assertContains(response, final_payment.tracking_code)
        self.assertContains(response, 'ثبت بازرگانی')
        self.assertContains(response, 'تایید نهایی')
        self.assertContains(response, 'flag-orange')
        self.assertContains(response, 'flag-green')

    def test_customer_timeline_hides_staff_identity_and_internal_steps(self):
        payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=100000,
            pay_date=jdatetime.date(1405, 2, 8),
            tracking_code='CUSTOMER-TIMELINE',
            status=PaymentRecord.STATUS_INCOMPLETE,
        )
        _log_activity = __import__('payments.views', fromlist=['_log_activity'])._log_activity
        _log_activity(payment, self.customer_user, PaymentActivityLog.ACTION_CREATED, to_status=PaymentRecord.STATUS_PENDING)
        _log_activity(payment, self.commercial_user, PaymentActivityLog.ACTION_VIEWED, note='internal view')
        _log_activity(
            payment,
            self.commercial_user,
            PaymentActivityLog.ACTION_STATUS_CHANGED,
            from_status=PaymentRecord.STATUS_PENDING,
            to_status=PaymentRecord.STATUS_COMMERCIAL_REVIEW,
        )
        _log_activity(
            payment,
            self.commercial_user,
            PaymentActivityLog.ACTION_STATUS_CHANGED,
            from_status=PaymentRecord.STATUS_COMMERCIAL_REVIEW,
            to_status=PaymentRecord.STATUS_INCOMPLETE,
            note='تصویر فیش واضح نیست',
        )

        self.client.login(username='customer1', password='pass1234')
        response = self.client.get(reverse('payment_timeline', args=[payment.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'فیش توسط مشتری ثبت شد.')
        self.assertContains(response, 'سند مشاهده شد.')
        self.assertContains(response, 'نقص مدارک ثبت شد.')
        self.assertContains(response, 'تصویر فیش واضح نیست')
        self.assertNotContains(response, self.commercial_user.username)
        self.assertNotContains(response, 'بررسی بازرگانی')
        self.assertNotContains(response, 'وضعیت سند را')

    def test_payment_list_shows_commercial_and_finance_status_columns_for_customer(self):
        PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=100000,
            pay_date=jdatetime.date(1405, 2, 8),
            tracking_code='DUAL-STATUS',
            status=PaymentRecord.STATUS_APPROVED,
        )

        self.client.login(username='customer1', password='pass1234')
        response = self.client.get(reverse('payment_create'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'وضعیت بازرگانی')
        self.assertContains(response, 'وضعیت مالی')
        self.assertContains(response, 'ثبت بازرگانی')
        self.assertContains(response, 'در انتظار ثبت مالی')

    def test_staff_status_choices_for_commercial_role(self):
        choices = _staff_status_choices_for_role('commercial')
        self.assertEqual(
            choices,
            [
                (PaymentRecord.STATUS_APPROVED, 'ثبت بازرگانی'),
                (PaymentRecord.STATUS_INCOMPLETE, 'ناقص'),
                (PaymentRecord.STATUS_REJECTED, 'رد شده'),
            ]
        )

    def test_staff_status_choices_for_finance_role(self):
        choices = _staff_status_choices_for_role('finance')
        self.assertEqual(
            choices,
            [
                (PaymentRecord.STATUS_FINAL_APPROVED, 'تایید نهایی'),
                (PaymentRecord.STATUS_RETURNED_TO_COMMERCIAL, 'عودت به بازرگانی'),
            ]
        )

    def test_finance_can_see_pending_but_cannot_change_until_commercial_approval(self):
        payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=100000,
            pay_date=jdatetime.date(1405, 2, 8),
            tracking_code='WF-PENDING',
        )

        self.client.login(username='finance1', password='pass1234')
        response = self.client.get(reverse('submit'))
        self.assertContains(response, 'WF-PENDING')

        response = self.client.post(
            reverse('staff_update_status', args=[payment.id]),
            {'status': PaymentRecord.STATUS_FINAL_APPROVED, 'note': '', 'next': reverse('submit')},
        )
        self.assertEqual(response.status_code, 302)
        payment.refresh_from_db()
        self.assertEqual(payment.status, PaymentRecord.STATUS_PENDING)

    def test_excel_export_defaults_to_all_fields_and_accepts_selected_fields(self):
        for index in range(11):
            PaymentRecord.objects.create(
                user=self.customer_user,
                first_name='Ali',
                last_name='Customer',
                organization='Alpha',
                city='Tehran',
                phone='09120000002',
                amount=100000 + index,
                pay_date=jdatetime.date(1405, 2, 8),
                tracking_code=f'EXPORT-{index + 1}',
            )

        self.client.login(username='commercial1', password='pass1234')
        response = self.client.get(reverse('export_data', args=['payments']))
        self.assertEqual(response.status_code, 200)
        self.assertRegex(
            response['Content-Disposition'],
            r'filename="payments_\d{8}_\d{6}\.xlsx"',
        )
        workbook = load_workbook(BytesIO(response.content))
        headers = [cell.value for cell in workbook.active[1]]
        self.assertIn('کد پیگیری', headers)
        self.assertIn('مبلغ', headers)

        response = self.client.get(
            reverse('export_data', args=['payments']),
            {'fields': ['tracking_code', 'amount']},
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        headers = [cell.value for cell in workbook.active[1]]
        self.assertEqual(headers, ['مبلغ', 'کد پیگیری'])

        response = self.client.get(
            reverse('export_data', args=['payments']),
            {'fields': ['tracking_code'], 'scope': 'page', 'per_page': '10', 'page': '1'},
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[0], ('کد پیگیری',))
        self.assertEqual(len(rows), 11)

    def test_daily_payment_plans_support_day_week_month_views_and_export_period(self):
        base_date = jdatetime.date(1405, 2, 10)
        DailyPaymentPlan.objects.create(
            deposit_date=base_date,
            bank_name='Bank',
            account_number='ACC-DAY',
            total_expected_amount=100000,
            created_by=self.commercial_user,
        )
        DailyPaymentPlan.objects.create(
            deposit_date=base_date + jdatetime.timedelta(days=1),
            bank_name='Bank',
            account_number='ACC-MONTH',
            total_expected_amount=200000,
            created_by=self.commercial_user,
        )
        DailyPaymentPlan.objects.create(
            deposit_date=base_date - jdatetime.timedelta(days=20),
            bank_name='Bank',
            account_number='ACC-PAST',
            total_expected_amount=150000,
            created_by=self.commercial_user,
        )
        DailyPaymentPlan.objects.create(
            deposit_date=jdatetime.date(1405, 3, 1),
            bank_name='Bank',
            account_number='ACC-NEXT',
            total_expected_amount=300000,
            created_by=self.commercial_user,
        )

        self.client.login(username='commercial1', password='pass1234')
        response = self.client.get(reverse('daily_payment_plans'), {'mode': 'day', 'date': '1405/02/10'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ACC-DAY')
        self.assertNotContains(response, 'ACC-MONTH')

        response = self.client.get(reverse('daily_payment_plans'), {'mode': 'week', 'date': '1405/02/10'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ACC-DAY')
        self.assertNotContains(response, 'ACC-MONTH')
        self.assertNotContains(response, 'ACC-PAST')

        response = self.client.get(reverse('daily_payment_plans'), {'mode': 'month', 'date': '1405/02/10'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ACC-DAY')
        self.assertContains(response, 'ACC-PAST')
        self.assertNotContains(response, 'ACC-MONTH')
        self.assertNotContains(response, 'ACC-NEXT')

        response = self.client.get(
            reverse('daily_payment_plans'),
            {'mode': 'range', 'date': '1405/02/11', 'start_date': '1405/02/11', 'end_date': '1405/03/01'},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'ACC-MONTH')
        self.assertContains(response, 'ACC-NEXT')
        self.assertNotContains(response, 'ACC-DAY')
        self.assertNotContains(response, 'ACC-PAST')

        response = self.client.get(
            reverse('export_data', args=['daily_plans']),
            {
                'mode': 'range',
                'date': '1405/02/11',
                'start_date': '1405/02/11',
                'end_date': '1405/03/01',
                'fields': ['account_number'],
            },
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        exported_accounts = [row[0] for row in rows[1:]]
        self.assertIn('ACC-MONTH', exported_accounts)
        self.assertIn('ACC-NEXT', exported_accounts)
        self.assertNotIn('ACC-DAY', exported_accounts)

    def test_daily_assignment_payment_notifies_staff_roles_and_report_exports_status(self):
        plan = DailyPaymentPlan.objects.create(
            deposit_date=jdatetime.date(1405, 2, 8),
            bank_name='Bank',
            account_number='PLAN-ACC',
            total_expected_amount=500000,
            created_by=self.commercial_user,
        )
        assignment = DailyPaymentAssignment.objects.create(
            plan=plan,
            customer=self.customer_user,
            expected_amount=500000,
        )
        DailyPaymentAssignment.objects.create(
            plan=plan,
            customer=self.other_customer,
            expected_amount=300000,
        )

        payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=500000,
            pay_date=jdatetime.date(1405, 2, 8),
            tracking_code='PLAN-PAY',
            daily_assignment=assignment,
        )
        _notify_payment_created = __import__('payments.views', fromlist=['_notify_payment_created'])._notify_payment_created
        _notify_payment_created(payment, self.customer_user)
        self.assertTrue(UserNotification.objects.filter(user=self.commercial_user, title='واریز برنامه‌ریزی‌شده ثبت شد').exists())
        self.assertTrue(UserNotification.objects.filter(user=self.finance_user, title='واریز برنامه‌ریزی‌شده ثبت شد').exists())
        self.assertTrue(UserNotification.objects.filter(user=self.sales_user, title='واریز برنامه‌ریزی‌شده ثبت شد').exists())

        self.client.logout()
        self.client.login(username='finance1', password='pass1234')
        response = self.client.get(reverse('daily_payment_plan_detail', args=[plan.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'تکمیل شده')
        self.assertContains(response, 'بدون واریز')

        response = self.client.get(
            reverse('export_data', args=['daily_assignments']),
            {'plan_id': str(plan.id), 'fields': ['customer_name', 'plan_status', 'paid_percent']},
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        headers = rows[0]
        self.assertIn('وضعیت برنامه', headers)
        self.assertIn('درصد تحقق', headers)
        values = [cell for row in rows[1:] for cell in row]
        self.assertIn('تکمیل شده', values)
        self.assertIn('بدون واریز', values)

    def test_data_entry_user_can_complete_payment_details_and_changes_are_logged(self):
        payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=0,
            tracking_code='',
        )

        self.client.login(username='dataentry1', password='pass1234')
        response = self.client.get(reverse('submit'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse('staff_edit_payment_details', args=[payment.id]))

        response = self.client.post(
            reverse('staff_edit_payment_details', args=[payment.id]),
            {
                'payer_account_number': '111222333',
                'payer_full_name': 'Ali Customer',
                'payer_bank_name': 'Melli',
                'beneficiary_bank_name': 'Saderat',
                'beneficiary_account_number': '444555666',
                'beneficiary_account_owner': 'Company',
                'amount': '250000',
                'tracking_code': 'TRACK-250',
                'pay_date': '1405/02/08',
                'next': reverse('submit'),
            },
        )
        self.assertEqual(response.status_code, 302)
        payment.refresh_from_db()
        self.assertEqual(payment.status, PaymentRecord.STATUS_PENDING)
        self.assertEqual(payment.tracking_code, 'TRACK-250')
        self.assertEqual(payment.amount, 250000)

        log = PaymentActivityLog.objects.filter(
            payment=payment,
            actor=self.data_entry_user,
            action=PaymentActivityLog.ACTION_EDITED,
        ).latest('created_at')
        self.assertEqual(log.from_status, PaymentRecord.STATUS_PENDING)
        self.assertEqual(log.to_status, PaymentRecord.STATUS_PENDING)
        self.assertIn('کد پیگیری', log.note)
        self.assertIn('TRACK-250', log.note)

    def test_data_entry_user_cannot_change_status_and_only_sees_detail_edit_action(self):
        payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=100000,
            pay_date=jdatetime.date(1405, 2, 8),
            tracking_code='DATAENTRY-NO-STATUS',
            status=PaymentRecord.STATUS_PENDING,
        )

        self.client.login(username='dataentry1', password='pass1234')
        response = self.client.get(reverse('submit'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse('staff_edit_payment_details', args=[payment.id]))
        self.assertNotContains(response, reverse('staff_update_status', args=[payment.id]))
        self.assertNotContains(response, 'تغییر وضعیت و توضیح')

        response = self.client.post(
            reverse('staff_update_status', args=[payment.id]),
            {'status': PaymentRecord.STATUS_APPROVED, 'note': '', 'next': reverse('submit')},
        )
        self.assertEqual(response.status_code, 302)
        payment.refresh_from_db()
        self.assertEqual(payment.status, PaymentRecord.STATUS_PENDING)

    def test_user_without_payment_detail_permission_cannot_edit_customer_payment_details(self):
        payment = PaymentRecord.objects.create(
            user=self.customer_user,
            first_name='Ali',
            last_name='Customer',
            organization='Alpha',
            city='Tehran',
            phone='09120000002',
            amount=100000,
            tracking_code='NO-EDIT',
        )

        self.client.login(username='commercial1', password='pass1234')
        response = self.client.get(reverse('staff_edit_payment_details', args=[payment.id]))
        self.assertEqual(response.status_code, 403)

    def test_uploaded_files_are_not_served_from_direct_media_url(self):
        invoice = InvoiceRecord.objects.create(
            customer=self.customer_user,
            uploaded_by=self.commercial_user,
            amount=500000,
            invoice_date=jdatetime.date(1405, 2, 8),
            invoice_number='INV-MEDIA',
            reference_number='REF-MEDIA',
            attachment=SimpleUploadedFile('media-direct.pdf', b'%PDF-1.4 media', content_type='application/pdf'),
        )

        self.client.login(username='customer1', password='pass1234')
        response = self.client.get(invoice.attachment.url)
        self.assertEqual(response.status_code, 404)

    def test_price_list_upload_history_and_customer_latest_file_scope(self):
        older = PriceList.objects.create(
            customer=self.customer_user,
            uploaded_by=self.commercial_user,
            title='old',
            file=SimpleUploadedFile('old.pdf', b'%PDF-1.4 old', content_type='application/pdf'),
        )
        latest_batch = uuid.uuid4()
        latest = PriceList.objects.create(
            customer=self.customer_user,
            uploaded_by=self.sales_user,
            title='latest',
            file=SimpleUploadedFile('latest.pdf', b'%PDF-1.4 latest', content_type='application/pdf'),
            batch_id=latest_batch,
        )
        latest_second = PriceList.objects.create(
            customer=self.customer_user,
            uploaded_by=self.sales_user,
            title='latest second',
            file=SimpleUploadedFile('latest-2.pdf', b'%PDF-1.4 latest 2', content_type='application/pdf'),
            batch_id=latest_batch,
        )
        other = PriceList.objects.create(
            customer=self.other_customer,
            uploaded_by=self.finance_user,
            title='other',
            file=SimpleUploadedFile('other.pdf', b'%PDF-1.4 other', content_type='application/pdf'),
        )

        self.client.login(username='customer1', password='pass1234')
        response = self.client.get(reverse('price_lists'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, latest.title)
        self.assertContains(response, latest_second.title)
        self.assertNotContains(response, older.title)
        self.assertNotContains(response, other.title)
        self.assertContains(response, 'مشاهده')
        self.assertContains(response, reverse('price_list_file', args=[latest.id]))
        self.assertContains(response, reverse('price_list_file', args=[latest_second.id]))
        self.assertContains(response, '?download=1')
        self.assertEqual(self.client.get(reverse('price_list_file', args=[latest.id])).status_code, 200)
        self.assertEqual(self.client.get(reverse('price_list_file', args=[latest_second.id])).status_code, 200)
        self.assertEqual(self.client.get(reverse('price_list_file', args=[older.id])).status_code, 403)
        self.assertEqual(self.client.get(reverse('price_list_file', args=[other.id])).status_code, 403)

        self.client.logout()
        self.client.login(username='finance1', password='pass1234')
        response = self.client.get(reverse('price_lists'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, latest.title)
        self.assertContains(response, older.title)
        self.assertContains(response, other.title)
        self.assertContains(response, 'preview-toggle')
        self.assertEqual(self.client.get(reverse('price_list_file', args=[other.id])).status_code, 200)

        response = self.client.get(reverse('price_lists'), {'city': 'اصفهان'})
        self.assertContains(response, other.title)
        self.assertNotContains(response, latest.title)

    def test_commercial_sales_finance_can_upload_price_list_but_customer_cannot(self):
        uploads = [
            SimpleUploadedFile('price.pdf', b'%PDF-1.4 price', content_type='application/pdf'),
            SimpleUploadedFile('price-2.pdf', b'%PDF-1.4 price 2', content_type='application/pdf'),
        ]
        CustomerSalesAssignment.objects.create(
            customer=self.customer_user,
            sales_user=self.sales_user,
            assigned_by=self.commercial_user,
        )
        CustomerSalesAssignment.objects.create(
            customer=self.other_customer,
            sales_user=self.sales_user,
            assigned_by=self.commercial_user,
        )

        self.client.login(username='sales1', password='pass1234')
        response = self.client.post(
            reverse('price_lists'),
            {
                'customers': [str(self.customer_profile.id), str(self.other_customer.profile.id)],
                'title': 'sales price',
                'files': uploads,
                'note': 'internal',
            },
        )
        self.assertEqual(response.status_code, 302)
        self.assertEqual(PriceList.objects.filter(customer=self.customer_user, title='sales price').count(), 2)
        self.assertEqual(PriceList.objects.filter(customer=self.other_customer, title='sales price').count(), 2)

        self.client.logout()
        self.client.login(username='customer1', password='pass1234')
        response = self.client.post(
            reverse('price_lists'),
            {
                'customers': [str(self.customer_profile.id)],
                'title': 'customer price',
                'files': SimpleUploadedFile('customer.pdf', b'%PDF-1.4 customer', content_type='application/pdf'),
            },
        )
        self.assertEqual(response.status_code, 403)

    def test_customer_file_views_mark_seen_and_staff_can_delete_documents(self):
        invoice = InvoiceRecord.objects.create(
            customer=self.customer_user,
            uploaded_by=self.commercial_user,
            amount=700000,
            invoice_date=jdatetime.date(1405, 2, 8),
            invoice_number='DEL-INV',
            attachment=SimpleUploadedFile('delete-invoice.pdf', b'%PDF-1.4 invoice', content_type='application/pdf'),
        )
        price_list = PriceList.objects.create(
            customer=self.customer_user,
            uploaded_by=self.sales_user,
            title='DEL-PRICE',
            file=SimpleUploadedFile('delete-price.pdf', b'%PDF-1.4 price', content_type='application/pdf'),
        )
        proforma = ProformaInvoice.objects.create(
            customer=self.customer_user,
            issued_by=self.commercial_user,
            title='DEL-PF',
            valid_until=jdatetime.date(1405, 12, 29),
            file=SimpleUploadedFile('delete-proforma.pdf', b'%PDF-1.4 proforma', content_type='application/pdf'),
        )

        self.client.login(username='customer1', password='pass1234')
        self.assertEqual(self.client.get(reverse('price_list_file', args=[price_list.id])).status_code, 200)
        self.assertEqual(self.client.get(reverse('proforma_file', args=[proforma.id])).status_code, 200)
        price_list.refresh_from_db()
        proforma.refresh_from_db()
        self.assertIsNotNone(price_list.customer_seen_at)
        self.assertIsNotNone(proforma.customer_seen_at)

        response = self.client.post(reverse('price_list_delete', args=[price_list.id]))
        self.assertEqual(response.status_code, 403)

        self.client.logout()
        self.client.login(username='commercial1', password='pass1234')
        self.assertEqual(self.client.post(reverse('invoice_delete', args=[invoice.id])).status_code, 302)
        self.assertEqual(self.client.post(reverse('price_list_delete', args=[price_list.id])).status_code, 302)
        self.assertEqual(self.client.post(reverse('proforma_delete', args=[proforma.id])).status_code, 302)
        self.assertFalse(InvoiceRecord.objects.filter(id=invoice.id).exists())
        self.assertFalse(PriceList.objects.filter(id=price_list.id).exists())
        self.assertFalse(ProformaInvoice.objects.filter(id=proforma.id).exists())

    def test_commercial_can_issue_proforma_customer_view_and_approve_logs_notification(self):
        self.client.login(username='commercial1', password='pass1234')
        response = self.client.post(
            reverse('proformas'),
            {
                'customers': [str(self.customer_profile.id)],
                'title': 'PF-1',
                'valid_until': '1405/12/29',
                'files': [SimpleUploadedFile('pf.pdf', b'%PDF-1.4 proforma', content_type='application/pdf')],
                'note': 'internal',
            },
        )
        self.assertEqual(response.status_code, 302)
        proforma = ProformaInvoice.objects.get(title='PF-1')
        self.assertEqual(proforma.issued_by, self.commercial_user)

        self.client.logout()
        self.client.login(username='customer1', password='pass1234')
        response = self.client.get(reverse('proforma_detail', args=[proforma.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'مشاهده فایل')
        self.assertContains(response, reverse('proforma_file', args=[proforma.id]))
        response = self.client.get(reverse('proforma_file', args=[proforma.id]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            ProformaInvoiceLog.objects.filter(proforma=proforma, actor=self.customer_user, action=ProformaInvoiceLog.ACTION_VIEWED).count(),
            1,
        )
        self.assertEqual(
            ProformaInvoiceLog.objects.filter(proforma=proforma, actor=self.customer_user, action=ProformaInvoiceLog.ACTION_FILE_VIEWED).count(),
            1,
        )

        response = self.client.post(reverse('proforma_detail', args=[proforma.id]), {'action': 'approve'})
        self.assertEqual(response.status_code, 302)
        proforma.refresh_from_db()
        self.assertEqual(proforma.status, ProformaInvoice.STATUS_APPROVED)
        self.assertIsNotNone(proforma.approved_at)
        self.assertTrue(
            ProformaInvoiceLog.objects.filter(proforma=proforma, actor=self.customer_user, action=ProformaInvoiceLog.ACTION_APPROVED).exists()
        )
        self.assertTrue(
            UserNotification.objects.filter(user=self.commercial_user, title='تایید پیش فاکتور').exists()
        )

    def test_commercial_can_issue_multiple_proforma_files_for_selected_customers(self):
        self.client.login(username='commercial1', password='pass1234')
        response = self.client.post(
            reverse('proformas'),
            {
                'customers': [str(self.customer_profile.id), str(self.other_customer.profile.id)],
                'title': 'PF-BATCH',
                'valid_until': '1405/12/29',
                'files': [
                    SimpleUploadedFile('pf-1.pdf', b'%PDF-1.4 proforma 1', content_type='application/pdf'),
                    SimpleUploadedFile('pf-2.pdf', b'%PDF-1.4 proforma 2', content_type='application/pdf'),
                ],
                'note': 'batch',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(ProformaInvoice.objects.filter(title='PF-BATCH').count(), 4)
        self.assertEqual(ProformaInvoice.objects.filter(customer=self.customer_user, title='PF-BATCH').count(), 2)
        self.assertEqual(ProformaInvoice.objects.filter(customer=self.other_customer, title='PF-BATCH').count(), 2)
        self.assertTrue(UserNotification.objects.filter(user=self.customer_user, title='پیش فاکتور جدید').exists())
        self.assertTrue(UserNotification.objects.filter(user=self.other_customer, title='پیش فاکتور جدید').exists())

    def test_customer_cannot_access_other_or_expired_proforma_approval(self):
        expired = ProformaInvoice.objects.create(
            customer=self.customer_user,
            issued_by=self.commercial_user,
            title='expired',
            valid_until=jdatetime.date(1404, 1, 1),
            file=SimpleUploadedFile('expired.pdf', b'%PDF-1.4 expired', content_type='application/pdf'),
        )
        other = ProformaInvoice.objects.create(
            customer=self.other_customer,
            issued_by=self.commercial_user,
            title='other',
            valid_until=jdatetime.date(1405, 12, 29),
            file=SimpleUploadedFile('other-pf.pdf', b'%PDF-1.4 other', content_type='application/pdf'),
        )

        self.client.login(username='customer1', password='pass1234')
        self.assertEqual(self.client.get(reverse('proforma_file', args=[other.id])).status_code, 403)
        response = self.client.post(reverse('proforma_detail', args=[expired.id]), {'action': 'approve'})
        self.assertEqual(response.status_code, 403)
        expired.refresh_from_db()
        self.assertEqual(expired.status, ProformaInvoice.STATUS_PENDING)

    def test_customer_profile_changes_wait_for_staff_approval(self):
        self.client.login(username='customer1', password='pass1234')
        response = self.client.post(
            reverse('profile_edit'),
            {
                'email': 'customer1@example.com',
                'phone': '02122222222',
                'second_mobile': '09124444444',
                'address': 'آدرس اصلی مشتری',
                'second_address': 'آدرس دوم مشتری',
                'organization': 'مجموعه جدید',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.customer_user.refresh_from_db()
        self.customer_profile.refresh_from_db()
        self.assertEqual(self.customer_user.email, '')
        self.assertEqual(self.customer_profile.phone, '09120000002')
        self.assertEqual(self.customer_profile.second_mobile, '')
        self.assertEqual(self.customer_profile.organization, 'شرکت آلفا')

        change = ProfileChangeRequest.objects.get(user=self.customer_user)
        self.assertEqual(change.status, ProfileChangeRequest.STATUS_PENDING)
        self.assertEqual(change.changes['email']['new'], 'customer1@example.com')
        self.assertEqual(change.changes['organization']['new'], 'مجموعه جدید')

        log = SystemActivityLog.objects.get(action=SystemActivityLog.ACTION_PROFILE_UPDATED)
        self.assertEqual(log.actor, self.customer_user)
        self.assertEqual(log.target_user, self.customer_user)
        self.assertIn('شماره همراه دوم', log.description)
        self.assertIn('آدرس دوم', log.description)

        self.client.logout()
        self.client.login(username='finance1', password='pass1234')
        response = self.client.post(
            reverse('profile_change_request_review', args=[change.id]),
            {'action': 'approve'},
        )
        self.assertEqual(response.status_code, 302)
        change.refresh_from_db()
        self.customer_user.refresh_from_db()
        self.customer_profile.refresh_from_db()
        self.assertEqual(change.status, ProfileChangeRequest.STATUS_APPROVED)
        self.assertEqual(self.customer_user.email, 'customer1@example.com')
        self.assertEqual(self.customer_profile.phone, '02122222222')
        self.assertEqual(self.customer_profile.second_mobile, '09124444444')
        self.assertEqual(self.customer_profile.organization, 'مجموعه جدید')

    def test_customer_profile_optional_fields_can_be_empty(self):
        self.customer_user.email = 'old@example.com'
        self.customer_user.save(update_fields=['email'])
        self.customer_profile.phone = '02122222222'
        self.customer_profile.mobile = '09123333333'
        self.customer_profile.second_mobile = '09124444444'
        self.customer_profile.address = 'آدرس اصلی'
        self.customer_profile.second_address = 'آدرس دوم'
        self.customer_profile.save()

        self.client.login(username='customer1', password='pass1234')
        response = self.client.post(
            reverse('profile_edit'),
            {
                'email': '',
                'phone': '',
                'second_mobile': '',
                'organization': '',
                'address': '',
                'second_address': '',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.customer_user.refresh_from_db()
        self.customer_profile.refresh_from_db()
        self.assertEqual(self.customer_user.email, 'old@example.com')
        self.assertEqual(self.customer_profile.phone, '02122222222')
        self.assertEqual(self.customer_profile.mobile, '09123333333')
        change = ProfileChangeRequest.objects.get(user=self.customer_user)
        self.assertEqual(change.changes['email']['new'], '')
        self.assertEqual(change.changes['second_mobile']['new'], '')


class UserManagementTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username='admin1',
            password='pass1234',
            email='admin@example.com',
        )
        self.admin_user.profile.phone = '02100000000'
        self.admin_user.profile.save()

    def test_superuser_can_create_user_from_simple_management_page(self):
        self.client.login(username='admin1', password='pass1234')
        response = self.client.post(
            reverse('users_manage'),
            {
                'username': 'customer_new',
                'first_name': 'مهدی',
                'last_name': 'محمدی',
                'email': 'customer-new@example.com',
                'phone': '02111111111',
                'mobile': '09121111111',
                'province': 'تهران',
                'city': 'تهران',
                'address': 'خیابان نمونه',
                'organization': 'شرکت نمونه',
                'password': 'Ab123',
                'role': 'customer',
                'active_from': '1405/02/08',
                'valid_until': '1405/12/29',
                'force_password_change': 'on',
                'is_active': 'on',
            },
        )
        self.assertEqual(response.status_code, 302)
        created_user = User.objects.get(username='09121111111')
        self.assertEqual(created_user.first_name, 'مهدی')
        self.assertEqual(created_user.email, 'customer-new@example.com')
        self.assertEqual(created_user.profile.mobile, '09121111111')
        self.assertEqual(created_user.profile.role, 'customer')

    def test_user_edit_hides_and_ignores_password_field(self):
        target = User.objects.create_user(
            username='09122222222',
            password='Old123',
            email='customer-edit@example.com',
            first_name='Old',
            last_name='User',
        )
        target.profile.phone = '02122222222'
        target.profile.mobile = '09122222222'
        target.profile.province = 'Tehran'
        target.profile.city = 'Tehran'
        target.profile.organization = 'Old Org'
        target.profile.role = 'customer'
        target.profile.save()
        old_password_hash = target.password

        self.client.login(username='admin1', password='pass1234')
        response = self.client.get(reverse('user_edit', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="password"')
        self.assertNotContains(response, 'id="use-suggested-password"')

        response = self.client.post(
            reverse('user_edit', args=[target.id]),
            {
                'first_name': 'New',
                'last_name': 'User',
                'email': 'customer-edit-new@example.com',
                'phone': '02122222222',
                'mobile': '09122222222',
                'province': 'Tehran',
                'city': 'Tehran',
                'address': 'Address',
                'organization': 'New Org',
                'password': 'New123',
                'role': 'customer',
                'active_from': '1405/02/08',
                'valid_until': '1405/12/29',
                'is_active': 'on',
            },
        )

        self.assertEqual(response.status_code, 302)
        target.refresh_from_db()
        self.assertEqual(target.password, old_password_hash)
        self.assertTrue(target.check_password('Old123'))
        self.assertEqual(target.first_name, 'New')

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_password_reset_emails_password_logs_without_storing_password_and_forces_change(self):
        target = User.objects.create_user(
            username='customer_reset',
            password='Old123',
            email='customer-reset@example.com',
        )
        target.profile.role = 'customer'
        target.profile.force_password_change = False
        target.profile.save()

        self.client.login(username='admin1', password='pass1234')
        response = self.client.post(reverse('reset_user_password', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        temp_password = payload['temp_password']
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn(temp_password, mail.outbox[0].body)
        self.assertEqual(mail.outbox[0].to, ['customer-reset@example.com'])

        target.refresh_from_db()
        self.assertTrue(target.check_password(temp_password))
        self.assertTrue(target.profile.force_password_change)

        log = SystemActivityLog.objects.get(action=SystemActivityLog.ACTION_PASSWORD_RESET)
        self.assertEqual(log.actor, self.admin_user)
        self.assertEqual(log.target_user, target)
        self.assertNotIn(temp_password, log.description)

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend')
    def test_password_reset_without_target_email_still_changes_password_and_logs_it(self):
        target = User.objects.create_user(
            username='customer_no_email',
            password='Old123',
            email='',
        )
        target.profile.role = 'customer'
        target.profile.force_password_change = False
        target.profile.save()

        self.client.login(username='admin1', password='pass1234')
        response = self.client.post(reverse('reset_user_password', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        temp_password = payload['temp_password']
        target.refresh_from_db()
        self.assertTrue(target.check_password(temp_password))
        self.assertTrue(target.profile.force_password_change)
        self.assertEqual(len(mail.outbox), 0)
        log = SystemActivityLog.objects.get(target_user=target, action=SystemActivityLog.ACTION_PASSWORD_RESET)
        self.assertNotIn(temp_password, log.description)
        self.assertIn('ایمیل ثبت نشده است', log.description)

    @override_settings(EMAIL_BACKEND='django.core.mail.backends.smtp.EmailBackend', EMAIL_HOST='invalid.localhost')
    def test_password_reset_continues_when_email_delivery_fails(self):
        target = User.objects.create_user(
            username='customer_email_fails',
            password='Old123',
            email='customer-fail@example.com',
        )
        target.profile.role = 'customer'
        target.profile.force_password_change = False
        target.profile.save()

        self.client.login(username='admin1', password='pass1234')
        response = self.client.post(reverse('reset_user_password', args=[target.id]))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        temp_password = payload['temp_password']
        self.assertIn('ارسال ایمیل انجام نشد', payload['message'])

        target.refresh_from_db()
        self.assertTrue(target.check_password(temp_password))
        self.assertTrue(target.profile.force_password_change)

        log = SystemActivityLog.objects.get(target_user=target, action=SystemActivityLog.ACTION_PASSWORD_RESET)
        self.assertNotIn(temp_password, log.description)
        self.assertIn('ارسال ایمیل انجام نشد', log.description)


class CustomerOrderTests(TestCase):
    def setUp(self):
        self.customer = User.objects.create_user(
            username='customer-order',
            password='pass1234',
            first_name='Customer',
            last_name='One',
        )
        self.customer.profile.role = 'customer'
        self.customer.profile.organization = 'Customer Org'
        self.customer.profile.city = 'Tehran'
        self.customer.profile.province = 'Tehran'
        self.customer.profile.force_password_change = False
        self.customer.profile.save()

        self.sales = User.objects.create_user(
            username='sales-order',
            password='pass1234',
            first_name='Sales',
            last_name='One',
        )
        self.sales.profile.role = 'sales'
        self.sales.profile.force_password_change = False
        self.sales.profile.save()

    def test_customer_can_create_order_with_items_and_sales_expert(self):
        self.client.login(username='customer-order', password='pass1234')
        response = self.client.get(reverse('orders'))
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            reverse('orders'),
            {
                'title': 'Spring order',
                'requested_sales_expert': str(self.sales.id),
                'customer_note': 'Please prepare proforma.',
                'items-TOTAL_FORMS': '3',
                'items-INITIAL_FORMS': '0',
                'items-MIN_NUM_FORMS': '1',
                'items-MAX_NUM_FORMS': '1000',
                'items-0-product_name': 'Product A',
                'items-0-quantity': '2',
                'items-0-unit': 'carton',
                'items-0-note': 'Blue',
                'items-1-product_name': '',
                'items-1-quantity': '',
                'items-1-unit': '',
                'items-1-note': '',
                'items-2-product_name': '',
                'items-2-quantity': '',
                'items-2-unit': '',
                'items-2-note': '',
            },
        )

        self.assertEqual(response.status_code, 302)
        order = CustomerOrder.objects.get(customer=self.customer)
        self.assertEqual(order.sales_expert, self.sales)
        self.assertEqual(order.items.count(), 1)
        self.assertEqual(order.items.first().product_name, 'Product A')
        self.assertEqual(CustomerSalesAssignment.objects.get(customer=self.customer).sales_user, self.sales)
        response = self.client.get(reverse('order_detail', args=[order.id]))
        self.assertEqual(response.status_code, 200)

    def test_sales_can_issue_order_proforma_for_single_order_customer(self):
        order = CustomerOrder.objects.create(
            customer=self.customer,
            sales_expert=self.sales,
            title='Order for proforma',
        )
        order.items.create(product_name='Product A', quantity='1', unit='piece')

        self.client.login(username='sales-order', password='pass1234')
        response = self.client.post(
            reverse('order_detail', args=[order.id]),
            {
                'action': 'issue_proforma',
                'title': 'Order proforma',
                'valid_until': '1405/12/29',
                'note': 'Internal note',
                'files': SimpleUploadedFile('proforma.pdf', b'%PDF-1.4 sample', content_type='application/pdf'),
            },
        )

        self.assertEqual(response.status_code, 302)
        order.refresh_from_db()
        self.assertEqual(order.status, CustomerOrder.STATUS_PROFORMA_SENT)
        proforma = ProformaInvoice.objects.get(order=order)
        self.assertEqual(proforma.customer, self.customer)
        self.assertEqual(proforma.issued_by, self.sales)
